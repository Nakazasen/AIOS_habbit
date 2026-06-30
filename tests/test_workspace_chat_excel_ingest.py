from io import BytesIO
import socket
import zipfile
from pathlib import Path

import openpyxl
import pytest

from aios_habit.workspace_chat_excel import (
    EMPTY_WORKBOOK_MESSAGE,
    GENERIC_READ_ERROR_MESSAGE,
    MAX_UPLOAD_BYTES,
    MAX_ZIP_UNCOMPRESSED_BYTES,
    TRUNCATED_MESSAGE,
    XLS_UNSUPPORTED_MESSAGE,
    extract_xlsx_text,
)
from aios_habit.workspace_chat_models import (
    WORKSPACE_CHAT_SOURCE_PREVIEW_LIMIT,
    WORKSPACE_CHAT_SOURCE_TEXT_LIMIT_BYTES,
)


def workbook_bytes(workbook):
    bio = BytesIO()
    workbook.save(bio)
    return bio.getvalue()


def test_extract_one_sheet_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kế hoạch"
    ws["A1"] = "Mã lệnh"
    ws["B1"] = "Sản phẩm"
    ws["A2"] = "MO-001"
    ws["B2"] = 120

    result = extract_xlsx_text(workbook_bytes(wb), "Ke_hoach_san_xuat.xlsx")

    assert result.ok is True
    assert result.filename == "Ke_hoach_san_xuat.xlsx"
    assert result.sheet_names == ("Kế hoạch",)
    assert "File Excel: Ke_hoach_san_xuat.xlsx" in result.text
    assert "Trang tính: Kế hoạch" in result.text
    assert "A1=Mã lệnh | B1=Sản phẩm" in result.text
    assert "A2=MO-001 | B2=120" in result.text


def test_extract_multi_sheet_preserves_order():
    wb = openpyxl.Workbook()
    wb.active.title = "Một"
    wb.active["A1"] = "first"
    ws2 = wb.create_sheet("Hai")
    ws2["A1"] = "second"

    result = extract_xlsx_text(workbook_bytes(wb), "multi.xlsx")

    assert result.ok is True
    assert result.sheet_names == ("Một", "Hai")
    assert result.text.index("Trang tính: Một") < result.text.index("Trang tính: Hai")


def test_empty_cells_skipped_but_zero_and_false_kept():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = None
    ws["B1"] = "   "
    ws["C1"] = 0
    ws["D1"] = False

    result = extract_xlsx_text(workbook_bytes(wb), "data.xlsx")

    assert result.ok is True
    assert "A1=" not in result.text
    assert "B1=" not in result.text
    assert "C1=0" in result.text
    assert "D1=False" in result.text


def test_unicode_values_preserved():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Tiếng Việt"
    ws["B1"] = "日本語"

    result = extract_xlsx_text(workbook_bytes(wb), "unicode.XLSX")

    assert result.ok is True
    assert "Tiếng Việt" in result.text
    assert "日本語" in result.text


def test_formula_text_preserved_not_recalculated():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["B1"] = 2
    ws["C1"] = "=SUM(A1:B1)"

    result = extract_xlsx_text(workbook_bytes(wb), "formula.xlsx")

    assert result.ok is True
    assert "C1==SUM(A1:B1)" in result.text
    assert "recalculate" not in result.text.lower()
    assert "tính lại" not in result.text.lower()


def test_merged_cells_anchor_once_no_crash():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("A1:B2")
    ws["A1"] = "Merged"

    result = extract_xlsx_text(workbook_bytes(wb), "merged.xlsx")

    assert result.ok is True
    assert result.text.count("A1=Merged") == 1
    assert "B1=Merged" not in result.text


def test_empty_workbook_friendly_failure():
    wb = openpyxl.Workbook()

    result = extract_xlsx_text(workbook_bytes(wb), "empty.xlsx")

    assert result.ok is False
    assert result.owner_message == EMPTY_WORKBOOK_MESSAGE


def test_corrupt_xlsx_friendly_no_raw_exception():
    result = extract_xlsx_text(b"not-a-zip", "bad.xlsx")

    assert result.ok is False
    assert result.owner_message == GENERIC_READ_ERROR_MESSAGE
    assert "BadZipFile" not in result.owner_message
    assert "Traceback" not in result.owner_message


def test_xls_unsupported_does_not_call_openpyxl(monkeypatch):
    def fail_load(*args, **kwargs):
        raise AssertionError("openpyxl should not be called for .xls")

    monkeypatch.setattr(openpyxl, "load_workbook", fail_load)
    result = extract_xlsx_text(b"legacy", "old.xls")

    assert result.ok is False
    assert result.owner_message == XLS_UNSUPPORTED_MESSAGE


def test_uppercase_xlsx_allowed():
    wb = openpyxl.Workbook()
    wb.active["A1"] = "ok"

    result = extract_xlsx_text(workbook_bytes(wb), "UPPER.XLSX")

    assert result.ok is True


def test_other_extension_rejected_friendly():
    result = extract_xlsx_text(b"abc", "notes.csv")

    assert result.ok is False
    assert result.owner_message == GENERIC_READ_ERROR_MESSAGE


def test_upload_too_large_rejected():
    result = extract_xlsx_text(b"0" * (MAX_UPLOAD_BYTES + 1), "big.xlsx")

    assert result.ok is False
    assert result.owner_message == GENERIC_READ_ERROR_MESSAGE
    assert result.truncated is True


def test_zip_uncompressed_too_large_rejected():
    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("xl/worksheets/sheet1.xml", "0" * (MAX_ZIP_UNCOMPRESSED_BYTES + 1))

    result = extract_xlsx_text(bio.getvalue(), "zipbomb.xlsx")

    assert result.ok is False
    assert result.owner_message == GENERIC_READ_ERROR_MESSAGE
    assert result.truncated is True


def test_sheet_limit_sets_truncated():
    wb = openpyxl.Workbook()
    wb.active.title = "S0"
    wb.active["A1"] = "v0"
    for i in range(1, 14):
        ws = wb.create_sheet(f"S{i}")
        ws["A1"] = f"v{i}"

    result = extract_xlsx_text(workbook_bytes(wb), "sheets.xlsx")

    assert result.ok is True
    assert result.truncated is True
    assert len(result.sheet_names) == 12
    assert TRUNCATED_MESSAGE in result.owner_message


def test_row_cell_preview_and_text_limits():
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in range(1, 1050):
        ws.cell(row=row, column=1).value = "Dữ liệu rất dài " * 20

    result = extract_xlsx_text(workbook_bytes(wb), "large.xlsx")

    assert result.ok is True
    assert result.truncated is True
    assert len(result.preview) <= WORKSPACE_CHAT_SOURCE_PREVIEW_LIMIT
    assert len(result.text.encode("utf-8")) <= WORKSPACE_CHAT_SOURCE_TEXT_LIMIT_BYTES
    assert TRUNCATED_MESSAGE in result.owner_message


def test_no_filesystem_write_network_or_ai(monkeypatch, tmp_path):
    writes = []
    original_write_bytes = Path.write_bytes
    original_write_text = Path.write_text

    def track_write_bytes(self, data):
        writes.append(self)
        return original_write_bytes(self, data)

    def track_write_text(self, data, *args, **kwargs):
        writes.append(self)
        return original_write_text(self, data, *args, **kwargs)

    monkeypatch.setattr(Path, "write_bytes", track_write_bytes)
    monkeypatch.setattr(Path, "write_text", track_write_text)
    monkeypatch.setattr(socket.socket, "connect", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("network not allowed")))

    wb = openpyxl.Workbook()
    wb.active["A1"] = "local"
    result = extract_xlsx_text(workbook_bytes(wb), "local.xlsx")

    assert result.ok is True
    assert writes == []


def test_cell_count_limit_activates_before_row_limit():
    from aios_habit.workspace_chat_excel import MAX_NON_EMPTY_CELLS, MAX_ROWS_PER_SHEET

    wb = openpyxl.Workbook()
    ws = wb.active
    # 1000 rows x 21 columns = 21,000 non-empty cells: below row cap, above cell cap.
    for row in range(1, MAX_ROWS_PER_SHEET + 1):
        for col in range(1, 22):
            ws.cell(row=row, column=col).value = "x"

    result = extract_xlsx_text(workbook_bytes(wb), "cell-limit.xlsx")

    assert MAX_NON_EMPTY_CELLS == 20_000
    assert result.ok is True
    assert result.truncated is True
    assert len(result.text.encode("utf-8")) <= WORKSPACE_CHAT_SOURCE_TEXT_LIMIT_BYTES
    assert result.owner_message == TRUNCATED_MESSAGE


def test_xlsm_rejected_without_parsing(monkeypatch):
    def fail_load(*args, **kwargs):
        raise AssertionError("openpyxl should not be called for .xlsm")

    monkeypatch.setattr(openpyxl, "load_workbook", fail_load)

    result = extract_xlsx_text(b"dummy", "macro.xlsm")

    assert result.ok is False
    assert result.owner_message == GENERIC_READ_ERROR_MESSAGE
    assert "Traceback" not in result.owner_message
    assert "AssertionError" not in result.owner_message


def test_near_text_cap_then_truncation_notice_still_within_cap():
    wb = openpyxl.Workbook()
    ws = wb.active
    # Large enough to fill the text cap while row 1001 triggers truncation.
    for row in range(1, 1002):
        ws.cell(row=row, column=1).value = "A" * 260

    result = extract_xlsx_text(workbook_bytes(wb), "near-cap.xlsx")

    assert result.ok is True
    assert result.truncated is True
    assert result.owner_message == TRUNCATED_MESSAGE
    assert len(result.text.encode("utf-8")) <= WORKSPACE_CHAT_SOURCE_TEXT_LIMIT_BYTES


def test_workbook_close_called_on_success(monkeypatch):
    closed = []

    class FakeCell:
        value = "ok"
        coordinate = "A1"
        row = 1
        column = 1

    class FakeSheet:
        def iter_rows(self):
            yield [FakeCell()]

    class FakeWorkbook:
        sheetnames = ["Sheet1"]

        def __getitem__(self, name):
            return FakeSheet()

        def close(self):
            closed.append(True)

    monkeypatch.setattr(openpyxl, "load_workbook", lambda *args, **kwargs: FakeWorkbook())
    wb = openpyxl.Workbook()
    wb.active["A1"] = "placeholder"

    result = extract_xlsx_text(workbook_bytes(wb), "close-success.xlsx")

    assert result.ok is True
    assert closed == [True]


def test_workbook_close_called_when_iteration_fails(monkeypatch):
    closed = []

    class BrokenSheet:
        def iter_rows(self):
            raise RuntimeError("broken sheet")

    class FakeWorkbook:
        sheetnames = ["Sheet1"]

        def __getitem__(self, name):
            return BrokenSheet()

        def close(self):
            closed.append(True)

    monkeypatch.setattr(openpyxl, "load_workbook", lambda *args, **kwargs: FakeWorkbook())
    wb = openpyxl.Workbook()
    wb.active["A1"] = "placeholder"

    result = extract_xlsx_text(workbook_bytes(wb), "close-failure.xlsx")

    assert result.ok is False
    assert result.owner_message == GENERIC_READ_ERROR_MESSAGE
    assert closed == [True]
