import os
import tempfile
from types import SimpleNamespace
from unittest.mock import patch
import pytest
import zipfile
import openpyxl
from typing import List

from aios_habit.rag_v2.schema import DocumentElement, ExtractionStatus, ElementType
from aios_habit.rag_v2.adapters import ConversionContext
from aios_habit.rag_v2.converters import (
    TextDocumentConverterAdapter,
    HTMLDocumentConverterAdapter,
    ImageOCRDocumentConverterAdapter,
    PDFDocumentConverterAdapter,
    ExcelDocumentConverterAdapter,
    WordDocumentConverterAdapter,
    PowerPointDocumentConverterAdapter,
)
from aios_habit.rag_v2.registry import ConverterRegistry


def test_adapter_capabilities():
    registry = ConverterRegistry()
    caps = registry.list_capabilities()
    assert len(caps) == 7
    names = {c["adapter_name"] for c in caps}
    assert "TextDocumentConverterAdapter" in names
    assert "HTMLDocumentConverterAdapter" in names
    assert "ImageOCRDocumentConverterAdapter" in names
    assert "PDFDocumentConverterAdapter" in names
    assert "ExcelDocumentConverterAdapter" in names
    assert "WordDocumentConverterAdapter" in names
    assert "PowerPointDocumentConverterAdapter" in names


def test_text_document_converter_success():
    adapter = TextDocumentConverterAdapter()
    
    with tempfile.NamedTemporaryFile(suffix=".txt", mode="w", encoding="utf-8", delete=False) as f:
        f.write("# Heading 1\n\nParagraph 1.\nParagraph 1 line 2.\n\n# Heading 2\n\nParagraph 2.")
        txt_path = f.name

    try:
        ctx = ConversionContext(
            document_id="doc-text",
            privacy_labels=("internal",),
            fail_soft=True
        )
        elements = adapter.convert(txt_path, ctx)
        
        assert len(elements) == 4
        assert elements[0].element_type == ElementType.HEADING
        assert elements[0].text == "# Heading 1"
        assert elements[0].privacy_labels == ("internal",)
        assert elements[0].source_name == os.path.basename(txt_path)
        assert elements[0].file_type == "txt"
        assert elements[0].extractor == "TextDocumentConverterAdapter"
        assert elements[0].extraction_status == ExtractionStatus.SUCCESS

        assert elements[1].element_type == ElementType.TEXT
        assert elements[1].text == "Paragraph 1.\nParagraph 1 line 2."

        assert elements[2].element_type == ElementType.HEADING
        assert elements[2].text == "# Heading 2"

        assert elements[3].element_type == ElementType.TEXT
        assert elements[3].text == "Paragraph 2."

    finally:
        os.remove(txt_path)


def test_html_document_converter_success():
    adapter = HTMLDocumentConverterAdapter()
    
    with tempfile.NamedTemporaryFile(suffix=".html", mode="w", encoding="utf-8", delete=False) as f:
        f.write("<html><head><title>Test</title><style>body {color: red;}</style></head><body><h1>Header</h1><p>Para 1</p></body></html>")
        html_path = f.name

    try:
        ctx = ConversionContext(document_id="doc-html")
        elements = adapter.convert(html_path, ctx)
        
        assert len(elements) >= 2
        texts = [e.text for e in elements]
        assert "Header" in texts
        assert "Para 1" in texts
        assert all(e.extraction_status == ExtractionStatus.SUCCESS for e in elements)
    finally:
        os.remove(html_path)


def test_pdf_document_converter_success():
    adapter = PDFDocumentConverterAdapter()
    
    # Try importing fitz to see if we can create a real PDF, otherwise mock
    try:
        import fitz
        has_fitz = True
    except ImportError:
        has_fitz = False

    if has_fitz:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            pdf_path = f.name
        
        doc = fitz.open()
        page = doc.new_page()
        page.insert_text((50, 50), "Hello PDF World\n\nSecond Paragraph")
        doc.save(pdf_path)
        doc.close()

        try:
            ctx = ConversionContext(document_id="doc-pdf")
            elements = adapter.convert(pdf_path, ctx)
            assert len(elements) >= 1
            assert "Hello PDF World" in elements[0].text
            assert "Second Paragraph" in elements[0].text
            assert elements[0].page == 1
            assert elements[0].extraction_status == ExtractionStatus.SUCCESS
        finally:
            os.remove(pdf_path)
    else:
        # Mocking test
        ctx = ConversionContext(document_id="doc-pdf", fail_soft=True)
        elements = adapter.convert("non_existent_file.pdf", ctx)
        assert len(elements) == 1
        assert elements[0].extraction_status == ExtractionStatus.FAILED


def test_registry_routes_supported_images_to_ocr_adapter():
    registry = ConverterRegistry()
    assert isinstance(
        registry.get_adapter_for_file("scan.TIFF"),
        ImageOCRDocumentConverterAdapter,
    )


def test_image_ocr_adapter_maps_usable_and_failed_quality_results(tmp_path):
    image_path = tmp_path / "scan.png"
    image_path.write_bytes(b"not-decoded-because-extractor-is-mocked")
    adapter = ImageOCRDocumentConverterAdapter()
    context = ConversionContext(document_id="doc-image", language_hints=["vie"])

    usable = [{
        "text": "Nội dung OCR đáng tin cậy",
        "file_type": ".png",
        "extractor_name": "local_ocr",
        "extraction_status": "ocr_partial",
        "warning": "OCR confidence is moderate; review important claims",
        "ocr_confidence": 52.0,
    }]
    with patch(
        "aios_habit.document_extractors.extract_text_chunks_from_file",
        return_value=usable,
    ):
        element = adapter.convert(str(image_path), context)[0]
    assert element.extraction_status == ExtractionStatus.PARTIAL
    assert element.text == "Nội dung OCR đáng tin cậy"
    assert element.confidence == 0.52
    assert element.language_hint == "vie"

    rejected = [{
        "text": "",
        "file_type": ".png",
        "extractor_name": "local_ocr",
        "extraction_status": "failed_with_reason",
        "warning": "OCR confidence 20.00 is below usable threshold 35.00",
        "ocr_confidence": 20.0,
    }]
    with patch(
        "aios_habit.document_extractors.extract_text_chunks_from_file",
        return_value=rejected,
    ):
        element = adapter.convert(str(image_path), context)[0]
    assert element.extraction_status == ExtractionStatus.FAILED
    assert element.text is None
    assert "below usable threshold" in element.extraction_warning


def test_pdf_scan_page_uses_ocr_and_never_emits_empty_success(tmp_path):
    import fitz

    pdf_path = tmp_path / "scan.pdf"
    document = fitz.open()
    document.new_page()
    document.save(pdf_path)
    document.close()
    ocr_result = SimpleNamespace(
        text="Recovered scanned text",
        extraction_status="ocr_success",
        warning="",
        ocr_confidence=91.0,
        ocr_engine="rapidocr",
        extractor_name="local_ocr",
        file_type="pdf",
        ocr_lang="vie+eng",
        ocr_confidence_samples=50,
        ocr_preprocessing="original",
        ocr_attempts=1,
        ocr_quality_reason="passed_quality_gate",
    )

    with patch(
        "aios_habit.document_extractors._ocr_image_object",
        return_value=ocr_result,
    ) as ocr:
        elements = PDFDocumentConverterAdapter().convert(
            str(pdf_path), ConversionContext(document_id="doc-scan")
        )

    assert ocr.call_count == 1
    assert len(elements) == 1
    assert elements[0].text == "Recovered scanned text"
    assert elements[0].extraction_status == ExtractionStatus.SUCCESS
    assert elements[0].extractor.endswith("+local_ocr")



def test_excel_document_converter_success():
    adapter = ExcelDocumentConverterAdapter()
    
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx_path = f.name

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SheetTest"
    ws["A1"] = "ColA"
    ws["B1"] = "ColB"
    ws["A2"] = 10
    ws["B2"] = 20
    wb.save(xlsx_path)
    wb.close()

    try:
        ctx = ConversionContext(document_id="doc-excel")
        elements = adapter.convert(xlsx_path, ctx)
        assert len(elements) == 1
        elem = elements[0]
        assert elem.element_type == ElementType.TABLE
        assert elem.sheet == "SheetTest"
        assert elem.table is not None
        assert elem.table.headers == ["ColA", "ColB"]
        assert any(row == ["ColA", "ColB"] for row in elem.table.rows)
        assert any("10" in str(row) for row in elem.table.rows)
        assert len(elem.table.cells) >= 2
        header_cells = [c for c in elem.table.cells if c.is_header]
        assert len(header_cells) >= 1
        assert header_cells[0].text == "ColA"
    finally:
        os.remove(xlsx_path)


def test_word_document_converter_success():
    adapter = WordDocumentConverterAdapter()
    
    with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as f:
        docx_path = f.name

    with zipfile.ZipFile(docx_path, "w") as z:
        z.writestr(
            "word/document.xml",
            '<?xml version="1.0" encoding="UTF-8"?><w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body><w:p><w:r><w:t>Hello Docx</w:t></w:r></w:p><w:p><w:r><w:t>World Docx</w:t></w:r></w:p></w:body></w:document>'
        )

    try:
        ctx = ConversionContext(document_id="doc-docx")
        elements = adapter.convert(docx_path, ctx)
        assert len(elements) == 2
        assert elements[0].text == "Hello Docx"
        assert elements[1].text == "World Docx"
    finally:
        os.remove(docx_path)


def test_powerpoint_document_converter_success():
    adapter = PowerPointDocumentConverterAdapter()
    
    with tempfile.NamedTemporaryFile(suffix=".pptx", delete=False) as f:
        pptx_path = f.name

    with zipfile.ZipFile(pptx_path, "w") as z:
        z.writestr(
            "ppt/slides/slide1.xml",
            '<?xml version="1.0" encoding="UTF-8"?><p:sld xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"><a:t>Hello Slide 1</a:t></p:sld>'
        )
        z.writestr(
            "ppt/slides/slide2.xml",
            '<?xml version="1.0" encoding="UTF-8"?><p:sld xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"><a:t>Hello Slide 2</a:t></p:sld>'
        )

    try:
        ctx = ConversionContext(document_id="doc-pptx")
        elements = adapter.convert(pptx_path, ctx)
        assert len(elements) == 2
        assert elements[0].text == "Hello Slide 1"
        assert elements[0].slide == 1
        assert elements[1].text == "Hello Slide 2"
        assert elements[1].slide == 2
    finally:
        os.remove(pptx_path)


def test_unsupported_and_missing_files():
    registry = ConverterRegistry()
    ctx_soft = ConversionContext(fail_soft=True)
    ctx_hard = ConversionContext(fail_soft=False)

    # Unsupported format
    elems = registry.convert_document("test.unsupported", ctx_soft)
    assert len(elems) == 1
    assert elems[0].extraction_status == ExtractionStatus.UNSUPPORTED
    assert "No supported adapter found" in elems[0].extraction_warning

    with pytest.raises(ValueError, match="No supported adapter found"):
        registry.convert_document("test.unsupported", ctx_hard)

    # Missing file
    elems = registry.convert_document("missing_file.txt", ctx_soft)
    assert len(elems) == 1
    assert elems[0].extraction_status == ExtractionStatus.FAILED
    assert "File not found" in elems[0].extraction_warning

    with pytest.raises(FileNotFoundError):
        registry.convert_document("missing_file.txt", ctx_hard)


def test_deterministic_failed_element_id():
    registry = ConverterRegistry()
    ctx = ConversionContext(fail_soft=True)
    
    elems1 = registry.convert_document("missing_file.txt", ctx)
    elems2 = registry.convert_document("missing_file.txt", ctx)
    
    assert elems1[0].element_id == elems2[0].element_id
    assert elems1[0].element_id.startswith("failed_")


def test_from_dict_ignores_unknown_future_fields():
    d = {
        "element_id": "test-future-1",
        "document_id": "doc-future",
        "source_path": "test.txt",
        "source_name": "test.txt",
        "file_type": "txt",
        "extractor": "future",
        "extraction_status": "success",
        "element_type": "text",
        "text": "Future Proof",
        "future_field_xyz": "ignored_value",
        "nested_unsupported": {"a": 1}
    }
    
    elem = DocumentElement.from_dict(d)
    assert elem.element_id == "test-future-1"
    assert elem.text == "Future Proof"
    assert not hasattr(elem, "future_field_xyz")
