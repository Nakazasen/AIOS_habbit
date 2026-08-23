from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
import json
from pathlib import Path
from types import SimpleNamespace

import pytest

from aios_habit.workspace_chat_ai_answer import WorkspaceAIContextSource
import aios_habit.workspace_chat_rag_v2_adapter as adapter
import aios_habit.workspace_chat_rag_v2_deployment as deployment_module


def _source(text: str, *, privacy_label: str = "local_only") -> WorkspaceAIContextSource:
    return WorkspaceAIContextSource(
        source_id="source-1",
        source_scope="temporary",
        source_type="pasted_text",
        title="owner-notes.txt",
        privacy_label=privacy_label,
        text=text,
        included_chars=len(text),
        truncated=False,
    )


@pytest.fixture(autouse=True)
def _close_canary_runtimes():
    adapter.close_workspace_chat_rag_v2_runtimes()
    yield
    adapter.close_workspace_chat_rag_v2_runtimes()


def test_canary_config_defaults_off_and_requires_explicit_enablement(monkeypatch):
    monkeypatch.setattr(
        adapter,
        "load_workspace_chat_rag_v2_deployment",
        lambda **_kwargs: None,
    )
    config = adapter.WorkspaceChatRagV2CanaryConfig.from_env({})

    assert config.enabled is False
    assert config.requested_profile == "bge_m3_hybrid"
    assert config.fail_closed_on_error is True
    assert not hasattr(config, "lexical_fallback_enabled")
    assert not hasattr(config, "semantic_progressive")


def test_forget_deleted_source_removes_materialized_text_and_readiness(tmp_path, monkeypatch):
    config = adapter.WorkspaceChatRagV2CanaryConfig(enabled=False, runtime_root=tmp_path / "rag")
    source = _source("Nội dung cần xóa")
    document_id = adapter._document_id(source)
    materialized = config.runtime_root / "materialized_sources" / f"{document_id}.txt"
    materialized.parent.mkdir(parents=True)
    materialized.write_text(source.text, encoding="utf-8")
    key = adapter._preparation_key(config, source)
    adapter._PREPARATION_REGISTRY[key] = adapter._preparation_entry(config, source, "ready")
    deleted_document_ids = []
    monkeypatch.setattr(
        adapter._SUBPROCESS_CLIENT,
        "delete_documents",
        lambda document_ids: deleted_document_ids.extend(document_ids) or 3,
    )

    assert adapter.forget_workspace_chat_sources([source], config=config) == 3
    assert not materialized.exists()
    assert key not in adapter._PREPARATION_REGISTRY
    assert deleted_document_ids == [document_id]


def test_explicit_local_recovery_environment_can_enable_real_deep_search(monkeypatch, tmp_path):
    """The documented local override must carry the complete reranker pin."""
    monkeypatch.setattr(
        adapter,
        "load_workspace_chat_rag_v2_deployment",
        lambda **_kwargs: None,
    )
    config = adapter.WorkspaceChatRagV2CanaryConfig.from_env(
        {
            adapter.CANARY_ENABLED_ENV: "1",
            adapter.BGE_MODEL_PATH_ENV: str(tmp_path / "bge-m3"),
            adapter.BGE_MODEL_REVISION_ENV: "bge-revision",
            adapter.BGE_MODEL_CHECKSUM_ENV: "sha256:" + "a" * 64,
            adapter.ADAPTIVE_ENABLED_ENV: "1",
            adapter.RERANKER_MODEL_PATH_ENV: str(tmp_path / "reranker"),
            adapter.RERANKER_MODEL_REVISION_ENV: "reranker-revision",
            adapter.RERANKER_MODEL_CHECKSUM_ENV: "sha256:" + "b" * 64,
            adapter.DEEP_TIMEOUT_MS_ENV: "120000",
            adapter.DEEP_RERANK_LIMIT_ENV: "9",
        }
    )

    assert config.enabled is True
    assert config.adaptive_enabled is True
    assert config.bge_reranker_model_path == tmp_path / "reranker"
    assert config.bge_reranker_model_revision == "reranker-revision"
    assert config.deep_timeout_ms == 120000
    assert config.deep_rerank_limit == 9


def test_precise_operational_question_prepares_only_matching_sources():
    manual = WorkspaceAIContextSource(
        source_id="manual", source_scope="notebook", source_type="txt",
        title="Matecon manual", privacy_label="cloud_safe",
        text="Manual Matecon controls ACR and CTU.", included_chars=37,
        truncated=False,
    )
    unrelated = _source("Finance process and purchase planning notes.")

    selected = adapter._select_semantic_candidate_sources(
        "Chế độ Manual Matecon ACR/CTU hoạt động như thế nào?",
        (manual, unrelated),
    )

    assert selected == (manual,)


def test_preparation_scope_refuses_broad_question_for_large_source_set():
    sources = tuple(
        WorkspaceAIContextSource(
            source_id=f"source-{index}", source_scope="notebook", source_type="txt",
            title=f"Document {index}", privacy_label="local_only",
            text=f"Operational content {index}.", included_chars=22, truncated=False,
        )
        for index in range(4)
    )

    scope = adapter.select_workspace_chat_preparation_scope("Giải thích giúp tôi", sources)

    assert scope.bounded is False
    assert scope.sources == ()
    assert scope.reason == "question_too_broad"


def test_preparation_scope_keeps_precise_question_bounded():
    manual = WorkspaceAIContextSource(
        source_id="manual", source_scope="notebook", source_type="txt",
        title="Matecon manual", privacy_label="local_only",
        text="Manual Matecon controls ACR and CTU.", included_chars=37, truncated=False,
    )
    sources = (manual,) + tuple(
        WorkspaceAIContextSource(
            source_id=f"source-{index}", source_scope="notebook", source_type="txt",
            title=f"Other document {index}", privacy_label="local_only",
            text=f"Unrelated finance procedure {index}.", included_chars=30, truncated=False,
        )
        for index in range(3)
    )

    scope = adapter.select_workspace_chat_preparation_scope(
        "Chế độ Manual Matecon ACR/CTU hoạt động như thế nào?", sources,
    )

    assert scope.bounded is True
    assert scope.sources == (manual,)
    assert scope.reason == "matched_sources"


def test_preparation_scope_can_limit_interactive_preparation_to_one_source():
    primary = WorkspaceAIContextSource(
        source_id="primary", source_scope="notebook", source_type="pdf",
        title="Matecon network connection", privacy_label="local_only",
        text="Matecon network connection configuration.", included_chars=40, truncated=False,
    )
    secondary = WorkspaceAIContextSource(
        source_id="secondary", source_scope="notebook", source_type="pdf",
        title="Matecon network guide", privacy_label="local_only",
        text="Matecon network connection guide.", included_chars=32, truncated=False,
    )

    scope = adapter.select_workspace_chat_preparation_scope(
        "Matecon network connection", (primary, secondary), limit=1
    )

    assert scope.bounded is True
    assert len(scope.sources) == 1


def test_existing_complete_semantic_index_is_ready_after_process_restart(monkeypatch, tmp_path):
    config = _semantic_config(tmp_path)
    source = _source("Already indexed BGE source.")
    executor = _ImmediateExecutor()
    monkeypatch.setattr(adapter, "_get_executor", lambda: executor)
    monkeypatch.setattr(adapter, "_durable_semantic_coverage_ready", lambda *_args: True)

    adapter.schedule_workspace_chat_source_preparation((source,), config=config)

    assert executor.submissions == []
    assert adapter.get_workspace_chat_source_preparation_status(
        (source,), config=config
    ) == {"temporary:source-1": "ready"}


def test_feature_flag_off_returns_no_evidence(tmp_path: Path):
    source = _source("Nội dung chủ sở hữu chỉ xử lý cục bộ.")
    config = adapter.WorkspaceChatRagV2CanaryConfig(
        enabled=False,
        runtime_root=tmp_path,
    )

    result = adapter.retrieve_workspace_chat_evidence(
        "Câu hỏi",
        (source,),
        config=config,
    )

    assert result["retrieval_available"] is False
    assert result["summary_count"] == 0
    assert result["evidence_items"] == []
    assert result["rag_v2_canary"]["backend"] == "unavailable"
    assert adapter.get_workspace_chat_source_preparation_status(
        (source,), config=config
    ) == {"temporary:source-1": "unavailable"}


def test_explicit_deep_never_silently_degrades_to_hybrid(tmp_path: Path):
    """The deep UI promise must not become a base-hybrid answer without reranking."""
    source = _source("Manual mode uses ctrlMode=1 before Matecon startup.")
    config = adapter.WorkspaceChatRagV2CanaryConfig(
        enabled=True,
        adaptive_enabled=False,
        runtime_root=tmp_path / "rag",
    )

    result = adapter.retrieve_workspace_chat_evidence(
        "How does Manual Matecon mode work?",
        (source,),
        config=config,
        search_preference="deep",
    )

    assert result["status"] == "quality_search_unavailable"
    assert result["retrieval_applied"] is False
    assert result["summary_count"] == 0
    assert result["rag_v2_canary"]["fallback_reason"] == "deep_search_unavailable"


def test_deep_search_availability_distinguishes_disabled_reranking_from_source_readiness(tmp_path: Path):
    config = adapter.WorkspaceChatRagV2CanaryConfig(
        enabled=True,
        adaptive_enabled=False,
        runtime_root=tmp_path / "rag",
    )

    availability = adapter.get_workspace_chat_deep_search_availability(config=config)

    assert availability.available is False
    assert availability.reason == "deep_disabled"


def test_deep_search_availability_requires_a_pinned_local_reranker(tmp_path: Path):
    reranker = tmp_path / "reranker"
    reranker.mkdir()
    config = adapter.WorkspaceChatRagV2CanaryConfig(
        enabled=True,
        adaptive_enabled=True,
        runtime_root=tmp_path / "rag",
        bge_reranker_model_path=reranker,
        bge_reranker_model_revision="reranker-revision",
        bge_reranker_model_checksum="sha256:" + "b" * 64,
    )

    availability = adapter.get_workspace_chat_deep_search_availability(config=config)

    assert availability.available is True


def test_missing_bge_pins_fail_closed_and_schedule_preparation(monkeypatch, tmp_path: Path):
    source = _source("ORCHID-731 là mã duy nhất của hồ sơ này.")
    executor = _ImmediateExecutor()
    monkeypatch.setattr(adapter, "_get_executor", lambda: executor)
    config = adapter.WorkspaceChatRagV2CanaryConfig(
        enabled=True,
        runtime_root=tmp_path,
    )

    result = adapter.retrieve_workspace_chat_evidence(
        "Mã dự án ORCHID-731 là gì?",
        (source,),
        config=config,
    )

    assert len(executor.submissions) == 1
    assert result["retrieval_available"] is False
    assert result["summary_count"] == 0
    assert result["rag_v2_canary"]["effective_profile"] == "unavailable"
    assert result["rag_v2_canary"]["fallback_applied"] is False
    assert not (tmp_path / "bge_m3_hybrid" / "workspace_chat.sqlite").exists()


def test_unprepared_query_remains_fail_closed_across_threads(monkeypatch, tmp_path: Path):
    source = _source("THREAD-882 là mã kiểm tra.")
    config = _semantic_config(tmp_path)
    executor = _ImmediateExecutor()
    monkeypatch.setattr(adapter, "_get_executor", lambda: executor)

    first = adapter.retrieve_workspace_chat_evidence("THREAD-882?", (source,), config=config)
    with ThreadPoolExecutor(max_workers=1) as callers:
        second = callers.submit(
            adapter.retrieve_workspace_chat_evidence,
            "THREAD-882?",
            (source,),
            config=config,
        ).result()

    assert first["retrieval_available"] is False
    assert second["retrieval_available"] is False
    assert len(executor.submissions) == 1
    assert adapter._RUNTIME_CACHE == {}


def test_changed_source_content_changes_preparation_key(tmp_path: Path):
    config = _semantic_config(tmp_path)
    first = _source("ALPHA-440 là trạng thái cũ.")
    updated = _source("BETA-991 là trạng thái mới.")

    assert adapter._preparation_key(config, first) != adapter._preparation_key(config, updated)


def test_runtime_cleanup_closes_and_evicts_cached_pipeline():
    assert adapter._RUNTIME_CACHE == {}
    adapter.close_workspace_chat_rag_v2_runtimes()
    assert adapter._RUNTIME_CACHE == {}


def test_runtime_cleanup_waits_for_background_preparation(monkeypatch):
    class RecordingExecutor:
        def __init__(self):
            self.wait = None

        def shutdown(self, *, wait):
            self.wait = wait

    executor = RecordingExecutor()
    monkeypatch.setattr(adapter, "_PREPARATION_EXECUTOR", executor)

    adapter.close_workspace_chat_rag_v2_runtimes()

    assert executor.wait is True
    assert adapter._PREPARATION_EXECUTOR is None


def test_failure_telemetry_does_not_expose_exception_text_or_paths(
    tmp_path: Path,
):
    secret_path = r"C:\private\owner\document.txt"
    source = _source("private owner text")
    config = _semantic_config(tmp_path)
    adapter.seed_workspace_chat_source_preparation(
        (source,),
        config=config,
        expected_source_fingerprints=[adapter._source_fingerprint(source)],
    )

    def failing_factory(_config):
        raise OSError(f"cannot open {secret_path}")

    result = adapter.retrieve_workspace_chat_evidence(
        "question",
        (source,),
        config=config,
        pipeline_factory=failing_factory,
    )

    telemetry = result["rag_v2_canary"]
    assert result["retrieval_available"] is False
    assert result["summary_count"] == 0
    assert telemetry["effective_profile"] == "unavailable"
    assert telemetry["fallback_applied"] is False
    assert secret_path not in str(telemetry)
    assert "private owner text" not in str(telemetry)


def test_adapter_keeps_provider_and_consent_outside_retrieval_boundary():
    source = Path("src/aios_habit/workspace_chat_rag_v2_adapter.py").read_text(
        encoding="utf-8"
    )
    app = Path("src/aios_habit/workspace_chat_app.py").read_text(encoding="utf-8")

    assert "RealWorkspaceAIProviderClient" not in source
    assert "BrainGateway" not in source
    assert "provider_client" not in source
    assert "retrieve_workspace_chat_evidence as retrieve_local_evidence" in app
    assert "retrieve_local_evidence(" in app
    assert app.index("retrieve_local_evidence(") < app.index(
        "generate_workspace_ai_answer(req, RealWorkspaceAIProviderClient())"
    )



def test_activated_manifest_is_authoritative_over_legacy_environment(
    monkeypatch,
    tmp_path: Path,
):
    model_path = tmp_path / "approved-model"
    model_path.mkdir()
    deployment = SimpleNamespace(
        activated=True,
        requested_profile="bge_m3_hybrid",
        runtime_root=tmp_path / "production-runtime",
        model_path=model_path,
        model_revision="approved-revision",
        model_checksum="sha256:" + "a" * 64,
        retrieval_device="cpu",
    )
    monkeypatch.setattr(
        adapter,
        "load_workspace_chat_rag_v2_deployment",
        lambda **_kwargs: deployment,
    )

    config = adapter.WorkspaceChatRagV2CanaryConfig.from_env(
        {
            adapter.CANARY_ENABLED_ENV: "0",
            adapter.PROFILE_ENV: "lexical_baseline",
            adapter.BGE_MODEL_CHECKSUM_ENV: "sha256:" + "b" * 64,
        }
    )

    assert config.enabled is True
    assert config.requested_profile == "bge_m3_hybrid"
    assert config.bge_m3_model_checksum == deployment.model_checksum
    assert config.fail_closed_on_error is True


def test_invalid_activated_manifest_returns_no_evidence_without_legacy(
    monkeypatch,
):
    monkeypatch.setenv(adapter.LOCAL_PILOT_ENABLED_ENV, "0")

    def fail_manifest(**_kwargs):
        raise adapter.DeploymentManifestError(
            r"checksum mismatch at C:\private\owner\model"
        )

    monkeypatch.setattr(adapter, "load_workspace_chat_rag_v2_deployment", fail_manifest)
    result = adapter.retrieve_workspace_chat_evidence(
        "question",
        (_source("private source contents"),),
    )

    assert result["summary_count"] == 0
    assert result["retrieval_available"] is False
    assert result["rag_v2_canary"]["backend"] == "unavailable"
    assert result["rag_v2_canary"]["fallback_applied"] is False
    assert "private" not in str(result["rag_v2_canary"])


def test_explicit_local_pilot_can_recover_when_historical_manifest_is_invalid(
    monkeypatch,
    tmp_path: Path,
):
    monkeypatch.setattr(
        adapter,
        "load_workspace_chat_rag_v2_deployment",
        lambda **_kwargs: (_ for _ in ()).throw(
            adapter.DeploymentManifestError("deployment_evidence_report_unavailable")
        ),
    )

    config = adapter.WorkspaceChatRagV2CanaryConfig.from_env(
        {
            adapter.LOCAL_PILOT_ENABLED_ENV: "1",
            adapter.CANARY_ENABLED_ENV: "1",
            adapter.BGE_MODEL_PATH_ENV: str(tmp_path / "bge-m3"),
            adapter.BGE_MODEL_REVISION_ENV: "test-revision",
            adapter.BGE_MODEL_CHECKSUM_ENV: "sha256:" + "a" * 64,
        }
    )

    assert config.enabled is True
    assert config.bge_m3_model_path == tmp_path / "bge-m3"


def test_activated_runtime_failure_never_falls_back(tmp_path: Path):
    config = _semantic_config(tmp_path)
    source = _source("private source contents")
    adapter.seed_workspace_chat_source_preparation(
        (source,),
        config=config,
        expected_source_fingerprints=[adapter._source_fingerprint(source)],
    )

    def failing_factory(_config):
        raise OSError(r"cannot open C:\private\owner\model")

    result = adapter.retrieve_workspace_chat_evidence(
        "question",
        (source,),
        config=config,
        pipeline_factory=failing_factory,
    )

    assert result["summary_count"] == 0
    assert result["retrieval_available"] is False
    assert result["rag_v2_canary"]["effective_profile"] == "unavailable"
    assert result["rag_v2_canary"]["fallback_applied"] is False
    assert "private" not in str(result["rag_v2_canary"])


def test_staged_manifest_is_inert_even_before_all_activation_fields_exist(
    tmp_path: Path,
):
    manifest = tmp_path / "deployment.json"
    manifest.write_text(
        json.dumps({"schema_version": 2, "activation_state": "staged"}),
        encoding="utf-8",
    )

    loaded = deployment_module.load_workspace_chat_rag_v2_deployment(
        env={deployment_module.DEPLOYMENT_MANIFEST_ENV: str(manifest)},
        require_activated=True,
    )

    assert loaded is None


def test_activated_manifest_seals_evidence_and_benchmark_files(tmp_path: Path):
    evidence_run_id = deployment_module.SELECTED_EVIDENCE_RUN_PREFIX + "test-corpus"
    corpus_fingerprint = "c" * 64
    model = tmp_path / "model"
    model.mkdir()
    (model / "model.bin").write_bytes(b"test model placeholder")
    runtime = (tmp_path / "production-runtime").resolve()
    runtime.mkdir()
    evidence_report = tmp_path / "selected_profile_report.json"
    evidence_report.write_text(
        json.dumps(
            {
                "status": "PASS",
                "qualification_passed": True,
                "qualification_id": evidence_run_id,
                "selected_profile": deployment_module.EXPECTED_PROFILE,
                "decision": "ADVANCE_TO_CANARY",
                "canary_allowed": True,
                "corpus_fingerprint": corpus_fingerprint,
            }
        ),
        encoding="utf-8",
    )
    benchmark_report = tmp_path / "benchmark_report.json"
    benchmark_fields = {
        "status": "PASS",
        "runtime_root": str(runtime),
        "effective_profile": deployment_module.EXPECTED_PROFILE,
        "fallback_applied": False,
        "warm_p95_ms": 2100.0,
        "runtime_init_count": 1,
        "memory_safe": True,
    }
    benchmark_report.write_text(json.dumps(benchmark_fields), encoding="utf-8")
    manifest = tmp_path / "deployment.json"
    manifest.write_text(
        json.dumps(
            {
                "schema_version": 2,
                "activation_state": "activated",
                "requested_profile": deployment_module.EXPECTED_PROFILE,
                "runtime": {"root": str(runtime)},
                "model": {
                    "path": str(model.resolve()),
                    "revision": deployment_module.EXPECTED_MODEL_REVISION,
                    "checksum": deployment_module.EXPECTED_MODEL_CHECKSUM,
                    "device": "cpu",
                },
                "policy": {
                    "lexical_fallback_enabled": False,
                    "semantic_progressive": False,
                    "fail_closed": True,
                },
                "evidence": {
                    "run_id": evidence_run_id,
                    "corpus_fingerprint": corpus_fingerprint,
                    "report_path": str(evidence_report),
                    "report_sha256": deployment_module.sha256_file(evidence_report),
                },
                "benchmark": {
                    **benchmark_fields,
                    "report_path": str(benchmark_report),
                    "report_sha256": deployment_module.sha256_file(benchmark_report),
                },
            }
        ),
        encoding="utf-8",
    )

    loaded = deployment_module.load_workspace_chat_rag_v2_deployment(
        manifest,
        require_activated=True,
    )
    assert loaded is not None
    assert loaded.activated is True
    assert loaded.requested_profile == deployment_module.EXPECTED_PROFILE
    identity = deployment_module.production_candidate_identity(loaded)
    assert identity["requested_profile"] == deployment_module.EXPECTED_PROFILE
    assert identity["model_revision"] == deployment_module.EXPECTED_MODEL_REVISION
    assert identity["model_checksum"] == deployment_module.EXPECTED_MODEL_CHECKSUM
    assert identity["benchmark_status"] == "PASS"
    assert identity["identity_sha256"].startswith("sha256:")
    assert "model_path" not in identity
    assert "runtime_root" not in identity

    benchmark_report.write_text(
        json.dumps({**benchmark_fields, "warm_p95_ms": 2999.0}),
        encoding="utf-8",
    )
    with pytest.raises(
        deployment_module.DeploymentManifestError,
        match="deployment_benchmark_report_changed",
    ):
        deployment_module.load_workspace_chat_rag_v2_deployment(
            manifest,
            require_activated=True,
        )


def test_production_identity_rejects_inactive_deployment(tmp_path: Path):
    deployment = deployment_module.WorkspaceChatRagV2Deployment(
        manifest_path=tmp_path / "deployment.json",
        activation_state=deployment_module.STAGED_STATE,
        requested_profile=deployment_module.EXPECTED_PROFILE,
        runtime_root=tmp_path / "runtime",
        model_path=tmp_path / "model",
        model_revision=deployment_module.EXPECTED_MODEL_REVISION,
        model_checksum=deployment_module.EXPECTED_MODEL_CHECKSUM,
        retrieval_device="cpu",
        fail_closed=True,
        evidence_run_id=deployment_module.SELECTED_EVIDENCE_RUN_PREFIX + "test-corpus",
        benchmark_status="PASS",
    )

    with pytest.raises(
        deployment_module.DeploymentManifestError,
        match="deployment_identity_requires_activated_manifest",
    ):
        deployment_module.production_candidate_identity(deployment)


def _enabled_config(tmp_path: Path, **overrides) -> adapter.WorkspaceChatRagV2CanaryConfig:
    values = {
        "enabled": True,
        "requested_profile": "bge_m3_hybrid",
        "runtime_root": tmp_path,
        "bge_m3_model_path": tmp_path / "pinned-bge-m3",
        "bge_m3_model_revision": "test-revision",
        "bge_m3_model_checksum": "sha256:" + "a" * 64,
        "fail_closed_on_error": True,
    }
    values.update(overrides)
    return adapter.WorkspaceChatRagV2CanaryConfig(**values)


class _ImmediateExecutor:
    def __init__(self):
        self.submissions = []

    def submit(self, fn, **kwargs):
        self.submissions.append((fn, kwargs))
        return SimpleNamespace()


def test_preparation_schedule_deduplicates_and_retry_is_explicit(monkeypatch, tmp_path):
    config = _enabled_config(tmp_path)
    source = _source("PREP-101 must be indexed once.")
    executor = _ImmediateExecutor()
    monkeypatch.setattr(adapter, "_get_executor", lambda: executor)

    with ThreadPoolExecutor(max_workers=8) as callers:
        futures = [
            callers.submit(
                adapter.schedule_workspace_chat_source_preparation,
                (source,),
                config=config,
            )
            for _ in range(24)
        ]
        for future in futures:
            future.result()

    assert len(executor.submissions) == 1
    assert adapter.get_workspace_chat_source_preparation_status(
        (source,), config=config
    ) == {"temporary:source-1": "pending"}

    key = adapter._preparation_key(config, source)
    adapter._PREPARATION_REGISTRY[key]["status"] = "failed"
    adapter.schedule_workspace_chat_source_preparation((source,), config=config)
    assert len(executor.submissions) == 1

    adapter.retry_workspace_chat_source_preparation((source,), config=config)
    assert len(executor.submissions) == 2
    assert adapter._PREPARATION_REGISTRY[key]["status"] == "pending"


def test_preparation_fingerprint_changes_for_content_privacy_and_config(tmp_path):
    source = _source("fingerprint content", privacy_label="local_only")
    changed_content = _source("fingerprint content changed", privacy_label="local_only")
    changed_privacy = _source("fingerprint content", privacy_label="cloud_safe")
    first_config = _enabled_config(tmp_path / "runtime-a")
    changed_config = _enabled_config(tmp_path / "runtime-b")

    keys = {
        adapter._preparation_key(first_config, source),
        adapter._preparation_key(first_config, changed_content),
        adapter._preparation_key(first_config, changed_privacy),
        adapter._preparation_key(changed_config, source),
    }

    assert len(keys) == 4


def test_preparation_status_identity_includes_scope(tmp_path):
    config = _enabled_config(tmp_path)
    temporary = _source("temporary text")
    notebook = WorkspaceAIContextSource(
        source_id=temporary.source_id,
        source_scope="notebook",
        source_type=temporary.source_type,
        title="notebook copy",
        privacy_label=temporary.privacy_label,
        text="notebook text",
        included_chars=len("notebook text"),
        truncated=False,
    )

    statuses = adapter.get_workspace_chat_source_preparation_status(
        (temporary, notebook), config=config
    )

    assert statuses == {
        "temporary:source-1": "not_prepared",
        "notebook:source-1": "not_prepared",
    }


def test_query_returns_no_evidence_while_semantic_is_queued(monkeypatch, tmp_path):
    config = _semantic_config(tmp_path)
    source = _source("READY-GATE-77 awaits semantic preparation.")
    executor = _ImmediateExecutor()
    monkeypatch.setattr(adapter, "_get_executor", lambda: executor)

    result = adapter.retrieve_workspace_chat_evidence(
        "What is READY-GATE-77?",
        (source,),
        config=config,
    )

    assert len(executor.submissions) == 1
    assert result["retrieval_available"] is False
    assert result["summary_count"] == 0
    assert result["rag_v2_canary"]["effective_profile"] == "unavailable"


def test_prepare_then_query_uses_prepared_registry(monkeypatch, tmp_path):
    config = _enabled_config(tmp_path)
    source = _source("PREPARED-808 is available after preparation.")

    class FakeReport:
        failed_count = 0
        unsupported_count = 0
        empty_count = 0
        converted_count = 1
        skipped_count = 0
        indexed_chunk_count = 1

    class FakePipeline:
        def ingest(self, _specs):
            return FakeReport()

        def close(self):
            pass

    result = adapter.prepare_workspace_chat_sources(
        (source,), config=config, pipeline_factory=lambda _config: FakePipeline()
    )

    assert result["status"] == "ok"
    assert adapter.get_workspace_chat_source_preparation_status(
        (source,), config=config
    ) == {"temporary:source-1": "ready"}


def _semantic_config(tmp_path: Path) -> adapter.WorkspaceChatRagV2CanaryConfig:
    return adapter.WorkspaceChatRagV2CanaryConfig(
        enabled=True,
        requested_profile="bge_m3_hybrid",
        runtime_root=tmp_path,
        bge_m3_model_path=tmp_path / "pinned-bge-m3",
        bge_m3_model_revision="test-revision",
        bge_m3_model_checksum="sha256:" + "a" * 64,
        fail_closed_on_error=True,
    )


def _sources(count: int) -> tuple[WorkspaceAIContextSource, ...]:
    return tuple(
        WorkspaceAIContextSource(
            source_id=f"source-{index}", source_scope="temporary",
            source_type="pasted_text", title=f"source-{index}.txt",
            privacy_label="local_only", text=f"evidence {index}",
            included_chars=len(f"evidence {index}"), truncated=False,
        )
        for index in range(count)
    )


def test_semantic_preparation_uses_bounded_incremental_batches(monkeypatch, tmp_path):
    config = _semantic_config(tmp_path)
    sources = _sources(5)
    staged_documents: list[str] = []
    initialized: list[object] = []

    def initialize_worker(_config):
        initialized.append(_config)
        return {
            "status": "ok", "reused": False, "init_latency_ms": 12.5,
            "readiness": {
                "protocol_version": "1", "retrieval_profile": "bge_m3_hybrid",
                "model": {"model_id": "BAAI/bge-m3", "fingerprint": "model-fingerprint"},
            },
        }

    def prepare_staged_source(spec, _config, *, group_size):
        staged_documents.append(spec.document_id)
        assert group_size == 4
        return {
            "converted_count": 1, "skipped_count": 0,
            "failed_count": 0, "indexed_chunk_count": len(staged_documents) * 10,
        }

    monkeypatch.setattr(
        adapter._SUBPROCESS_CLIENT, "initialize_worker", initialize_worker
    )
    monkeypatch.setattr(
        adapter._SUBPROCESS_CLIENT, "prepare_staged_source", prepare_staged_source
    )

    result = adapter.prepare_workspace_chat_sources(sources, config=config)

    assert len(initialized) == 1
    assert len(staged_documents) == 5
    assert len(set(staged_documents)) == 5
    assert all(document_id.startswith("wsc-") for document_id in staged_documents)
    assert result["status"] == "ok"
    assert result["report"]["initialization"] == {
        "status": "ok", "reused": False, "init_latency_ms": 12.5,
        "readiness": {
            "protocol_version": "1", "retrieval_profile": "bge_m3_hybrid",
            "model": {"model_id": "BAAI/bge-m3", "fingerprint": "model-fingerprint"},
        },
    }
    assert result["prepared_count"] == 5
    assert {key: value for key, value in result["report"].items() if key != "initialization"} == {
        "converted_count": 5,
        "skipped_count": 0,
        "failed_count": 0,
        "indexed_chunk_count": 50,
        "batch_count": 5,
        "batch_size": 1,
    }
    assert set(adapter.get_workspace_chat_source_preparation_status(sources, config=config).values()) == {"ready"}


def test_semantic_preparation_batch_failure_preserves_prior_commits(monkeypatch, tmp_path):
    config = _semantic_config(tmp_path)
    sources = _sources(5)
    calls = 0

    def prepare_staged_source(_spec, _config, *, group_size):
        nonlocal calls
        assert group_size == 4
        calls += 1
        if calls == 2:
            raise RuntimeError("sensitive source details must stay private")
        return {
            "converted_count": 1, "skipped_count": 0,
            "failed_count": 0, "indexed_chunk_count": 10,
        }

    def initialize_worker(_config):
        return {"status": "ok", "reused": True, "init_latency_ms": 0.0}

    monkeypatch.setattr(
        adapter._SUBPROCESS_CLIENT, "initialize_worker", initialize_worker
    )
    monkeypatch.setattr(
        adapter._SUBPROCESS_CLIENT, "prepare_staged_source", prepare_staged_source
    )

    with pytest.raises(
        RuntimeError,
        match=r"^preparation_batch_002_document_[0-9a-f]{12}_runtimeerror$",
    ):
        adapter.prepare_workspace_chat_sources(sources, config=config)

    assert calls == 2
    states = adapter.get_workspace_chat_source_preparation_status(sources, config=config)
    assert states["temporary:source-0"] == "ready"
    assert {states[f"temporary:source-{index}"] for index in range(1, 5)} == {"failed"}
    for source in sources[1:]:
        entry = adapter._PREPARATION_REGISTRY[adapter._preparation_key(config, source)]
        assert entry["reason"] == "runtimeerror"
        assert "sensitive" not in str(entry)


def test_semantic_preparation_resume_skips_completed_sources_and_emits_safe_progress(monkeypatch, tmp_path):
    config = _semantic_config(tmp_path)
    sources = _sources(3)
    completed_document_id = adapter._document_id(sources[0])
    staged_documents: list[str] = []
    progress_events: list[dict[str, object]] = []

    monkeypatch.setattr(
        adapter._SUBPROCESS_CLIENT,
        "initialize_worker",
        lambda _config: {"status": "ok", "reused": True, "init_latency_ms": 0.0},
    )

    def prepare_staged_source(spec, _config, *, group_size, **kwargs):
        assert group_size == 4
        assert kwargs["source_timeout_s"] == 30.0
        staged_documents.append(spec.document_id)
        return {
            "converted_count": 1, "skipped_count": 0,
            "failed_count": 0, "indexed_chunk_count": len(staged_documents),
        }

    monkeypatch.setattr(adapter._SUBPROCESS_CLIENT, "prepare_staged_source", prepare_staged_source)

    result = adapter.prepare_workspace_chat_sources(
        sources,
        config=config,
        completed_document_ids=(completed_document_id,),
        source_timeout_s=30.0,
        progress_callback=progress_events.append,
    )

    assert staged_documents == [adapter._document_id(sources[1]), adapter._document_id(sources[2])]
    assert result["prepared_count"] == 3
    assert result["resumed_count"] == 1
    assert progress_events == [
        {"document_id": adapter._document_id(sources[1]), "completed_count": 2, "total_sources": 3},
        {"document_id": adapter._document_id(sources[2]), "completed_count": 3, "total_sources": 3},
    ]
    assert all("source-" not in str(event) for event in progress_events)


def test_semantic_preparation_rejects_unknown_resume_document(monkeypatch, tmp_path):
    config = _semantic_config(tmp_path)
    source = _sources(1)

    monkeypatch.setattr(
        adapter._SUBPROCESS_CLIENT,
        "initialize_worker",
        lambda _config: {"status": "ok", "reused": True, "init_latency_ms": 0.0},
    )

    with pytest.raises(ValueError, match="completed_document_ids_unknown"):
        adapter.prepare_workspace_chat_sources(
            source,
            config=config,
            completed_document_ids=("wsc-not-a-current-source",),
        )


def test_semantic_preparation_initialization_failure_is_phase_safe(monkeypatch, tmp_path):
    config = _semantic_config(tmp_path)
    sources = _sources(2)

    def initialize_worker(_config):
        raise adapter.SemanticBackendError("bge_worker_init_timeout")

    monkeypatch.setattr(adapter._SUBPROCESS_CLIENT, "initialize_worker", initialize_worker)

    with pytest.raises(RuntimeError, match=r"^preparation_init_bge_worker_init_timeout$"):
        adapter.prepare_workspace_chat_sources(sources, config=config)

    assert set(adapter.get_workspace_chat_source_preparation_status(sources, config=config).values()) == {"failed"}
    for source in sources:
        entry = adapter._PREPARATION_REGISTRY[adapter._preparation_key(config, source)]
        assert entry["reason"] == "runtimeerror"


def test_batch_failure_reason_is_opaque_and_deterministic(tmp_path):
    config = _semantic_config(tmp_path)
    source = _sources(1)[0]
    spec = adapter._materialize_sources((source,), config.runtime_root)[0][0]

    reason = adapter._batch_failure_reason(7, (spec,), RuntimeError("private details"))

    assert reason.startswith("preparation_batch_007_document_")
    assert reason.endswith("_runtimeerror")
    assert source.source_id not in reason
    assert source.title not in reason
    assert source.text not in reason
    assert str(spec.path) not in reason


def test_query_config_is_read_only_and_uses_production_index(tmp_path):
    config = _semantic_config(tmp_path)

    preparation = adapter._pipeline_config(config, "bge_m3_hybrid")
    query = adapter._pipeline_config(
        config, "bge_m3_hybrid", read_only=True
    )

    assert preparation.index_read_only is False
    assert preparation.ensure_embeddings_on_open is True
    assert query.index_read_only is True
    assert query.ensure_embeddings_on_open is False
    assert query.index_path == preparation.index_path
    assert query.rerank_limit == config.deep_rerank_limit

    deep_query = adapter._pipeline_config(
        config, "bge_m3_hybrid", read_only=True, include_reranker=True
    )
    assert deep_query.rerank_limit == 30
    assert deep_query.retrieval_profile == "bge_m3_hybrid_rerank_expand"


def test_seeded_preparation_enables_read_only_query_without_preparing(tmp_path):
    config = _semantic_config(tmp_path)
    sources = _sources(2)
    fingerprints = [adapter._source_fingerprint(source) for source in sources]

    adapter.seed_workspace_chat_source_preparation(
        sources,
        config=config,
        expected_source_fingerprints=fingerprints,
    )

    assert set(
        adapter.get_workspace_chat_source_preparation_status(sources, config=config).values()
    ) == {"ready"}


def test_worker_initialization_accepts_staging_specific_timeout(tmp_path, monkeypatch):
    config = _semantic_config(tmp_path)
    observed = {}

    def initialize_worker(_config, *, timeout_s=300.0):
        observed["timeout_s"] = timeout_s
        return {"status": "ok", "reused": False, "init_latency_ms": 1.0, "readiness": {}}

    monkeypatch.setattr(adapter._SUBPROCESS_CLIENT, "initialize_worker", initialize_worker)

    result = adapter.initialize_workspace_chat_rag_v2_worker(config, timeout_s=601.0)

    assert result["status"] == "ok"
    assert observed["timeout_s"] == 601.0


def test_structured_excel_query_routes_before_semantic_preparation(tmp_path, monkeypatch):
    from openpyxl import Workbook
    import aios_habit.workspace_chat_source_ingest as ingest

    managed_root = tmp_path / "managed_workbooks"
    managed_root.mkdir()
    workbook_path = managed_root / "sales.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Region", "Revenue"])
    sheet.append(["North", 10])
    sheet.append(["North", 20])
    sheet.append(["South", 5])
    workbook.save(workbook_path)
    monkeypatch.setattr(ingest, "MANAGED_WORKBOOK_ROOT", managed_root)
    monkeypatch.setattr(
        adapter,
        "schedule_workspace_chat_source_preparation",
        lambda *_args, **_kwargs: pytest.fail("semantic preparation should not run"),
    )
    source = WorkspaceAIContextSource(
        source_id="excel-source",
        source_scope="temporary",
        source_type="xlsx",
        title="sales.xlsx",
        privacy_label="machine_only",
        text="Region | Revenue",
        included_chars=16,
        truncated=False,
        managed_path=str(workbook_path.resolve()),
    )
    config = adapter.WorkspaceChatRagV2CanaryConfig(enabled=True, runtime_root=tmp_path / "rag")

    result = adapter.retrieve_workspace_chat_evidence(
        "Tính tổng Revenue theo Region",
        (source,),
        config=config,
    )

    assert result["status"] == "structured_excel_query"
    assert result["rag_v2_canary"]["backend"] == "structured_excel_sqlite"
    assert result["summary_count"] == 1
    assert "North | 30" in result["evidence_items"][0]["text"]
    retrieved = result["retrieved_context_sources"][0]
    assert (retrieved.source_id, retrieved.source_scope) == ("excel-source", "temporary")
    assert retrieved.privacy_label == "machine_only"
def test_structured_excel_query_cites_all_contributing_sheets(tmp_path, monkeypatch):
    from openpyxl import Workbook
    import aios_habit.workspace_chat_source_ingest as ingest

    managed_root = tmp_path / "managed_workbooks"
    managed_root.mkdir()
    workbook_path = managed_root / "regional-sales.xlsx"
    workbook = Workbook()
    east = workbook.active
    east.title = "East"
    east.append(["Region", "Revenue"])
    east.append(["North", 10])
    west = workbook.create_sheet(title="West")
    west.append(["Region", "Revenue"])
    west.append(["South", 20])
    workbook.save(workbook_path)
    workbook.close()

    monkeypatch.setattr(ingest, "MANAGED_WORKBOOK_ROOT", managed_root)
    source = WorkspaceAIContextSource(
        source_id="regional-excel-source",
        source_scope="temporary",
        source_type="xlsx",
        title="regional-sales.xlsx",
        privacy_label="machine_only",
        text="Region | Revenue",
        included_chars=16,
        truncated=False,
        managed_path=str(workbook_path.resolve()),
    )
    config = adapter.WorkspaceChatRagV2CanaryConfig(enabled=True, runtime_root=tmp_path / "rag")

    result = adapter.retrieve_workspace_chat_evidence(
        "Tính tổng Revenue trên tất cả các sheet",
        (source,),
        config=config,
    )

    evidence = result["evidence_items"][0]
    assert result["status"] == "structured_excel_query"
    assert evidence["location_info"] == "Sheets: East, West"
    assert "Structured Excel result — multi-region (East, West)" in evidence["text"]
    assert result["citations"][0]["location"] == "Sheets: East, West"


def test_structured_excel_retains_priority_over_deep_preference(tmp_path, monkeypatch):
    from openpyxl import Workbook
    import aios_habit.workspace_chat_source_ingest as ingest

    managed_root = tmp_path / "managed_workbooks"
    managed_root.mkdir()
    workbook_path = managed_root / "metrics.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    ws.append(["Total", 100])
    workbook.save(workbook_path)
    workbook.close()

    monkeypatch.setattr(ingest, "MANAGED_WORKBOOK_ROOT", managed_root)
    source = WorkspaceAIContextSource(
        source_id="excel-src",
        source_scope="temporary",
        source_type="xlsx",
        title="metrics.xlsx",
        privacy_label="machine_only",
        text="Metric | Value",
        included_chars=14,
        truncated=False,
        managed_path=str(workbook_path.resolve()),
    )
    config = adapter.WorkspaceChatRagV2CanaryConfig(
        enabled=True,
        adaptive_enabled=True,
        runtime_root=tmp_path / "rag",
    )

    result = adapter.retrieve_workspace_chat_evidence(
        "Tính tổng Value theo Metric",
        (source,),
        config=config,
        search_preference="deep",
    )

    assert result["status"] == "structured_excel_query"
    assert result["rag_v2_canary"]["backend"] == "structured_excel_sqlite"


def test_search_preference_deep_overrides_fast_pre_gate(tmp_path, monkeypatch):
    source = _source("Hướng dẫn chấm công nhân sự theo quy định.")
    config = adapter.WorkspaceChatRagV2CanaryConfig(

        enabled=True,
        adaptive_enabled=True,
        runtime_root=tmp_path / "rag",
        bge_m3_model_path=tmp_path / "bge_m3",
        bge_m3_model_revision="rev",
        bge_m3_model_checksum="sha256:00",
        bge_reranker_model_path=tmp_path / "reranker",
        bge_reranker_model_revision="rerank-rev",
        bge_reranker_model_checksum="sha256:11",
    )

    captured_requests = []
    def mock_query_ready(question, specs, pipe_config, timeout_s=30.0, expansion=None, rerank_requested=False, routing_reason_codes=(), policy_version="adaptive-reranking-v1"):
        captured_requests.append({
            "rerank_requested": rerank_requested,
            "routing_reason_codes": routing_reason_codes,
            "timeout_s": timeout_s,
        })
        return {
            "summary": {
                "candidate_count": 5,
                "returned_count": 1,
                "filtered_as_stale_count": 0,
                "indexed_chunk_count": 1,
                "candidate_backend": "bge_m3_hybrid",
                "evidence_set_term_coverage": 0.9,
                "rerank_latency_ms": 5.0 if rerank_requested else 0.0,
            },
            "insufficiency_reasons": [],
            "items": [
                {
                    "document_id": "doc1",
                    "text": "Đoạn văn bản mẫu.",
                    "score": 0.95,
                    "citation_id": "c1",
                    "evidence_id": "e1",
                }
            ],
            "routing": {
                "reranker_requested": rerank_requested,
                "reranker_applied": rerank_requested,
                "effective_path": "hybrid_rerank" if rerank_requested else "hybrid",
                "degraded": False,
                "degraded_reason": "",
                "rerank_latency_ms": 5.0 if rerank_requested else 0.0,
                "policy_version": policy_version,
            },
            "synthesis": {
                "answer": "Câu trả lời.",
                "citation_ids": ["c1"],
                "claims": [],
                "grounded": True,
                "abstained": False,
                "abstention_reasons": [],
                "answer_mode": "local_evidence",
                "limitation_reasons": [],
                "provider_used": False,
                "mode": "local_retrieval_evidence",
            },
        }

    adapter.seed_workspace_chat_source_preparation(
        (source,),
        config=config,
        expected_source_fingerprints=[adapter._source_fingerprint(source)],
    )
    monkeypatch.setattr(adapter._SUBPROCESS_CLIENT, "query_ready", mock_query_ready)
    monkeypatch.setattr(
        adapter, "initialize_workspace_chat_rag_v2_worker", lambda *_args, **_kwargs: {}
    )
    monkeypatch.setattr(adapter, "_semantic_readiness", lambda sources, config: (adapter._PREPARATION_READY_STATE, ""))


    # Simple query (would be Fast in Auto), but user explicitly requested Deep
    result = adapter.retrieve_workspace_chat_evidence(
        "chấm công",
        (source,),
        config=config,
        search_preference="deep",
    )

    assert len(captured_requests) == 1
    assert captured_requests[0]["rerank_requested"] is True
    assert captured_requests[0]["timeout_s"] == config.deep_timeout_ms / 1000.0
    assert "user_requested_deep" in captured_requests[0]["routing_reason_codes"]
    assert result["rag_v2_canary"]["reranker_requested"] is True
    assert result["rag_v2_canary"]["reranker_applied"] is True
    assert "Đã tìm kỹ" in result["safe_owner_message"]


def test_adapter_degraded_reranker_telemetry_and_safe_owner_copy(tmp_path, monkeypatch):
    source = _source("Quy trình chấm công và tính lương.")
    config = adapter.WorkspaceChatRagV2CanaryConfig(
        enabled=True,
        adaptive_enabled=True,
        runtime_root=tmp_path / "rag",
        bge_m3_model_path=tmp_path / "bge_m3",
        bge_m3_model_revision="rev",
        bge_m3_model_checksum="sha256:00",
        bge_reranker_model_path=tmp_path / "reranker",
        bge_reranker_model_revision="rerank-rev",
        bge_reranker_model_checksum="sha256:11",
    )

    doc_id = adapter._document_id(source)

    def mock_query_ready(question, specs, pipe_config, timeout_s=30.0, expansion=None, rerank_requested=False, routing_reason_codes=(), policy_version="adaptive-reranking-v1"):
        return {
            "summary": {
                "candidate_count": 5,
                "returned_count": 2,
                "filtered_as_stale_count": 0,
                "indexed_chunk_count": 1,
                "candidate_backend": "bge_m3_hybrid",
                "evidence_set_term_coverage": 0.9,
                "rerank_latency_ms": 0.0,
            },
            "insufficiency_reasons": [],
            "items": [
                {
                    "document_id": doc_id,
                    "text": "Đoạn 1",
                    "score": 0.9,
                    "citation_id": "c1",
                    "evidence_id": "e1",
                },
                {
                    "document_id": doc_id,
                    "text": "Đoạn 2",
                    "score": 0.8,
                    "citation_id": "c2",
                    "evidence_id": "e2",
                }
            ],


            "routing": {
                "reranker_requested": True,
                "reranker_applied": False,
                "effective_path": "hybrid",
                "degraded": True,
                "degraded_reason": "reranker_backend_timeout",
                "rerank_latency_ms": 0.0,
                "policy_version": policy_version,
            },
            "synthesis": {
                "answer": "Câu trả lời.",
                "citation_ids": ["c1", "c2"],
                "claims": [],
                "grounded": True,
                "abstained": False,
                "abstention_reasons": [],
                "answer_mode": "local_evidence",
                "limitation_reasons": [],
                "provider_used": False,
                "mode": "local_retrieval_evidence",
            },
        }

    adapter.seed_workspace_chat_source_preparation(
        (source,),
        config=config,
        expected_source_fingerprints=[adapter._source_fingerprint(source)],
    )
    monkeypatch.setattr(adapter._SUBPROCESS_CLIENT, "query_ready", mock_query_ready)
    monkeypatch.setattr(
        adapter, "initialize_workspace_chat_rag_v2_worker", lambda *_args, **_kwargs: {}
    )
    monkeypatch.setattr(adapter, "_semantic_readiness", lambda sources, config: (adapter._PREPARATION_READY_STATE, ""))

    result = adapter.retrieve_workspace_chat_evidence(
        "chấm công",
        (source,),
        config=config,
        search_preference="deep",
    )

    telemetry = result["rag_v2_canary"]
    assert telemetry["reranker_requested"] is True
    assert telemetry["reranker_applied"] is False
    assert telemetry["degraded"] is True
    assert telemetry["degraded_reason"] == "reranker_backend_timeout"
    assert telemetry["effective_profile"] == "bge_m3_hybrid"
    assert "Đã tìm kỹ" not in result["safe_owner_message"]
    assert "Đã dùng 2 đoạn liên quan từ 1 nguồn." == result["safe_owner_message"]


def test_sqlite_preparation_ledger_crud_and_schema(tmp_path: Path):
    config = _enabled_config(tmp_path)
    db_path = adapter._get_ledger_db_path(config)
    adapter._init_preparation_ledger_db(db_path)

    row = adapter.SourcePreparationLedgerRow(
        source_scope="temporary",
        source_id="src-101",
        source_fingerprint="fp-101",
        model_id="BAAI/bge-m3",
        model_revision=config.bge_m3_model_revision,
        state=adapter.PREP_STATE_PENDING,
        priority=adapter.PREP_PRIORITY_NORMAL,
        document_id="doc-101",
        created_at=100.0,
        updated_at=100.0,
    )
    adapter._upsert_ledger_row(db_path, row)

    loaded = adapter._load_ledger_row(db_path, "temporary", "src-101")
    assert loaded is not None
    assert loaded.source_id == "src-101"
    assert loaded.state == adapter.PREP_STATE_PENDING
    assert loaded.priority == adapter.PREP_PRIORITY_NORMAL


def test_sqlite_stale_processing_recovery_on_startup(tmp_path: Path):
    config = _enabled_config(tmp_path)
    db_path = adapter._get_ledger_db_path(config)

    # Insert row directly in processing state
    adapter._init_preparation_ledger_db(db_path)
    row = adapter.SourcePreparationLedgerRow(
        source_scope="temporary",
        source_id="src-stale",
        source_fingerprint="fp-stale",
        model_id="BAAI/bge-m3",
        model_revision=config.bge_m3_model_revision,
        state=adapter.PREP_STATE_PROCESSING,
        priority=adapter.PREP_PRIORITY_NORMAL,
        document_id="doc-stale",
        created_at=50.0,
        updated_at=50.0,
    )
    adapter._upsert_ledger_row(db_path, row)

    # Re-init simulates process restart
    adapter._init_preparation_ledger_db(db_path)
    recovered = adapter._load_ledger_row(db_path, "temporary", "src-stale")
    assert recovered is not None
    assert recovered.state == adapter.PREP_STATE_PENDING


def test_sqlite_preparation_priority_ordering_and_atomic_claim(tmp_path: Path):
    config = _enabled_config(tmp_path)
    db_path = adapter._get_ledger_db_path(config)
    adapter._init_preparation_ledger_db(db_path)

    # Insert normal, backfill, interactive items
    for sid, prio, ts in [
        ("src-normal", adapter.PREP_PRIORITY_NORMAL, 10.0),
        ("src-backfill", adapter.PREP_PRIORITY_BACKFILL, 5.0),
        ("src-interactive", adapter.PREP_PRIORITY_INTERACTIVE, 20.0),
    ]:
        adapter._upsert_ledger_row(
            db_path,
            adapter.SourcePreparationLedgerRow(
                source_scope="temporary",
                source_id=sid,
                source_fingerprint=f"fp-{sid}",
                model_id="BAAI/bge-m3",
                model_revision=config.bge_m3_model_revision,
                state=adapter.PREP_STATE_PENDING,
                priority=prio,
                document_id=f"doc-{sid}",
                created_at=ts,
                updated_at=ts,
            ),
        )

    # First claim should get interactive
    item1 = adapter._claim_next_preparation_item(
        db_path, "BAAI/bge-m3", config.bge_m3_model_revision
    )
    assert item1 is not None
    assert item1.source_id == "src-interactive"
    assert item1.state == adapter.PREP_STATE_PROCESSING

    # While item1 is processing, next claim returns the same active item (single worker lock)
    item_dup = adapter._claim_next_preparation_item(
        db_path, "BAAI/bge-m3", config.bge_m3_model_revision
    )
    assert item_dup is not None
    assert item_dup.source_id == "src-interactive"

    # Commit item1 as ready
    adapter._commit_preparation_result(
        db_path, "temporary", "src-interactive", adapter.PREP_STATE_READY
    )

    # Next claim gets normal
    item2 = adapter._claim_next_preparation_item(
        db_path, "BAAI/bge-m3", config.bge_m3_model_revision
    )
    assert item2 is not None
    assert item2.source_id == "src-normal"
    adapter._commit_preparation_result(
        db_path, "temporary", "src-normal", adapter.PREP_STATE_READY
    )

    # Next claim gets backfill
    item3 = adapter._claim_next_preparation_item(
        db_path, "BAAI/bge-m3", config.bge_m3_model_revision
    )
    assert item3 is not None
    assert item3.source_id == "src-backfill"
    adapter._commit_preparation_result(
        db_path, "temporary", "src-backfill", adapter.PREP_STATE_READY
    )

    # Queue empty
    item4 = adapter._claim_next_preparation_item(
        db_path, "BAAI/bge-m3", config.bge_m3_model_revision
    )
    assert item4 is None


def test_reconcile_and_enqueue_preserves_ready_and_deduplicates(tmp_path: Path, monkeypatch):
    config = _enabled_config(tmp_path)
    executor = _ImmediateExecutor()
    monkeypatch.setattr(adapter, "_get_executor", lambda: executor)
    source_ready = _source("Ready source content", privacy_label="local_only")
    source_new = _source("New unread source content", privacy_label="local_only")
    source_new = adapter.WorkspaceAIContextSource(
        source_id="src-new",
        source_scope="temporary",
        source_type="pasted_text",
        title="new.txt",
        privacy_label="local_only",
        text="New unread source content",
        included_chars=24,
        truncated=False,
    )

    db_path = adapter._get_ledger_db_path(config)
    adapter._init_preparation_ledger_db(db_path)
    # Seed source_ready in ledger as ready
    adapter._upsert_ledger_row(
        db_path,
        adapter.SourcePreparationLedgerRow(
            source_scope=source_ready.source_scope,
            source_id=source_ready.source_id,
            source_fingerprint=adapter._source_fingerprint(source_ready),
            model_id="BAAI/bge-m3",
            model_revision=config.bge_m3_model_revision,
            state=adapter.PREP_STATE_READY,
            priority=adapter.PREP_PRIORITY_NORMAL,
            document_id=adapter._document_id(source_ready),
            created_at=10.0,
            updated_at=10.0,
        ),
    )

    # Reconcile both sources
    count = adapter.reconcile_and_enqueue_workspace_chat_sources(
        (source_ready, source_new),
        config=config,
    )
    # Only source_new needed enqueue
    assert count == 1


def test_promote_priority_to_interactive(tmp_path: Path):
    config = _enabled_config(tmp_path)
    db_path = adapter._get_ledger_db_path(config)
    adapter._init_preparation_ledger_db(db_path)

    adapter._upsert_ledger_row(
        db_path,
        adapter.SourcePreparationLedgerRow(
            source_scope="temporary",
            source_id="src-pending",
            source_fingerprint="fp-pending",
            model_id="BAAI/bge-m3",
            model_revision=config.bge_m3_model_revision,
            state=adapter.PREP_STATE_PENDING,
            priority=adapter.PREP_PRIORITY_NORMAL,
            document_id="doc-pending",
            created_at=10.0,
            updated_at=10.0,
        ),
    )

    promoted = adapter.promote_workspace_chat_source_priority(
        "temporary", "src-pending", adapter.PREP_PRIORITY_INTERACTIVE, config=config
    )
    assert promoted is True

    row = adapter._load_ledger_row(db_path, "temporary", "src-pending")
    assert row is not None
    assert row.priority == adapter.PREP_PRIORITY_INTERACTIVE


def test_get_workspace_chat_preparation_summary_aggregate(tmp_path: Path):
    config = _enabled_config(tmp_path)
    db_path = adapter._get_ledger_db_path(config)
    adapter._init_preparation_ledger_db(db_path)

    src1 = _source("Content 1")
    src2 = adapter.WorkspaceAIContextSource(
        source_id="src-2",
        source_scope="temporary",
        source_type="pasted_text",
        title="doc2.txt",
        privacy_label="local_only",
        text="Content 2",
        included_chars=9,
        truncated=False,
    )
    src3 = adapter.WorkspaceAIContextSource(
        source_id="src-3",
        source_scope="temporary",
        source_type="pasted_text",
        title="doc3.txt",
        privacy_label="local_only",
        text="Content 3",
        included_chars=9,
        truncated=False,
    )

    adapter._upsert_ledger_row(
        db_path,
        adapter.SourcePreparationLedgerRow(
            source_scope=src1.source_scope,
            source_id=src1.source_id,
            source_fingerprint=adapter._source_fingerprint(src1),
            model_id="BAAI/bge-m3",
            model_revision=config.bge_m3_model_revision,
            state=adapter.PREP_STATE_READY,
            priority=adapter.PREP_PRIORITY_NORMAL,
            document_id=adapter._document_id(src1),
            created_at=10.0,
            updated_at=10.0,
        ),
    )
    adapter._upsert_ledger_row(
        db_path,
        adapter.SourcePreparationLedgerRow(
            source_scope=src2.source_scope,
            source_id=src2.source_id,
            source_fingerprint=adapter._source_fingerprint(src2),
            model_id="BAAI/bge-m3",
            model_revision=config.bge_m3_model_revision,
            state=adapter.PREP_STATE_PENDING,
            priority=adapter.PREP_PRIORITY_NORMAL,
            document_id=adapter._document_id(src2),
            created_at=10.0,
            updated_at=10.0,
        ),
    )
    adapter._PREPARATION_REGISTRY[adapter._preparation_key(config, src2)] = (
        adapter._preparation_entry(config, src2, adapter.PREP_STATE_PROCESSING)
    )
    adapter._upsert_ledger_row(
        db_path,
        adapter.SourcePreparationLedgerRow(
            source_scope=src3.source_scope,
            source_id=src3.source_id,
            source_fingerprint=adapter._source_fingerprint(src3),
            model_id="BAAI/bge-m3",
            model_revision=config.bge_m3_model_revision,
            state=adapter.PREP_STATE_FAILED,
            priority=adapter.PREP_PRIORITY_NORMAL,
            last_error="backend_timeout",
            document_id=adapter._document_id(src3),
            created_at=10.0,
            updated_at=10.0,
        ),
    )

    summary = adapter.get_workspace_chat_preparation_summary(
        (src1, src2, src3), config=config
    )
    from aios_habit.workspace_chat_ui import format_preparation_summary_text
    summary_text = format_preparation_summary_text(summary, locale="vi")
    assert summary["total"] == 3
    assert summary["ready"] == 1
    assert summary["processing"] == 1
    assert summary["failed"] == 1
    assert summary["bge_available"] is True
    assert "BGE-M3: 1/3 sẵn sàng" in summary_text
    assert "đang đọc doc2.txt" in summary_text
    assert "1 lỗi" in summary_text


def test_forget_sources_deletes_ledger_rows(tmp_path: Path):
    config = _enabled_config(tmp_path)
    db_path = adapter._get_ledger_db_path(config)
    adapter._init_preparation_ledger_db(db_path)

    src = _source("To be deleted")
    adapter._upsert_ledger_row(
        db_path,
        adapter.SourcePreparationLedgerRow(
            source_scope=src.source_scope,
            source_id=src.source_id,
            source_fingerprint=adapter._source_fingerprint(src),
            model_id="BAAI/bge-m3",
            model_revision=config.bge_m3_model_revision,
            state=adapter.PREP_STATE_READY,
            priority=adapter.PREP_PRIORITY_NORMAL,
            document_id=adapter._document_id(src),
            created_at=10.0,
            updated_at=10.0,
        ),
    )

    assert adapter._load_ledger_row(db_path, src.source_scope, src.source_id) is not None
    adapter.forget_workspace_chat_sources((src,), config=config)
    assert adapter._load_ledger_row(db_path, src.source_scope, src.source_id) is None


def test_background_drain_queue_three_sources_sequential(tmp_path: Path, monkeypatch):
    config = _enabled_config(tmp_path)
    db_path = adapter._get_ledger_db_path(config)
    adapter._init_preparation_ledger_db(db_path)

    prepared_sources = []
    def fake_prepare(sources, *, config=None):
        for s in sources:
            prepared_sources.append(s.source_id)
            key = adapter._preparation_key(config, s)
            adapter._PREPARATION_REGISTRY[key] = adapter._preparation_entry(config, s, adapter.PREP_STATE_READY)
        return len(sources)

    monkeypatch.setattr(adapter, "prepare_workspace_chat_sources", fake_prepare)
    monkeypatch.setattr(adapter, "start_workspace_chat_background_drain", lambda *a, **kw: None)

    src1 = adapter.WorkspaceAIContextSource(
        source_id="src-1", source_scope="temporary", source_type="pasted_text",
        title="doc1.txt", privacy_label="local_only", text="Content 1", included_chars=9, truncated=False
    )
    src2 = adapter.WorkspaceAIContextSource(
        source_id="src-2", source_scope="temporary", source_type="pasted_text",
        title="doc2.txt", privacy_label="local_only", text="Content 2", included_chars=9, truncated=False
    )
    src3 = adapter.WorkspaceAIContextSource(
        source_id="src-3", source_scope="temporary", source_type="pasted_text",
        title="doc3.txt", privacy_label="local_only", text="Content 3", included_chars=9, truncated=False
    )

    enqueued = adapter.reconcile_and_enqueue_workspace_chat_sources(
        (src1, src2, src3), config=config, priority=adapter.PREP_PRIORITY_NORMAL
    )
    assert enqueued == 3

    adapter._drain_preparation_queue(config)

    assert len(prepared_sources) == 3
    for s in (src1, src2, src3):
        row = adapter._load_ledger_row(db_path, s.source_scope, s.source_id)
        assert row is not None
        assert row.state == adapter.PREP_STATE_READY


def test_background_drain_queue_upload_race_condition(tmp_path: Path, monkeypatch):
    config = _enabled_config(tmp_path)
    db_path = adapter._get_ledger_db_path(config)
    adapter._init_preparation_ledger_db(db_path)

    prepared_sources = []
    def fake_prepare(sources, *, config=None):
        for s in sources:
            prepared_sources.append(s.source_id)
            key = adapter._preparation_key(config, s)
            adapter._PREPARATION_REGISTRY[key] = adapter._preparation_entry(config, s, adapter.PREP_STATE_READY)
        return len(sources)

    monkeypatch.setattr(adapter, "prepare_workspace_chat_sources", fake_prepare)
    monkeypatch.setattr(adapter, "start_workspace_chat_background_drain", lambda *a, **kw: None)

    src1 = adapter.WorkspaceAIContextSource(
        source_id="src-1", source_scope="temporary", source_type="pasted_text",
        title="doc1.txt", privacy_label="local_only", text="Content 1", included_chars=9, truncated=False
    )
    src2 = adapter.WorkspaceAIContextSource(
        source_id="src-2", source_scope="temporary", source_type="pasted_text",
        title="doc2.txt", privacy_label="local_only", text="Content 2", included_chars=9, truncated=False
    )

    adapter.reconcile_and_enqueue_workspace_chat_sources((src1, src2), config=config)
    adapter._drain_preparation_queue(config)
    assert len(prepared_sources) == 2

    src3 = adapter.WorkspaceAIContextSource(
        source_id="src-3", source_scope="temporary", source_type="pasted_text",
        title="doc3.txt", privacy_label="local_only", text="Content 3", included_chars=9, truncated=False
    )
    adapter.reconcile_and_enqueue_workspace_chat_sources((src3,), config=config)
    adapter._drain_preparation_queue(config)

    assert len(prepared_sources) == 3
    row3 = adapter._load_ledger_row(db_path, src3.source_scope, src3.source_id)
    assert row3 is not None
    assert row3.state == adapter.PREP_STATE_READY


def test_background_drain_true_concurrency_race_condition(tmp_path: Path, monkeypatch):
    """Test true concurrent multi-threaded enqueueing while background worker is draining."""
    import threading
    import time

    config = _enabled_config(tmp_path)
    db_path = adapter._get_ledger_db_path(config)
    adapter._init_preparation_ledger_db(db_path)

    prepared_sources = []
    prep_lock = threading.Lock()

    def fake_prepare(sources, *, config=None):
        time.sleep(0.01)  # simulated preparation delay
        with prep_lock:
            for s in sources:
                prepared_sources.append(s.source_id)
                key = adapter._preparation_key(config, s)
                adapter._PREPARATION_REGISTRY[key] = adapter._preparation_entry(config, s, adapter.PREP_STATE_READY)
        return len(sources)

    monkeypatch.setattr(adapter, "prepare_workspace_chat_sources", fake_prepare)

    sources = [
        adapter.WorkspaceAIContextSource(
            source_id=f"concurrent-src-{i}",
            source_scope="temporary",
            source_type="pasted_text",
            title=f"doc_{i}.txt",
            privacy_label="local_only",
            text=f"Content for doc {i}",
            included_chars=20,
            truncated=False,
        )
        for i in range(10)
    ]

    def worker_producer(sub_sources):
        for s in sub_sources:
            time.sleep(0.005)
            adapter.reconcile_and_enqueue_workspace_chat_sources((s,), config=config)

    t1 = threading.Thread(target=worker_producer, args=(sources[:5],))
    t2 = threading.Thread(target=worker_producer, args=(sources[5:],))

    t1.start()
    t2.start()
    t1.join()
    t2.join()

    deadline = time.time() + 5.0
    while time.time() < deadline:
        with prep_lock:
            if len(prepared_sources) == 10:
                break
        time.sleep(0.05)

    with prep_lock:
        assert len(prepared_sources) == 10, f"Expected 10 prepared sources, got {len(prepared_sources)}"

    for s in sources:
        row = adapter._load_ledger_row(db_path, s.source_scope, s.source_id)
        assert row is not None
        assert row.state == adapter.PREP_STATE_READY


def test_drain_worker_double_check_lock_prevents_lost_enqueue_race(tmp_path: Path, monkeypatch):
    """Deterministically exercise the double-check race condition window."""
    import threading
    import time

    config = _enabled_config(tmp_path)
    db_path = adapter._get_ledger_db_path(config)
    adapter._init_preparation_ledger_db(db_path)

    prepared_sources = []
    def fake_prepare(sources, *, config=None):
        for s in sources:
            prepared_sources.append(s.source_id)
            key = adapter._preparation_key(config, s)
            adapter._PREPARATION_REGISTRY[key] = adapter._preparation_entry(config, s, adapter.PREP_STATE_READY)
        return len(sources)

    monkeypatch.setattr(adapter, "prepare_workspace_chat_sources", fake_prepare)

    src1 = adapter.WorkspaceAIContextSource(
        source_id="race-doc-1",
        source_scope="temporary",
        source_type="pasted_text",
        title="doc_1.txt",
        privacy_label="local_only",
        text="Content 1",
        included_chars=10,
        truncated=False,
    )
    src2 = adapter.WorkspaceAIContextSource(
        source_id="race-doc-2",
        source_scope="temporary",
        source_type="pasted_text",
        title="doc_2.txt",
        privacy_label="local_only",
        text="Content 2",
        included_chars=10,
        truncated=False,
    )

    orig_claim = adapter._claim_next_preparation_item
    first_empty_seen = threading.Event()
    second_item_enqueued = threading.Event()

    def instrumented_claim(p_db_path, model_id, model_revision):
        res = orig_claim(p_db_path, model_id, model_revision)
        if res is None and not first_empty_seen.is_set():
            first_empty_seen.set()
            second_item_enqueued.wait(timeout=2.0)
        return res

    monkeypatch.setattr(adapter, "_claim_next_preparation_item", instrumented_claim)

    def producer_thread():
        first_empty_seen.wait(timeout=2.0)
        adapter.reconcile_and_enqueue_workspace_chat_sources((src2,), config=config)
        second_item_enqueued.set()

    t_prod = threading.Thread(target=producer_thread, daemon=True)
    t_prod.start()

    adapter.reconcile_and_enqueue_workspace_chat_sources((src1,), config=config)

    t_prod.join(timeout=3.0)

    deadline = time.time() + 5.0
    while time.time() < deadline:
        if len(prepared_sources) == 2:
            break
        time.sleep(0.05)

    assert "race-doc-1" in prepared_sources
    assert "race-doc-2" in prepared_sources
    assert len(prepared_sources) == 2

    row1 = adapter._load_ledger_row(db_path, src1.source_scope, src1.source_id)
    row2 = adapter._load_ledger_row(db_path, src2.source_scope, src2.source_id)
    assert row1 is not None and row1.state == adapter.PREP_STATE_READY
    assert row2 is not None and row2.state == adapter.PREP_STATE_READY
