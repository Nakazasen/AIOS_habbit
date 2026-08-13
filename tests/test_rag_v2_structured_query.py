from pathlib import Path

import openpyxl
import pytest

from aios_habit.rag_v2.structured_query import (
    StructuredAggregate,
    StructuredFilter,
    StructuredOrder,
    StructuredQueryBoundsError,
    StructuredQueryPlan,
    StructuredQueryValidationError,
    execute_excel_query,
)


def _workbook(path: Path, rows, *, sheet: str = "Sales") -> Path:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def test_structured_query_normalizes_blank_and_duplicate_headers(tmp_path):
    path = _workbook(tmp_path / "headers.xlsx", [["Name", "", "Name"], ["A", 10, "North"]])
    result = execute_excel_query(path, StructuredQueryPlan(sheet="Sales", select_columns=("Name", "column_2", "Name_2")))
    assert result.columns == ("Name", "column_2", "Name_2")
    assert result.rows == (("A", 10, "North"),)
    assert result.source_name == "headers.xlsx"
    assert result.provenance[0].sheet == "Sales"
    assert result.provenance[0].cell_range == "A2:C2"


def test_structured_query_infers_numbers_and_applies_parameterized_filters(tmp_path):
    path = _workbook(tmp_path / "sales.xlsx", [["Region", "Revenue"], ["North", "10.5"], ["South", "7"], ["North", "bad"]])
    result = execute_excel_query(path, StructuredQueryPlan(select_columns=("Region", "Revenue"), filters=(StructuredFilter("Revenue", ">=", 8),), order_by=(StructuredOrder("Revenue", "desc"),)))
    assert result.rows == (("North", 10.5),)
    assert result.provenance[0].excel_rows == (2,)
    assert "North" in result.rendered_evidence
    assert "Sales!A1:B4" in result.rendered_evidence


def test_structured_query_groups_and_aggregates(tmp_path):
    path = _workbook(tmp_path / "aggregate.xlsx", [["Region", "Revenue"], ["North", 10], ["South", 7], ["North", 5]])
    result = execute_excel_query(path, StructuredQueryPlan(group_by=("Region",), aggregates=(StructuredAggregate("sum", "Revenue", "total_revenue"), StructuredAggregate("count", "*", "row_count")), order_by=(StructuredOrder("total_revenue", "desc"),)))
    assert result.columns == ("Region", "total_revenue", "row_count")
    assert result.rows == (("North", 15.0, 2), ("South", 7.0, 1))
    assert result.provenance[0].excel_rows == (2, 4)
    assert result.provenance[1].excel_rows == (3,)


def test_structured_query_rejects_identifier_and_operator_injection(tmp_path):
    path = _workbook(tmp_path / "safe.xlsx", [["Value"], [1]])
    with pytest.raises(StructuredQueryValidationError, match="Unknown column"):
        execute_excel_query(path, StructuredQueryPlan(select_columns=('Value"; DROP TABLE data; --',)))
    with pytest.raises(StructuredQueryValidationError, match="Unsupported filter operator"):
        execute_excel_query(path, StructuredQueryPlan(filters=(StructuredFilter("Value", "= 1; DROP TABLE data", 1),)))
    result = execute_excel_query(path, StructuredQueryPlan(filters=(StructuredFilter("Value", "=", "1' OR 1=1 --"),)))
    assert result.rows == ()


def test_structured_query_enforces_cell_and_result_bounds(tmp_path, monkeypatch):
    path = _workbook(tmp_path / "bounded.xlsx", [["A", "B"], [1, 2], [3, 4]])
    monkeypatch.setenv("AIOS_STRUCTURED_SQL_MAX_CELLS", "5")
    with pytest.raises(StructuredQueryBoundsError, match="cell limit"):
        execute_excel_query(path, StructuredQueryPlan())
    monkeypatch.setenv("AIOS_STRUCTURED_SQL_MAX_CELLS", "100")
    result = execute_excel_query(path, StructuredQueryPlan(limit=9999))
    assert result.row_count == 2
    assert result.truncated is False


def test_excel_engine_fails_soft_for_unsupported_path(tmp_path: Path):
    unsupported = execute_excel_query(
        tmp_path / "missing.csv",
        StructuredQueryPlan(select_columns=("x",)),
    )
    assert unsupported.applied is False
    assert unsupported.reason == "unsupported_or_missing_workbook"


def test_deterministic_planner_builds_grouped_average_plan():
    from aios_habit.rag_v2.structured_query import ExcelTableSchema, plan_excel_query

    planned = plan_excel_query(
        "Tính trung bình Revenue theo Region, lấy top 3",
        (ExcelTableSchema("Sales", ("Region", "Revenue")),),
    )

    assert planned.applied is True
    assert planned.plan is not None
    assert planned.plan.sheet == "Sales"
    assert planned.plan.group_by == ("Region",)
    assert planned.plan.select_columns == ("Region",)
    assert planned.plan.aggregates[0].function == "avg"
    assert planned.plan.aggregates[0].column == "Revenue"
    assert planned.plan.order_by[0].direction == "desc"
    assert planned.plan.limit == 3


def test_deterministic_planner_rejects_raw_sql_and_falls_back_for_prose():
    from aios_habit.rag_v2.structured_query import ExcelTableSchema, plan_excel_query

    schemas = (ExcelTableSchema("Sales", ("Region", "Revenue")),)
    raw_sql = plan_excel_query("SELECT * FROM Sales", schemas)
    prose = plan_excel_query("Hãy giải thích báo cáo này", schemas)

    assert raw_sql.applied is False
    assert raw_sql.reason == "raw_sql_rejected"
    assert prose.applied is False
    assert prose.reason == "unstructured_intent"


def test_structured_query_rejects_grouped_non_grouped_projection(tmp_path):
    path = _workbook(tmp_path / "grouped.xlsx", [["Region", "Revenue"], ["N", 1]])
    with pytest.raises(StructuredQueryValidationError, match="Grouped query selections"):
        execute_excel_query(path, StructuredQueryPlan(select_columns=("Revenue",), group_by=("Region",), aggregates=(StructuredAggregate("sum", "Revenue", "total"),)))


def test_planner_understands_vietnamese_synonyms_and_diacritics():
    from aios_habit.rag_v2.structured_query import ExcelTableSchema, plan_excel_query

    planned = plan_excel_query(
        "Tính tổng doanh thu theo khu vực",
        (ExcelTableSchema("Báo Cáo", ("Region", "Revenue")),),
    )
    assert planned.applied is True
    assert planned.plan is not None
    assert planned.plan.sheet == "Báo Cáo"
    assert planned.plan.group_by == ("Region",)
    assert planned.plan.aggregates[0].function == "sum"
    assert planned.plan.aggregates[0].column == "Revenue"


def test_planner_parses_date_range_and_multi_conditions():
    from aios_habit.rag_v2.structured_query import ExcelTableSchema, plan_excel_query

    planned = plan_excel_query(
        "Tổng doanh thu theo khu vực trong tháng 8/2026 và Region là North",
        (ExcelTableSchema("Sales", ("Region", "Revenue", "Date")),),
    )
    assert planned.applied is True
    assert planned.plan is not None
    assert len(planned.plan.filters) >= 3
    operators = [f.operator for f in planned.plan.filters]
    assert ">=" in operators
    assert "<=" in operators
    assert "=" in operators


def test_planner_rejects_or_filter_logic():
    from aios_habit.rag_v2.structured_query import ExcelTableSchema, plan_excel_query

    planned = plan_excel_query(
        "Liệt kê doanh thu cho Region là North hoặc Region là South",
        (ExcelTableSchema("Sales", ("Region", "Revenue")),),
    )
    assert planned.applied is False
    assert "filter_validation_error" in planned.reason


def test_planner_selects_correct_sheet_and_fails_soft_on_ambiguity():
    from aios_habit.rag_v2.structured_query import ExcelTableSchema, plan_excel_query

    schemas = (
        ExcelTableSchema("East", ("Region", "Revenue")),
        ExcelTableSchema("West", ("Branch", "Amount")),
    )
    planned = plan_excel_query("Tổng Revenue ở sheet East", schemas)
    assert planned.applied is True
    assert planned.plan.sheet == "East"

    ambiguous_schemas = (
        ExcelTableSchema("East", ("Region", "Revenue")),
        ExcelTableSchema("West", ("Region", "Sales")),
    )
    ambiguous = plan_excel_query("Tính tổng Revenue", ambiguous_schemas)
    assert ambiguous.applied is False
    assert ambiguous.reason == "ambiguous_sheet_table"


def test_structured_query_multi_region_execution(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "East"
    ws1.append(["Region", "Revenue"])
    ws1.append(["North", 100])

    ws2 = wb.create_sheet(title="West")
    ws2.append(["Region", "Revenue"])
    ws2.append(["South", 200])

    path = tmp_path / "multi_region.xlsx"
    wb.save(path)
    wb.close()

    result = execute_excel_query(
        path,
        StructuredQueryPlan(
            target_regions=("East!A1:B2", "West!A1:B2"),
            aggregates=(StructuredAggregate("sum", "Revenue", "total_rev"),),
        ),
    )
    assert result.applied is True
    assert result.rows[0][0] == 300.0
    assert len(result.provenance) >= 1

    detail_result = execute_excel_query(
        path,
        StructuredQueryPlan(
            target_regions=("East!A1:B2", "West!A1:B2"),
            select_columns=("Region", "Revenue"),
        ),
    )
    assert detail_result.applied is True
    assert detail_result.rows == (("North", 100), ("South", 200))
    assert detail_result.provenance[0].sheet == "East"
    assert detail_result.provenance[1].sheet == "West"


def test_month_slash_year_extraction_preserves_year():
    from aios_habit.rag_v2.structured_query import ExcelTableSchema, plan_excel_query

    planned = plan_excel_query(
        "Tổng doanh thu trong tháng 8/2025",
        (ExcelTableSchema("Sales", ("Revenue", "Date")),),
    )
    assert planned.applied is True
    assert planned.plan is not None
    filters = planned.plan.filters
    assert any(f.value == "2025-08-01" for f in filters)
    assert any(f.value == "2025-08-31" for f in filters)


def test_parse_iso_date_rejects_invalid_calendar_date():
    from aios_habit.rag_v2.structured_query import _parse_iso_date

    assert _parse_iso_date("2026-02-31") is None
    assert _parse_iso_date("2026-02-28") == "2026-02-28"


def test_identical_schema_ambiguity_rejected():
    from aios_habit.rag_v2.structured_query import ExcelTableSchema, plan_excel_query

    ambiguous_schemas = (
        ExcelTableSchema("East", ("Region", "Revenue"), "A1:B2"),
        ExcelTableSchema("West", ("Region", "Revenue"), "A1:B2"),
    )
    ambiguous = plan_excel_query("Tính tổng Revenue", ambiguous_schemas)
    assert ambiguous.applied is False
    assert ambiguous.reason == "ambiguous_sheet_table"

    all_planned = plan_excel_query("Tính tổng Revenue trên tất cả các sheet", ambiguous_schemas)
    assert all_planned.applied is True
    assert all_planned.plan.target_regions == ("East!A1:B2", "West!A1:B2")

    no_range_schemas = (
        ExcelTableSchema("East", ("Region", "Revenue")),
        ExcelTableSchema("West", ("Region", "Revenue")),
    )
    no_range_all = plan_excel_query("Tổng Revenue trên tất cả các sheet", no_range_schemas)
    assert no_range_all.applied is True
    assert no_range_all.plan.target_regions == ("East", "West")


def test_quantity_listing_intent_selects_column_instead_of_count_aggregate():
    from aios_habit.rag_v2.structured_query import ExcelTableSchema, plan_excel_query

    planned = plan_excel_query(
        "Liệt kê số lượng theo sản phẩm",
        (ExcelTableSchema("Inventory", ("Product", "Quantity")),),
    )
    assert planned.applied is True
    assert planned.plan is not None
    assert planned.plan.group_by == ("Product",)
    assert planned.plan.select_columns == ("Quantity", "Product") or "Quantity" in planned.plan.select_columns
    assert len(planned.plan.aggregates) == 0


def test_multi_sheet_aggregate_provenance_and_rendered_evidence(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "East"
    ws1.append(["Region", "Revenue"])
    ws1.append(["North", 100])

    ws2 = wb.create_sheet(title="West")
    ws2.append(["Region", "Revenue"])
    ws2.append(["North", 200])

    path = tmp_path / "grouped_multi.xlsx"
    wb.save(path)
    wb.close()

    result = execute_excel_query(
        path,
        StructuredQueryPlan(
            target_regions=("East", "West"),
            group_by=("Region",),
            aggregates=(StructuredAggregate("sum", "Revenue", "total_rev"),),
        ),
    )
    assert result.applied is True
    assert result.rows == (("North", 300.0),)
    provenance_sheets = tuple(p.sheet for p in result.provenance)
    assert "East" in provenance_sheets
    assert "West" in provenance_sheets
    assert "multi-region (East, West)" in result.rendered_evidence
def test_planner_does_not_treat_all_substring_as_all_sheets_intent():
    from aios_habit.rag_v2.structured_query import ExcelTableSchema, plan_excel_query

    schemas = (
        ExcelTableSchema("East", ("Revenue",)),
        ExcelTableSchema("West", ("Revenue",)),
    )
    planned = plan_excel_query("smallest Revenue", schemas)

    assert planned.applied is False
    assert planned.reason == "ambiguous_sheet_table"


def test_aggregate_provenance_preserves_comma_in_excel_sheet_name(tmp_path):
    wb = openpyxl.Workbook()
    comma_sheet = wb.active
    comma_sheet.title = "East,West"
    comma_sheet.append(["Revenue"])
    comma_sheet.append([100])
    north_sheet = wb.create_sheet(title="North")
    north_sheet.append(["Revenue"])
    north_sheet.append([200])

    path = tmp_path / "comma_sheet.xlsx"
    wb.save(path)
    wb.close()
    result = execute_excel_query(
        path,
        StructuredQueryPlan(
            target_regions=("East,West", "North"),
            aggregates=(StructuredAggregate("sum", "Revenue", "total_revenue"),),
        ),
    )

    assert result.applied is True
    assert result.rows == ((300.0,),)
    assert tuple(provenance.sheet for provenance in result.provenance) == ("East,West", "North")
