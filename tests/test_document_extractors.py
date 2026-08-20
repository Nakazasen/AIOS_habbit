import zipfile
from pathlib import Path
import openpyxl

from aios_habit.document_extractors import _extract_excel


def test_extract_excel_with_shapes(tmp_path):
    file_path = tmp_path / "mock.xlsx"
    
    # Create valid base workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Cell Text"
    wb.save(file_path)
    wb.close()
    
    # Append drawing XML to the existing zip
    drawing_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
        <xdr:twoCellAnchor>
            <xdr:sp>
                <xdr:txBody>
                    <a:p>
                        <a:r>
                            <a:t>Mock Shape Text</a:t>
                        </a:r>
                    </a:p>
                </xdr:txBody>
            </xdr:sp>
        </xdr:twoCellAnchor>
    </xdr:wsDr>
    """
    with zipfile.ZipFile(file_path, "a") as archive:
        archive.writestr("xl/drawings/drawing1.xml", drawing_xml)
        
    results = _extract_excel(file_path)
    assert any("Cell Text" in r.text for r in results)
    assert any("Mock Shape Text" in r.text for r in results)

import fitz
from aios_habit.document_extractors import _extract_pdf, extract_text_chunks_from_file

def test_extract_pdf_with_text(tmp_path):
    file_path = tmp_path / "mock.pdf"
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text(fitz.Point(50, 50), "Hello from fake PDF page 1")
    doc.save(str(file_path))
    doc.close()
    
    results = _extract_pdf(file_path)
    assert len(results) == 1
    assert "Hello from fake PDF page 1" in results[0].text
    assert results[0].page in {"1", ""}
    assert results[0].extraction_status == "extracted"
    assert results[0].element_type in {"pdf_page_text", "pdf_markdown_page"}

def test_extract_pdf_empty_scanned(tmp_path):
    file_path = tmp_path / "mock_empty.pdf"
    doc = fitz.open()
    doc.new_page()
    doc.save(str(file_path))
    doc.close()

    results = _extract_pdf(file_path)
    assert len(results) == 1
    assert results[0].text == ""
    assert results[0].extraction_status == "empty_text"

def test_extract_text_chunks_from_pdf(tmp_path):
    file_path = tmp_path / "mock.pdf"
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text(fitz.Point(50, 50), "Hello from fake PDF chunking")
    doc.save(str(file_path))
    doc.close()

    chunks = extract_text_chunks_from_file(file_path, root=tmp_path)
    assert len(chunks) == 1
    chunk = chunks[0]
    assert "Hello from fake PDF chunking" in chunk["text"]
    assert chunk["page"] in {"1", ""}
    assert chunk["extraction_status"] == "extracted"
    assert chunk["element_type"] in {"pdf_page_text", "pdf_markdown_page"}
    assert chunk.get("privacy_level") == "local_only"

def test_extract_pdf_missing_dependency(monkeypatch, tmp_path):
    import builtins
    import sys

    monkeypatch.setitem(sys.modules, "pdf_inspector", None)
    monkeypatch.setitem(sys.modules, "fitz", None)
    original_import = builtins.__import__

    def blocked_import(name, *args, **kwargs):
        if name in {"pdf_inspector", "fitz"}:
            raise ImportError(f"blocked {name}")
        return original_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", blocked_import)
    file_path = tmp_path / "mock.pdf"
    file_path.write_bytes(b"%PDF-1.4\n")

    results = _extract_pdf(file_path)
    assert len(results) == 1
    assert results[0].text == ""
    assert results[0].extraction_status == "dependency_missing"
    assert "PDF extraction unavailable" in results[0].warning


def test_route_pdf_pages_uses_pdf_inspector_markdown(monkeypatch, tmp_path):
    import sys
    from types import SimpleNamespace
    from aios_habit.document_extractors import route_pdf_pages

    fake_module = SimpleNamespace(
        extract_pages_markdown=lambda path: SimpleNamespace(
            pages=[SimpleNamespace(page=0, markdown="# Tiêu đề\n\nNội dung tiếng Việt", needs_ocr=False)],
            pages_with_tables=[1],
            pages_with_columns=[],
        )
    )
    monkeypatch.setitem(sys.modules, "pdf_inspector", fake_module)
    file_path = tmp_path / "native.pdf"
    file_path.write_bytes(b"not-read-by-fake-parser")

    routes = route_pdf_pages(file_path)
    assert len(routes) == 1
    assert routes[0].page == 1
    assert routes[0].extractor == "pdf_inspector"
    assert routes[0].needs_ocr is False
    assert routes[0].has_table is True
    assert "Nội dung tiếng Việt" in routes[0].text


def test_route_pdf_pages_falls_back_to_pymupdf(monkeypatch, tmp_path):
    import sys
    from types import SimpleNamespace
    from aios_habit.document_extractors import route_pdf_pages

    fake_module = SimpleNamespace(
        extract_pages_markdown=lambda path: (_ for _ in ()).throw(RuntimeError("parser failed"))
    )
    monkeypatch.setitem(sys.modules, "pdf_inspector", fake_module)
    file_path = tmp_path / "fallback.pdf"
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((50, 50), "Fallback native text works")
    doc.save(file_path)
    doc.close()

    routes = route_pdf_pages(file_path)
    assert routes[0].extractor == "pymupdf_fallback"
    assert routes[0].needs_ocr is False
    assert "Fallback native text works" in routes[0].text
    assert "parser failed" in routes[0].warning


def test_route_pdf_pages_rescues_false_positive_ocr(monkeypatch, tmp_path):
    import sys
    from types import SimpleNamespace
    from aios_habit.document_extractors import route_pdf_pages

    fake_module = SimpleNamespace(
        extract_pages_markdown=lambda path: SimpleNamespace(
            pages=[SimpleNamespace(page=0, markdown="", needs_ocr=True)],
            pages_with_tables=[],
            pages_with_columns=[],
        )
    )
    monkeypatch.setitem(sys.modules, "pdf_inspector", fake_module)
    file_path = tmp_path / "native-rescue.pdf"
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((50, 50), "Native text must avoid unnecessary OCR")
    doc.save(file_path)
    doc.close()

    routes = route_pdf_pages(file_path)
    assert routes[0].extractor == "pymupdf_native_rescue"
    assert routes[0].needs_ocr is False
    assert "Native text must avoid unnecessary OCR" in routes[0].text
    assert "requested OCR" in routes[0].warning



from aios_habit.document_extractors import _extract_docx, _extract_html, _extract_pptx, normalize_extracted_text, local_capabilities
from aios_habit.extractor_registry import adapter_status, extract_with_registry, metadata_only_fallback, registered_extensions


def _write_zip(path, members):
    with zipfile.ZipFile(path, "w") as z:
        for name, data in members.items():
            z.writestr(name, data)


def test_pptx_adapter_extracts_slide_text_and_notes(tmp_path):
    pptx = tmp_path / "deck.pptx"
    _write_zip(pptx, {
        "ppt/slides/slide1.xml": "<a:t>Slide Title</a:t><a:t>ManualShipping table text</a:t>",
        "ppt/notesSlides/notesSlide1.xml": "<a:t>Speaker note safe text</a:t>",
    })
    res = _extract_pptx(pptx)
    assert res.extraction_status == "extracted_success"
    assert "Slide Title" in res.text
    assert "Speaker note" in res.text


def test_docx_adapter_extracts_paragraphs_and_tables(tmp_path):
    docx = tmp_path / "doc.docx"
    _write_zip(docx, {"word/document.xml": '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body><w:p><w:r><w:t>Heading Text</w:t></w:r></w:p><w:tbl><w:tr><w:tc><w:p><w:r><w:t>Cell Value</w:t></w:r></w:p></w:tc></w:tr></w:tbl></w:body></w:document>'})
    res = _extract_docx(docx)
    assert res.extraction_status == "extracted_success"
    assert "Heading Text" in res.text
    assert "Cell Value" in res.text


def test_html_adapter_keeps_mermaid_and_removes_script(tmp_path):
    html_path = tmp_path / "diagram.html"
    html_path.write_text("<html><head><title>ERD</title><style>.x{}</style><script>secret()</script></head><body><h1>Diagram</h1><pre class='mermaid'>A-->B</pre><table><tr><td>Order</td></tr></table></body></html>", encoding="utf-8")
    res = _extract_html(html_path)
    assert res.extraction_status == "extracted_success"
    assert "Diagram" in res.text
    assert "A-->B" in res.text
    assert "secret" not in res.text


def test_registry_routes_and_fallback_metadata_only(tmp_path):
    assert ".pdf" in registered_extensions()
    status = adapter_status(".pdf")
    assert status.status == "available"
    unknown = tmp_path / "x.unknown"
    unknown.write_text("ignored", encoding="utf-8")
    elements = extract_with_registry(unknown, root=tmp_path)
    assert elements[0].extraction_status == "dependency_missing"
    assert elements[0].text == ""
    assert elements[0].metadata["_is_metadata_only"] == "True"


def test_metadata_only_fallback_schema(tmp_path):
    p = tmp_path / "x.bin"
    p.write_bytes(b"abc")
    el = metadata_only_fallback(p, tmp_path, status="parse_failed")[0]
    assert el.relative_path == "x.bin"
    assert el.element_type == "metadata_only"
    assert el.extraction_status == "parse_failed"


def test_image_metadata_or_ocr_safe(tmp_path):
    from PIL import Image
    img = tmp_path / "img.png"
    Image.new("RGB", (20, 10), "white").save(img)
    chunks = extract_text_chunks_from_file(img, root=tmp_path)
    assert chunks[0]["file_type"] == ".png"
    assert chunks[0]["extraction_status"] in {"unsupported_no_local_ocr", "ocr_partial", "ocr_success", "failed_with_reason"}
    assert local_capabilities()["cloud_ocr_used"] is False


def test_normalization_reduces_coordinate_garbage_and_preserves_multilingual():
    text = "10 20 30 40 50\nThiết kế 変更 ManualShipping\n   nhiều   khoảng trắng   "
    cleaned = normalize_extracted_text(text)
    assert "10 20 30 40 50" not in cleaned
    assert "Thiết kế 変更 ManualShipping" in cleaned


def test_ocr_router_prefers_rapidocr_and_stops_after_quality_pass(monkeypatch):
    import aios_habit.ocr_engines as engines

    calls = []
    monkeypatch.delenv("AIOS_OCR_ENGINE_ORDER", raising=False)
    monkeypatch.setattr(engines, "engine_availability", lambda: {
        "rapidocr": True, "paddleocr": True, "tesseract": True,
    })
    monkeypatch.setattr(engines, "run_rapidocr", lambda image: (
        calls.append("rapidocr") or engines.OCREngineResult(
            text="Xin chào tài liệu", confidence=92.0, confidence_samples=3, engine="rapidocr",
        )
    ))
    monkeypatch.setattr(engines, "run_paddleocr", lambda image: (
        calls.append("paddleocr") or engines.OCREngineResult(engine="paddleocr")
    ))

    result, attempts = engines.run_ocr_router(
        object(), meaningful=lambda text: bool(text.strip()), minimum_confidence=35.0,
        tesseract_fallback=lambda image: calls.append("tesseract"), mode="balanced",
    )

    assert result.engine == "rapidocr"
    assert attempts == 1
    assert calls == ["rapidocr"]


def test_ocr_router_escalates_after_failed_quality_gate(monkeypatch):
    import aios_habit.ocr_engines as engines

    calls = []
    monkeypatch.delenv("AIOS_OCR_ENGINE_ORDER", raising=False)
    monkeypatch.setattr(engines, "engine_availability", lambda: {
        "rapidocr": True, "paddleocr": True, "tesseract": True,
    })
    monkeypatch.setattr(engines, "run_rapidocr", lambda image: (
        calls.append("rapidocr") or engines.OCREngineResult(
            text="bad", confidence=10.0, confidence_samples=1, engine="rapidocr",
        )
    ))
    monkeypatch.setattr(engines, "run_paddleocr", lambda image: (
        calls.append("paddleocr") or engines.OCREngineResult(
            text="Bảng dữ liệu hợp lệ", confidence=88.0, confidence_samples=4, engine="paddleocr",
        )
    ))

    result, attempts = engines.run_ocr_router(
        object(), meaningful=lambda text: len(text) > 5, minimum_confidence=35.0,
        mode="balanced",
    )

    assert result.engine == "paddleocr"
    assert attempts == 2
    assert calls == ["rapidocr", "paddleocr"]


def test_fast_mode_never_loads_heavy_fallbacks(monkeypatch):
    from aios_habit.ocr_engines import configured_engine_order

    monkeypatch.delenv("AIOS_OCR_ENGINE_ORDER", raising=False)
    assert configured_engine_order("fast") == ["rapidocr"]


def test_deep_pdf_mode_is_opt_in(monkeypatch, tmp_path):
    from aios_habit.document_extractors import PDFPageRoute, _deep_pdf_result

    monkeypatch.setenv("AIOS_OCR_MODE", "balanced")
    result = _deep_pdf_result(tmp_path / "not-opened.pdf", [PDFPageRoute(page=1, has_table=True)])
    assert result is None


def test_excel_streaming_row_chunking_2000_rows(tmp_path):
    from aios_habit.excel_extractors import extract_excel

    file_path = tmp_path / "large_production_bom.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM_Master"

    headers = ["Part_ID", "Part_Name", "Quantity", "Unit_Cost", "Total_Value", "Status"]
    ws.append(headers)

    for i in range(1, 2001):
        ws.append([f"P-{i:04d}", f"Part Item {i}", i * 2, i * 1.5, i * 3.0, "ACTIVE" if i % 2 == 0 else "PENDING"])

    wb.save(file_path)
    wb.close()

    result = extract_excel(file_path)
    assert result.succeeded is True
    assert result.error == ""
    assert len(result.truncated_reasons) == 0
    assert len(result.regions) == 4

    expected_headers = tuple(headers)
    for index, region in enumerate(result.regions):
        assert region.headers == expected_headers
        assert region.chunk_index == index
        assert region.total_chunks == 4
        assert region.sheet == "BOM_Master"
        assert region.header_rows == (expected_headers,)
        assert region.rows[0] == expected_headers

    assert result.regions[0].row_range == (1, 501)
    assert result.regions[0].cell_range == "A1:F501"
    assert result.regions[0].rows[1][0] == "P-0001"
    assert result.regions[0].rows[-1][0] == "P-0500"

    assert result.regions[1].row_range == (502, 1001)
    assert result.regions[1].cell_range == "A502:F1001"
    assert result.regions[1].rows[1][0] == "P-0501"
    assert result.regions[1].rows[-1][0] == "P-1000"

    assert result.regions[2].row_range == (1002, 1501)
    assert result.regions[2].cell_range == "A1002:F1501"
    assert result.regions[2].rows[1][0] == "P-1001"
    assert result.regions[2].rows[-1][0] == "P-1500"

    assert result.regions[3].row_range == (1502, 2001)
    assert result.regions[3].cell_range == "A1502:F2001"
    assert result.regions[3].rows[1][0] == "P-1501"
    assert result.regions[3].rows[-1][0] == "P-2000"


def test_excel_no_cell_count_truncation_30k_cells(tmp_path):
    from aios_habit.excel_extractors import extract_excel

    file_path = tmp_path / "wide_30k_cells.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WideData"

    cols = [f"Col_{c}" for c in range(1, 31)]
    ws.append(cols)

    for r in range(1, 1001):
        ws.append([f"R{r}C{c}" for c in range(1, 31)])

    wb.save(file_path)
    wb.close()

    result = extract_excel(file_path)
    assert result.succeeded is True
    assert not any("cell limit" in r for r in result.truncated_reasons)
    assert len(result.regions) == 2
    assert result.regions[0].total_chunks == 2
    assert result.regions[1].total_chunks == 2


def test_document_extractors_excel_streaming_integration(tmp_path):
    file_path = tmp_path / "orders_1200.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    ws.append(["Order_ID", "Customer", "Amount", "Status"])
    for i in range(1, 1201):
        ws.append([f"ORD-{i:04d}", f"Cust_{i}", i * 100, "COMPLETED"])

    wb.save(file_path)
    wb.close()

    results = _extract_excel(file_path)
    assert len(results) == 3

    assert results[0].section == "table A1:D501 (chunk 1/3)"
    assert results[0].row_range == "1-501"
    assert "Table range: A1:D501 (Chunk 1/3)" in results[0].text
    assert "Columns: Order_ID | Customer | Amount | Status" in results[0].text
    assert "Row 2: ORD-0001 | Cust_1 | 100 | COMPLETED" in results[0].text
    assert "Row 501: ORD-0500 | Cust_500 | 50000 | COMPLETED" in results[0].text

    assert results[1].section == "table A502:D1001 (chunk 2/3)"
    assert results[1].row_range == "502-1001"
    assert "Table range: A502:D1001 (Chunk 2/3)" in results[1].text
    assert "Columns: Order_ID | Customer | Amount | Status" in results[1].text
    assert "Row 502: ORD-0501 | Cust_501 | 50100 | COMPLETED" in results[1].text
    assert "Row 1001: ORD-1000 | Cust_1000 | 100000 | COMPLETED" in results[1].text

    assert results[2].section == "table A1002:D1201 (chunk 3/3)"
    assert results[2].row_range == "1002-1201"
    assert "Table range: A1002:D1201 (Chunk 3/3)" in results[2].text
    assert "Columns: Order_ID | Customer | Amount | Status" in results[2].text
    assert "Row 1002: ORD-1001 | Cust_1001 | 10100 | COMPLETED" in results[2].text
    assert "Row 1201: ORD-1200 | Cust_1200 | 120000 | COMPLETED" in results[2].text


def test_excel_extraction_config_custom_chunk_size(tmp_path):
    from aios_habit.excel_extractors import ExcelExtractionConfig, extract_excel

    file_path = tmp_path / "custom_chunks.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Custom"
    ws.append(["ID", "Name"])
    for i in range(1, 251):
        ws.append([i, f"Name_{i}"])
    wb.save(file_path)
    wb.close()

    cfg = ExcelExtractionConfig(chunk_row_size=100)
    result = extract_excel(file_path, config=cfg)
    assert len(result.regions) == 3
    assert result.regions[0].row_range == (1, 101)
    assert result.regions[0].chunk_index == 0
    assert result.regions[0].total_chunks == 3
    assert result.regions[1].row_range == (102, 201)
    assert result.regions[1].chunk_index == 1
    assert result.regions[1].total_chunks == 3
    assert result.regions[2].row_range == (202, 251)
    assert result.regions[2].chunk_index == 2
    assert result.regions[2].total_chunks == 3
