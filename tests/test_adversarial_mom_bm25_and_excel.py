from __future__ import annotations

import ast
import math
from pathlib import Path
from typing import Any
import pytest
import openpyxl

from aios_habit.mom_local_index import (
    MomChunk,
    MomSearchHit,
    _tokens,
    build_mom_local_index,
    search_mom_index,
)
from aios_habit.excel_extractors import (
    ExcelCell,
    ExcelExtractionConfig,
    ExcelTableRegion,
    extract_excel,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]


# ===========================================================================
# 1. MOM BM25 Adversarial Challenge & Stress Tests
# ===========================================================================

def test_mom_bm25_empty_and_whitespace_queries(tmp_path, monkeypatch):
    """Stress test BM25 with empty strings, whitespace, newlines, and control chars."""
    monkeypatch.chdir(tmp_path)
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    (doc_dir / "sample.md").write_text("Alpha Bravo Charlie Delta Echo", encoding="utf-8")
    build_mom_local_index(doc_dir)

    for empty_q in ["", "   ", "\t", "\n\r", "   \t\n  ", "!@#$%^&*()_+~`", "   ...   "]:
        hits = search_mom_index(empty_q)
        assert hits == [], f"Query {empty_q!r} should return empty list, got {hits}"


def test_mom_bm25_single_character_queries(tmp_path, monkeypatch):
    """Stress test single ASCII, numeric, and CJK character queries."""
    monkeypatch.chdir(tmp_path)
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    (doc_dir / "a_doc.md").write_text("A B C single character test 1 2 3", encoding="utf-8")
    (doc_dir / "cjk_doc.md").write_text("製 造 履 歴 単 語", encoding="utf-8")
    build_mom_local_index(doc_dir)

    # Single ASCII character
    hits_a = search_mom_index("a", limit=5)
    assert len(hits_a) >= 1
    assert all(h.score >= 0.0 for h in hits_a)
    assert "a" in hits_a[0].matched_terms

    # Single Digit
    hits_1 = search_mom_index("1", limit=5)
    assert len(hits_1) >= 1
    assert hits_1[0].score > 0.0

    # Single CJK character
    hits_cjk = search_mom_index("製", limit=5)
    assert len(hits_cjk) >= 1
    assert "cjk_doc.md" in hits_cjk[0].chunk.relative_path
    assert hits_cjk[0].score > 0.0


def test_mom_bm25_rare_cjk_compounds_and_variations(tmp_path, monkeypatch):
    """Stress test rare and multi-syllable CJK compounds with n-gram decomposition."""
    monkeypatch.chdir(tmp_path)
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()

    # Document with 4-char and rare Japanese/Chinese manufacturing terms
    (doc_dir / "japan_mfg.md").write_text(
        "生産管理システムにおける自動化工程および製造履歴登録手順の解説。\n"
        "出荷指示データと倉庫管理(WMS)連携仕様。",
        encoding="utf-8",
    )
    (doc_dir / "general_spec.md").write_text(
        "General specification for standard interface and data structures.",
        encoding="utf-8",
    )
    build_mom_local_index(doc_dir)

    # 4-character CJK query
    hits_4char = search_mom_index("自動化工程", limit=3)
    assert len(hits_4char) >= 1
    assert "japan_mfg.md" in hits_4char[0].chunk.relative_path
    assert hits_4char[0].score > 5.0  # high score due to n-grams + exact phrase boost

    # 6-character CJK compound
    hits_6char = search_mom_index("製造履歴登録手順", limit=3)
    assert len(hits_6char) >= 1
    assert "japan_mfg.md" in hits_6char[0].chunk.relative_path

    # Partial CJK compound match should rank below full compound match
    hits_partial = search_mom_index("製造", limit=3)
    assert len(hits_partial) >= 1
    assert hits_4char[0].score > hits_partial[0].score or hits_6char[0].score > hits_partial[0].score


def test_mom_bm25_deeply_nested_underscore_identifiers(tmp_path, monkeypatch):
    """Stress test deeply nested snake_case identifiers, symbols, and token splits."""
    monkeypatch.chdir(tmp_path)
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()

    (doc_dir / "api_schema.md").write_text(
        "Schema definition: mom_prod_v2_order_registration_batch_async_handler\n"
        "Also contains sys_core_config_database_connection_pool_timeout_ms",
        encoding="utf-8",
    )
    (doc_dir / "other.md").write_text("Unrelated document about general order processing.", encoding="utf-8")
    build_mom_local_index(doc_dir)

    # Full nested identifier
    full_ident = "mom_prod_v2_order_registration_batch_async_handler"
    hits_full = search_mom_index(full_ident, limit=3)
    assert len(hits_full) >= 1
    assert "api_schema.md" in hits_full[0].chunk.relative_path
    assert full_ident in hits_full[0].matched_terms or "mom" in hits_full[0].matched_terms

    # Partial nested subterm search
    hits_sub = search_mom_index("registration_batch", limit=3)
    assert len(hits_sub) >= 1
    assert "api_schema.md" in hits_sub[0].chunk.relative_path

    # Individual sub-atom from deeply nested identifier
    hits_atom = search_mom_index("async_handler", limit=3)
    assert len(hits_atom) >= 1
    assert "api_schema.md" in hits_atom[0].chunk.relative_path


def test_mom_bm25_identical_scores_and_tie_breaking(tmp_path, monkeypatch):
    """Verify tie-breaking, deduplication, and file diversification when documents have identical scores."""
    monkeypatch.chdir(tmp_path)
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()

    # Create 5 distinct files with exact identical content
    for i in range(1, 6):
        (doc_dir / f"identical_doc_{i:02d}.md").write_text(
            f"Exact duplicate manufacturing procedure for part inspection.\nIndex doc copy {i}.",
            encoding="utf-8",
        )
    build_mom_local_index(doc_dir)

    hits = search_mom_index("manufacturing procedure part inspection", limit=5)
    assert len(hits) == 5

    # Scores should all be non-negative and identical or near-identical
    scores = [h.score for h in hits]
    assert all(s > 0.0 for s in scores)

    # Diversification ensures distinct files appear in the top 5
    files = [h.chunk.relative_path for h in hits]
    assert len(set(files)) == 5, f"Expected 5 distinct files, got {files}"


def test_mom_bm25_score_monotonicity_and_strict_non_negativity(tmp_path, monkeypatch):
    """Mathematically verify that BM25 scores are strictly monotonic with term frequency and strictly non-negative."""
    monkeypatch.chdir(tmp_path)
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()

    # doc1 has term 1 time, doc2 has term 5 times, doc3 has term 20 times
    (doc_dir / "low_freq.md").write_text("telemetry " + "filler " * 100, encoding="utf-8")
    (doc_dir / "mid_freq.md").write_text("telemetry " * 5 + "filler " * 100, encoding="utf-8")
    (doc_dir / "high_freq.md").write_text("telemetry " * 20 + "filler " * 100, encoding="utf-8")
    (doc_dir / "no_freq.md").write_text("filler " * 120, encoding="utf-8")

    build_mom_local_index(doc_dir)

    hits = search_mom_index("telemetry", limit=10)
    assert len(hits) == 3

    hit_map = {h.chunk.relative_path: h.score for h in hits}
    assert hit_map["high_freq.md"] > hit_map["mid_freq.md"] > hit_map["low_freq.md"] > 0.0

    # Test that un-matched terms get zero hits / strictly >= 0.0
    all_hits = search_mom_index("completely_nonexistent_token_xyz999", limit=5)
    assert all_hits == []


def test_mom_bm25_zero_hardcode_ast_verification():
    """Verify mom_local_index.py contains zero hardcoded q1/q2/q3 term sets or negative file penalties."""
    file_path = PROJECT_ROOT / "src/aios_habit/mom_local_index.py"
    source = file_path.read_text(encoding="utf-8")
    tree = ast.parse(source)

    forbidden_names = {"q1_terms", "q2_terms", "q3_terms"}
    for node in ast.walk(tree):
        if isinstance(node, ast.Name):
            assert node.id not in forbidden_names, f"Found forbidden identifier: {node.id}"
        elif isinstance(node, ast.Constant) and isinstance(node.value, str):
            for bad in ["q1_terms", "q2_terms", "q3_terms", "erd_kho_van_new.html"]:
                assert bad not in node.value.lower(), f"Found forbidden string literal: {node.value}"
        elif isinstance(node, ast.UnaryOp) and isinstance(node.op, ast.USub):
            if isinstance(node.operand, ast.Constant) and node.operand.value in (50, 50.0):
                pytest.fail("Found forbidden -50.0 penalty constant in AST")


# ===========================================================================
# 2. Excel Streaming Row-Chunking Adversarial Challenge & Stress Tests
# ===========================================================================

def test_excel_streaming_large_workbook_over_1500_rows(tmp_path):
    """Stress test large workbook with 1,850 rows, verifying streaming chunking without truncation."""
    file_path = tmp_path / "large_inventory_1850.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory_Records"

    headers = ["SKU_Code", "Item_Name", "Warehouse_Bin", "Quantity_On_Hand", "Reorder_Level", "Unit_Price"]
    ws.append(headers)

    total_rows = 1850
    for r in range(1, total_rows + 1):
        ws.append([
            f"SKU-{r:05d}",
            f"Component Part #{r}",
            f"BIN-{r % 50:02d}",
            r * 10,
            50,
            round(12.50 + (r % 100) * 0.25, 2),
        ])

    wb.save(file_path)
    wb.close()

    result = extract_excel(file_path)
    assert result.succeeded is True
    assert result.error == ""
    assert len(result.truncated_reasons) == 0

    # Default chunk_row_size is 500. 1850 data rows -> ceil(1850/500) = 4 chunks.
    assert len(result.regions) == 4

    expected_headers = tuple(headers)
    for idx, region in enumerate(result.regions):
        assert region.chunk_index == idx
        assert region.total_chunks == 4
        assert region.sheet == "Inventory_Records"
        assert region.headers == expected_headers
        assert region.header_rows == (expected_headers,)
        assert region.rows[0] == expected_headers

    # Verify row boundaries and exact data continuity
    # Chunk 0: 500 rows (data 1 to 500)
    assert result.regions[0].row_range == (1, 501)
    assert result.regions[0].cell_range == "A1:F501"
    assert result.regions[0].rows[1][0] == "SKU-00001"
    assert result.regions[0].rows[-1][0] == "SKU-00500"

    # Chunk 1: 500 rows (data 501 to 1000)
    assert result.regions[1].row_range == (502, 1001)
    assert result.regions[1].cell_range == "A502:F1001"
    assert result.regions[1].rows[1][0] == "SKU-00501"
    assert result.regions[1].rows[-1][0] == "SKU-01000"

    # Chunk 2: 500 rows (data 1001 to 1500)
    assert result.regions[2].row_range == (1002, 1501)
    assert result.regions[2].cell_range == "A1002:F1501"
    assert result.regions[2].rows[1][0] == "SKU-01001"
    assert result.regions[2].rows[-1][0] == "SKU-01500"

    # Chunk 3: 350 rows (data 1501 to 1850)
    assert result.regions[3].row_range == (1502, 1851)
    assert result.regions[3].cell_range == "A1502:F1851"
    assert result.regions[3].rows[1][0] == "SKU-01501"
    assert result.regions[3].rows[-1][0] == "SKU-01850"


def test_excel_multi_level_hierarchical_headers_2_and_3_rows(tmp_path):
    """Stress test 2-row and 3-row hierarchical header structures with merged column categories."""
    file_path = tmp_path / "multi_level_headers.xlsx"
    wb = openpyxl.Workbook()

    # Sheet 1: 2-level headers
    ws2 = wb.active
    ws2.title = "TwoLevel"
    ws2.append(["Financial Year 2026", "Financial Year 2026", "Operations", "Operations"])
    ws2.append(["Q1 Revenue", "Q2 Revenue", "Staff Count", "Shift Hours"])
    for i in range(1, 101):
        ws2.append([1000 * i, 1200 * i, 50 + i, 8])

    # Sheet 2: 3-level headers
    ws3 = wb.create_sheet(title="ThreeLevel")
    ws3.append(["Asia Pacific", "Asia Pacific", "Europe", "Europe"])
    ws3.append(["Vietnam", "Japan", "Germany", "France"])
    ws3.append(["Hanoi Plant", "Tokyo Plant", "Munich Plant", "Paris Plant"])
    for i in range(1, 101):
        ws3.append([i * 10, i * 20, i * 30, i * 40])

    wb.save(file_path)
    wb.close()

    result = extract_excel(file_path)
    assert result.succeeded is True
    assert len(result.regions) == 2

    # Verify Sheet 1 (2 levels)
    r2 = result.regions[0]
    assert r2.sheet == "TwoLevel"
    assert len(r2.header_rows) == 2
    assert r2.headers[0] == "Financial Year 2026 > Q1 Revenue"
    assert r2.headers[1] == "Financial Year 2026 > Q2 Revenue"
    assert r2.headers[2] == "Operations > Staff Count"
    assert r2.headers[3] == "Operations > Shift Hours"

    # Verify Sheet 2 (3 levels)
    r3 = result.regions[1]
    assert r3.sheet == "ThreeLevel"
    assert len(r3.header_rows) == 3
    assert r3.headers[0] == "Asia Pacific > Vietnam > Hanoi Plant"
    assert r3.headers[1] == "Asia Pacific > Japan > Tokyo Plant"
    assert r3.headers[2] == "Europe > Germany > Munich Plant"
    assert r3.headers[3] == "Europe > France > Paris Plant"


def test_excel_merged_cells_spanning_chunk_boundaries(tmp_path):
    """Stress test merged cell ranges that cross chunk boundaries (e.g. rows 490 to 520)."""
    file_path = tmp_path / "merged_boundary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MergedBoundarySheet"

    ws.append(["Batch_ID", "Process_Step", "Value"])
    for r in range(1, 1001):
        ws.append([f"BATCH-{(r // 50) + 1}", f"Step-{r % 10}", r * 5])

    # Merge Batch_ID column across rows 490 to 520 (which straddles chunk boundary 500/501)
    ws.merge_cells("A490:A520")

    wb.save(file_path)
    wb.close()

    # Extract with chunk_row_size=500
    cfg = ExcelExtractionConfig(chunk_row_size=500)
    result = extract_excel(file_path, config=cfg)
    assert result.succeeded is True
    assert len(result.regions) == 2

    # Chunk 0: rows 1 to 501
    reg0 = result.regions[0]
    assert reg0.chunk_index == 0
    assert any("A490:A520" in m for m in reg0.merged_ranges), (
        f"Chunk 0 should track merge A490:A520, got: {reg0.merged_ranges}"
    )

    # Chunk 1: rows 502 to 1001
    reg1 = result.regions[1]
    assert reg1.chunk_index == 1
    assert any("A490:A520" in m for m in reg1.merged_ranges), (
        f"Chunk 1 should track merge A490:A520, got: {reg1.merged_ranges}"
    )


def test_excel_custom_chunk_row_size_variations(tmp_path):
    """Stress test various custom chunk sizes (50, 100, 250, 1200) and edge sizes."""
    file_path = tmp_path / "chunk_variations.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["ID", "Name", "Val"])
    for i in range(1, 601):
        ws.append([i, f"Name_{i}", i * 10])
    wb.save(file_path)
    wb.close()

    # Size = 50 -> 600 / 50 = 12 chunks
    r50 = extract_excel(file_path, config=ExcelExtractionConfig(chunk_row_size=50))
    assert len(r50.regions) == 12
    assert all(reg.total_chunks == 12 for reg in r50.regions)

    # Size = 250 -> ceil(600 / 250) = 3 chunks (250, 250, 100)
    r250 = extract_excel(file_path, config=ExcelExtractionConfig(chunk_row_size=250))
    assert len(r250.regions) == 3
    assert r250.regions[0].row_range == (1, 251)
    assert r250.regions[1].row_range == (252, 501)
    assert r250.regions[2].row_range == (502, 601)

    # Size = 1200 (> total rows) -> 1 chunk
    r1200 = extract_excel(file_path, config=ExcelExtractionConfig(chunk_row_size=1200))
    assert len(r1200.regions) == 1
    assert r1200.regions[0].total_chunks == 1
    assert r1200.regions[0].row_range == (1, 601)


def test_excel_no_data_loss_and_chunk_boundary_continuity(tmp_path):
    """Verify that every single cell and row from start to finish is present across chunks with zero data loss."""
    file_path = tmp_path / "continuity_check.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Continuous"
    ws.append(["Row_Index", "Payload"])
    total_data_rows = 1234
    for i in range(1, total_data_rows + 1):
        ws.append([i, f"payload_value_{i:06d}"])
    wb.save(file_path)
    wb.close()

    cfg = ExcelExtractionConfig(chunk_row_size=300)
    result = extract_excel(file_path, config=cfg)
    assert result.succeeded is True

    # Expected chunks: ceil(1234/300) = 5 chunks (300, 300, 300, 300, 34)
    assert len(result.regions) == 5

    extracted_rows = []
    for region in result.regions:
        # Skip header row (row index 0 in matrix)
        for data_row in region.rows[1:]:
            extracted_rows.append((int(data_row[0]), data_row[1]))

    assert len(extracted_rows) == total_data_rows
    for idx, (row_idx, payload) in enumerate(extracted_rows, 1):
        assert row_idx == idx
        assert payload == f"payload_value_{idx:06d}"
