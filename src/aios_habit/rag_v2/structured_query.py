"""Bounded, read-only SQLite analytics for structured Excel tables."""
from __future__ import annotations

import calendar
from dataclasses import dataclass
import datetime
import os
import re
import sqlite3
import unicodedata
from pathlib import Path
from typing import Iterable, Sequence

from .adapters import ConversionContext
from .converters import ExcelDocumentConverterAdapter
from .schema import DocumentElement, ElementType

_NUMBER_RE = re.compile(r"^-?\d+(?:[.,]\d+)?$")
_ALLOWED_FILTER_OPERATORS = frozenset({"=", "!=", ">", ">=", "<", "<=", "contains"})
_ALLOWED_AGGREGATES = frozenset({"count", "sum", "avg", "min", "max"})


class StructuredQueryError(RuntimeError):
    """Base class for deterministic structured-query failures."""


class StructuredQueryValidationError(StructuredQueryError):
    """Raised when an internal query plan fails the strict allow-list."""


class StructuredQueryBoundsError(StructuredQueryError):
    """Raised when workbook or execution resource limits are exceeded."""


@dataclass(frozen=True)
class StructuredFilter:
    column: str
    operator: str
    value: str | float | int


@dataclass(frozen=True)
class StructuredAggregate:
    function: str
    column: str = "*"
    alias: str = ""


@dataclass(frozen=True)
class StructuredOrder:
    column: str
    direction: str = "asc"


@dataclass(frozen=True)
class StructuredQueryPlan:
    sheet: str = ""
    target_regions: tuple[str, ...] = ()
    select_columns: tuple[str, ...] = ()
    filters: tuple[StructuredFilter, ...] = ()
    group_by: tuple[str, ...] = ()
    aggregates: tuple[StructuredAggregate, ...] = ()
    order_by: tuple[StructuredOrder, ...] = ()
    limit: int = 20


@dataclass(frozen=True)
class StructuredProvenance:
    document_id: str
    source_name: str
    sheet: str
    cell_range: str
    excel_rows: tuple[int, ...] = ()


@dataclass(frozen=True)
class StructuredQueryResult:
    applied: bool
    rows: tuple[tuple[object, ...], ...] = ()
    columns: tuple[str, ...] = ()
    rendered_evidence: str = ""
    document_id: str = ""
    source_name: str = ""
    sheet: str = ""
    cell_range: str = ""
    provenance: tuple[StructuredProvenance, ...] = ()
    row_count: int = 0
    truncated: bool = False
    reason: str = ""

    @property
    def rendered_text(self) -> str:
        """Backward-compatible alias for callers created during initial development."""
        return self.rendered_evidence


def _strip_accents(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).replace("đ", "d").replace("Đ", "D")


def _canonical_text(text: str) -> str:
    stripped = _strip_accents(str(text or "").casefold())
    return " ".join(re.sub(r"[^\w\s]", " ", stripped).split())


def _parse_iso_date(value: str) -> str | None:
    text = str(value or "").strip()
    m1 = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})(?:\s+.*)?$", text)
    if m1:
        y, m, d = int(m1.group(1)), int(m1.group(2)), int(m1.group(3))
        try:
            dt = datetime.date(y, m, d)
            return dt.isoformat()
        except ValueError:
            return None
    m2 = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})(?:\s+.*)?$", text)
    if m2:
        d, m, y = int(m2.group(1)), int(m2.group(2)), int(m2.group(3))
        try:
            dt = datetime.date(y, m, d)
            return dt.isoformat()
        except ValueError:
            return None
    return None


def _bound(name: str, default: int, minimum: int, maximum: int) -> int:
    try:
        return max(minimum, min(maximum, int(os.environ.get(name, str(default)))))
    except (TypeError, ValueError):
        return default


def _normalize_headers(headers: Sequence[str], width: int) -> tuple[str, ...]:
    result: list[str] = []
    used: dict[str, int] = {}
    for index in range(width):
        raw = str(headers[index] if index < len(headers) else "").strip()
        generated_blank = re.fullmatch(r"Column\s+[A-Z]+", raw)
        duplicate = re.fullmatch(r"(.+?)\s+\((\d+)\)", raw)
        if generated_blank:
            base = f"column_{index + 1}"
        elif duplicate:
            base = f"{duplicate.group(1).strip()}_{duplicate.group(2)}"
        else:
            base = raw or f"column_{index + 1}"
        used[base.casefold()] = used.get(base.casefold(), 0) + 1
        occurrence = used[base.casefold()]
        result.append(base if occurrence == 1 else f"{base}_{occurrence}")
    return tuple(result)


def _coerce(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    text = str(value).strip()
    if not text:
        return None
    iso_date = _parse_iso_date(text)
    if iso_date is not None:
        return iso_date
    compact = text.replace(" ", "")
    if _NUMBER_RE.fullmatch(compact):
        try:
            number = float(compact.replace(",", "."))
            return int(number) if number.is_integer() else number
        except ValueError:
            pass
    return text


def _resolve_column(name: str, columns: Sequence[str], aliases: Sequence[str] = ()) -> str:
    wanted_raw = str(name or "").strip().casefold()
    wanted_canon = _canonical_text(name)
    matches = [
        column for column in (*columns, *aliases)
        if column.casefold() == wanted_raw or _canonical_text(column) == wanted_canon
    ]
    if len(matches) != 1:
        raise StructuredQueryValidationError(f"Unknown column: {name}")
    return matches[0]


def _quote(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def _load_table(
    conn: sqlite3.Connection,
    table_name: str,
    element: DocumentElement,
    max_cells: int,
) -> tuple[tuple[str, ...], int]:
    assert element.table is not None
    rows = [list(row) for row in element.table.rows]
    width = max([len(element.table.headers), *(len(row) for row in rows)], default=0)
    headers = _normalize_headers(element.table.headers, width)
    if width == 0:
        return (), 0
    header_values = tuple(str(value).strip() for value in element.table.headers)
    first_row_raw = tuple(str(value).strip() for value in rows[0][:len(header_values)]) if rows else ()
    first_row_normalized = _normalize_headers(first_row_raw, width) if rows else ()
    data_rows = rows[1:] if rows and first_row_normalized == headers else rows
    cell_count = (len(data_rows) + 1) * width
    if cell_count > max_cells:
        raise StructuredQueryBoundsError("Workbook cell limit exceeded")
    definitions = ", ".join(f"{_quote(name)}" for name in headers)
    conn.execute(f"CREATE TABLE {_quote(table_name)} (_excel_row INTEGER, {definitions})")
    placeholders = ", ".join("?" for _ in range(width + 1))
    start_row = element.row_range[0] if element.row_range else 1
    first_data_row = start_row + (1 if data_rows is not rows else 0)
    values = [
        (
            first_data_row + row_index,
            *[_coerce(row[column] if column < len(row) else None) for column in range(width)],
        )
        for row_index, row in enumerate(data_rows)
    ]
    conn.executemany(f"INSERT INTO {_quote(table_name)} VALUES ({placeholders})", values)
    return headers, cell_count


def _compile_query(
    plan: StructuredQueryPlan,
    table_name: str,
    columns: Sequence[str],
    max_rows: int,
) -> tuple[str, tuple[object, ...], tuple[str, ...], bool, int]:
    aggregate_aliases: list[str] = []
    aggregate_parts: list[str] = []
    for position, aggregate in enumerate(plan.aggregates, start=1):
        function = aggregate.function.strip().casefold()
        if function not in _ALLOWED_AGGREGATES:
            raise StructuredQueryValidationError(f"Unsupported aggregate: {aggregate.function}")
        if aggregate.column == "*":
            if function != "count":
                raise StructuredQueryValidationError("Only COUNT may target *")
            expression = "*"
        else:
            expression = _quote(_resolve_column(aggregate.column, columns))
        alias = aggregate.alias.strip() or f"{function}_{position}"
        if any(alias.casefold() == value.casefold() for value in (*columns, *aggregate_aliases)):
            raise StructuredQueryValidationError(f"Duplicate aggregate alias: {alias}")
        aggregate_aliases.append(alias)
        aggregate_parts.append(f"{function.upper()}({expression}) AS {_quote(alias)}")

    group_columns = tuple(_resolve_column(name, columns) for name in plan.group_by)
    selected = tuple(_resolve_column(name, columns) for name in plan.select_columns)
    if group_columns and any(column not in group_columns for column in selected):
        raise StructuredQueryValidationError("Grouped query selections must appear in group_by")
    if plan.aggregates and selected and any(column not in group_columns for column in selected):
        raise StructuredQueryValidationError("Aggregate query selections must appear in group_by")
    if group_columns and not plan.aggregates:
        raise StructuredQueryValidationError("Grouped queries require an aggregate")

    if plan.aggregates:
        projections = [*(_quote(column) for column in (selected or group_columns)), *aggregate_parts]
        sheet_expression = "'' AS __excel_sheet"
        range_expression = "'' AS __excel_cell_range"
        provenance_expression = (
            "GROUP_CONCAT(_excel_sheet || char(31) || _excel_cell_range "
            "|| char(31) || _excel_row, char(30)) AS __excel_provenance"
        )
    else:
        projection_columns = selected or tuple(columns)
        projections = [_quote(column) for column in projection_columns]
        sheet_expression = "_excel_sheet AS __excel_sheet"
        range_expression = "_excel_cell_range AS __excel_cell_range"
        provenance_expression = "CAST(_excel_row AS TEXT) AS __excel_provenance"
    projections.extend([sheet_expression, range_expression, provenance_expression])

    params: list[object] = []
    where_parts: list[str] = []
    for item in plan.filters:
        column = _resolve_column(item.column, columns)
        operator = item.operator.strip().casefold()
        if operator not in _ALLOWED_FILTER_OPERATORS:
            raise StructuredQueryValidationError(f"Unsupported filter operator: {item.operator}")
        if operator == "contains":
            where_parts.append(f"CAST({_quote(column)} AS TEXT) LIKE ? ESCAPE '\\'")
            escaped = str(item.value).replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
            params.append(f"%{escaped}%")
        else:
            if isinstance(item.value, (int, float)) and not isinstance(item.value, bool):
                where_parts.append(
                    f"typeof({_quote(column)}) IN ('integer', 'real') AND "
                    f"{_quote(column)} {operator} ?"
                )
            else:
                where_parts.append(f"{_quote(column)} {operator} ?")
            params.append(_coerce(item.value))

    aliases = tuple(aggregate_aliases)
    order_parts: list[str] = []
    for item in plan.order_by:
        column = _resolve_column(item.column, columns, aliases)
        direction = item.direction.strip().casefold()
        if direction not in {"asc", "desc"}:
            raise StructuredQueryValidationError(f"Unsupported sort direction: {item.direction}")
        order_parts.append(f"{_quote(column)} {direction.upper()}")

    requested_limit = max(1, int(plan.limit))
    effective_limit = min(requested_limit, max_rows)
    sql = f"SELECT {', '.join(projections)} FROM {_quote(table_name)}"
    if where_parts:
        sql += f" WHERE {' AND '.join(where_parts)}"
    if group_columns:
        sql += " GROUP BY " + ", ".join(_quote(column) for column in group_columns)
    if order_parts:
        sql += " ORDER BY " + ", ".join(order_parts)
    elif not plan.aggregates:
        sql += " ORDER BY _excel_row ASC"
    sql += " LIMIT ?"
    params.append(effective_limit + 1)
    output_columns = tuple(selected or group_columns) + aliases if plan.aggregates else tuple(selected or columns)
    return sql, tuple(params), output_columns, False, effective_limit


def execute_excel_query(
    path: str | Path,
    plan: StructuredQueryPlan,
    *,
    document_id: str = "",
) -> StructuredQueryResult:
    """Execute an allow-listed plan against one bounded Excel region in memory."""
    workbook = Path(path)
    if workbook.suffix.lower() not in {".xlsx", ".xlsm", ".xls"} or not workbook.is_file():
        return StructuredQueryResult(False, reason="unsupported_or_missing_workbook")
    max_cells = _bound("AIOS_STRUCTURED_SQL_MAX_CELLS", 100_000, 1, 1_000_000)
    max_rows = _bound("AIOS_STRUCTURED_SQL_MAX_RESULT_ROWS", 50, 1, 500)
    max_steps = _bound("AIOS_STRUCTURED_SQL_MAX_PROGRESS_STEPS", 250_000, 1_000, 5_000_000)
    context = ConversionContext(document_id=document_id or None, fail_soft=True)
    elements = ExcelDocumentConverterAdapter().convert(str(workbook), context)
    tables = [item for item in elements if item.element_type == ElementType.TABLE and item.table]
    if plan.target_regions:
        target_set = set(plan.target_regions)
        tables = [
            item for item in tables
            if (item.sheet or "") in target_set
            or f"{(item.sheet or '')}!{(item.cell_range or '')}" in target_set
        ]
    elif plan.sheet:
        tables = [item for item in tables if (item.sheet or "").casefold() == plan.sheet.casefold()]

    if not tables:
        return StructuredQueryResult(False, reason="workbook_has_no_matching_structured_table")

    conn = sqlite3.connect(":memory:")
    progress_steps = 0
    def _progress() -> int:
        nonlocal progress_steps
        progress_steps += 1_000
        return 1 if progress_steps > max_steps else 0
    conn.set_progress_handler(_progress, 1_000)
    loaded_cells = 0
    try:
        first_cols: tuple[str, ...] = ()
        all_compatible_elements: list[DocumentElement] = []
        element_by_sheet: dict[str, DocumentElement] = {}

        for element in tables:
            assert element.table is not None
            width = max([len(element.table.headers), *(len(row) for row in element.table.rows)], default=0)
            cols = _normalize_headers(element.table.headers, width)
            if not cols:
                continue
            if not first_cols:
                first_cols = cols
                all_compatible_elements.append(element)
                if element.sheet:
                    element_by_sheet[element.sheet] = element
            elif cols == first_cols:
                all_compatible_elements.append(element)
                if element.sheet and element.sheet not in element_by_sheet:
                    element_by_sheet[element.sheet] = element
            elif plan.sheet:
                break

        if not all_compatible_elements:
            return StructuredQueryResult(False, reason="no_matching_table_or_rows")

        table_name = "merged_region"
        definitions = ", ".join(f"{_quote(name)}" for name in first_cols)
        conn.execute(f"CREATE TABLE {_quote(table_name)} (_excel_row INTEGER, _excel_sheet TEXT, _excel_cell_range TEXT, {definitions})")

        for element in all_compatible_elements:
            assert element.table is not None
            rows = [list(row) for row in element.table.rows]
            width = len(first_cols)
            header_values = tuple(str(value).strip() for value in element.table.headers)
            first_row_raw = tuple(str(value).strip() for value in rows[0][:len(header_values)]) if rows else ()
            first_row_normalized = _normalize_headers(first_row_raw, width) if rows else ()
            data_rows = rows[1:] if rows and first_row_normalized == first_cols else rows
            cell_count = (len(data_rows) + 1) * width
            loaded_cells += cell_count
            if loaded_cells > max_cells:
                raise StructuredQueryBoundsError("Workbook cell limit exceeded")

            start_row = element.row_range[0] if element.row_range else 1
            first_data_row = start_row + (1 if data_rows is not rows else 0)

            placeholders = ", ".join("?" for _ in range(width + 3))
            insert_values = []
            for row_index, row in enumerate(data_rows):
                excel_row_num = first_data_row + row_index
                row_vals = [_coerce(row[column] if column < len(row) else None) for column in range(width)]
                insert_values.append((excel_row_num, element.sheet or "", element.cell_range or "", *row_vals))

            conn.executemany(f"INSERT INTO {_quote(table_name)} VALUES ({placeholders})", insert_values)

        try:
            sql, params, output_columns, limit_clamped, effective_limit = _compile_query(
                plan, table_name, first_cols, max_rows
            )
        except StructuredQueryValidationError as exc:
            raise exc

        try:
            cursor = conn.execute(sql, params)
            raw_rows = cursor.fetchall()
        except sqlite3.OperationalError as exc:
            if "interrupted" in str(exc).casefold():
                raise StructuredQueryBoundsError("Structured query execution limit exceeded") from exc
            raise

        if not raw_rows:
            return StructuredQueryResult(False, reason="no_matching_table_or_rows")

        truncated = limit_clamped or len(raw_rows) > effective_limit
        raw_rows = raw_rows[:effective_limit]
        rows = tuple(tuple(row[:-3]) for row in raw_rows)

        first_element = all_compatible_elements[0]
        provenance_items: list[StructuredProvenance] = []
        for row in raw_rows:
            sheet_val = str(row[-3] or "").strip()
            range_val = str(row[-2] or "").strip()
            provenance_val = row[-1]
            if plan.aggregates:
                records = _parse_aggregate_provenance(provenance_val)
                grouped_records: dict[tuple[str, str], list[int]] = {}
                for sheet, cell_range, excel_row in records:
                    grouped_records.setdefault((sheet, cell_range), []).append(excel_row)
                for (sheet, cell_range), excel_rows in grouped_records.items():
                    matched_element = element_by_sheet.get(sheet, first_element)
                    provenance_items.append(
                        StructuredProvenance(
                            document_id=matched_element.document_id,
                            source_name=matched_element.source_name,
                            sheet=sheet,
                            cell_range=cell_range or matched_element.cell_range or "",
                            excel_rows=tuple(sorted(set(excel_rows))),
                        )
                    )
                continue

            matched_element = element_by_sheet.get(sheet_val, first_element)
            provenance_items.append(
                StructuredProvenance(
                    document_id=matched_element.document_id,
                    source_name=matched_element.source_name,
                    sheet=sheet_val or (first_element.sheet or ""),
                    cell_range=_row_cell_range(matched_element, provenance_val),
                    excel_rows=_parse_excel_rows(provenance_val),
                )
            )

        provenance = tuple(provenance_items)
        rendered = _render_result(output_columns, rows, first_element, truncated, provenance=provenance)

        return StructuredQueryResult(
            applied=True,
            rows=rows,
            columns=output_columns,
            rendered_evidence=rendered,
            document_id=first_element.document_id,
            source_name=first_element.source_name,
            sheet=first_element.sheet or "",
            cell_range=first_element.cell_range or "",
            provenance=provenance,
            row_count=len(rows),
            truncated=truncated,
        )

    except (StructuredQueryError, ValueError, OverflowError):
        raise
    except sqlite3.Error:
        return StructuredQueryResult(False, reason="structured_query_failed")
    finally:
        conn.close()


def _parse_aggregate_provenance(value: object) -> tuple[tuple[str, str, int], ...]:
    """Decode internal SQLite provenance records using Excel-illegal control separators."""
    records: list[tuple[str, str, int]] = []
    for record in str(value or "").split("\x1e"):
        if not record:
            continue
        parts = record.split("\x1f")
        if len(parts) != 3:
            continue
        sheet, cell_range, raw_row = parts
        try:
            records.append((sheet, cell_range, int(raw_row)))
        except ValueError:
            continue
    return tuple(records)


def _parse_excel_rows(value: object) -> tuple[int, ...]:
    result = []
    for part in str(value or "").split(","):
        try:
            result.append(int(part))
        except ValueError:
            continue
    return tuple(sorted(set(result)))


def _row_cell_range(element: DocumentElement, value: object) -> str:
    rows = _parse_excel_rows(value)
    if not rows:
        return element.cell_range or ""
    start_column, end_column = element.column_range or (1, 1)
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(start_column)}{min(rows)}:{get_column_letter(end_column)}{max(rows)}"


def _render_result(
    columns: Sequence[str],
    rows: Iterable[Sequence[object]],
    element: DocumentElement,
    truncated: bool,
    provenance: Sequence[StructuredProvenance] = (),
) -> str:
    sheets = tuple(dict.fromkeys(p.sheet for p in provenance if p.sheet))
    if len(sheets) > 1:
        location = f"multi-region ({', '.join(sheets)})"
    else:
        location = element.sheet or "workbook"
        if element.cell_range:
            location += f"!{element.cell_range}"
    lines = [f"Structured Excel result — {location}", " | ".join(columns)]
    for row in rows:
        lines.append(" | ".join("" if value is None else str(value) for value in row))
    if truncated:
        lines.append("[Result truncated by local safety limits]")
    max_chars = _bound("AIOS_STRUCTURED_SQL_MAX_OUTPUT_CHARS", 12_000, 500, 100_000)
    return "\n".join(lines)[:max_chars]


# The deterministic planner below consumes table metadata and produces only the
# allow-listed dataclasses above. It intentionally never accepts or emits SQL.
_SQL_TEXT_RE = re.compile(
    r"(?:^|[;\s])(?:select|insert|update|delete|drop|alter|create|pragma|attach|detach)\b",
    re.IGNORECASE,
)
_AGGREGATE_TERMS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("avg", ("trung bình", "bình quân", "average", "avg", "mean")),
    ("sum", ("tổng cộng", "tổng", "sum", "total")),
    ("count", ("bao nhiêu", "đếm", "count", "how many")),
    ("max", ("lớn nhất", "cao nhất", "tối đa", "maximum", "max")),
    ("min", ("nhỏ nhất", "thấp nhất", "tối thiểu", "minimum", "min")),
)
_LIST_TERMS = ("liệt kê", "hiển thị", "cho xem", "danh sách", "list", "show")
_TOP_TERMS = ("top", "cao nhất", "lớn nhất", "nhiều nhất", "highest", "largest")
_BOTTOM_TERMS = ("bottom", "thấp nhất", "nhỏ nhất", "ít nhất", "lowest", "smallest")
_GROUP_MARKERS = ("theo", "mỗi", "từng", " by ", "per ")

_SQL_LEXICON_MAP: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("revenue", ("doanh thu", "doanh so", "sales", "revenue", "turnover", "tien", "doanh_thu")),
    ("region", ("khu vuc", "mien", "vung", "region", "area", "chi nhanh")),
    ("date", ("ngay", "thoi gian", "thoi diem", "date", "created_at", "ngay tao", "ngay lap")),
    ("month", ("thang", "month")),
    ("year", ("nam", "year")),
    ("quantity", ("so luong", "san luong", "quantity", "qty", "amount")),
    ("product", ("san pham", "mat hang", "hang hoa", "product", "item", "ten hang")),
    ("cost", ("chi phi", "gia", "don gia", "price", "cost")),
    ("customer", ("khach hang", "khach", "customer", "client")),
    ("status", ("trang thai", "tinh trang", "status")),
)


@dataclass(frozen=True)
class ExcelTableSchema:
    sheet: str
    columns: tuple[str, ...]
    cell_range: str = ""
    region_index: int = 0


@dataclass(frozen=True)
class ExcelQueryPlanningResult:
    applied: bool
    plan: StructuredQueryPlan | None = None
    reason: str = ""


def inspect_excel_schemas(
    path: str | Path,
    *,
    document_id: str = "",
) -> tuple[ExcelTableSchema, ...]:
    """Read bounded table metadata through the same converter as execution."""
    workbook = Path(path)
    if workbook.suffix.lower() not in {".xlsx", ".xlsm", ".xls"} or not workbook.is_file():
        return ()
    elements = ExcelDocumentConverterAdapter().convert(
        str(workbook),
        ConversionContext(document_id=document_id or None, fail_soft=True),
    )
    schemas: list[ExcelTableSchema] = []
    for index, element in enumerate(elements):
        if element.element_type != ElementType.TABLE or not element.table:
            continue
        rows = element.table.rows
        width = max([len(element.table.headers), *(len(row) for row in rows)], default=0)
        columns = _normalize_headers(element.table.headers, width)
        if columns:
            schemas.append(
                ExcelTableSchema(
                    sheet=element.sheet or "",
                    columns=columns,
                    cell_range=element.cell_range or "",
                    region_index=index,
                )
            )
    return tuple(schemas)


def plan_excel_query(
    question: str,
    schemas: Sequence[ExcelTableSchema],
    *,
    default_limit: int = 20,
) -> ExcelQueryPlanningResult:
    """Create an internal plan only when structured intent is unambiguous."""
    normalized = _normalize_question(question)
    canon_q = _canonical_text(question)
    if not normalized:
        return ExcelQueryPlanningResult(False, reason="empty_question")
    if _SQL_TEXT_RE.search(question):
        return ExcelQueryPlanningResult(False, reason="raw_sql_rejected")
    if not schemas:
        return ExcelQueryPlanningResult(False, reason="no_structured_tables")

    scored_schemas: list[tuple[int, ExcelTableSchema, dict[str, tuple[int, int]]]] = []
    for schema in schemas:
        matches = _column_matches(question, schema.columns)
        score = sum(end - start for start, end in matches.values())
        sheet_canon = _canonical_text(schema.sheet)
        if sheet_canon and sheet_canon in canon_q:
            score += len(sheet_canon) + 30
        scored_schemas.append((score, schema, matches))

    scored_schemas.sort(key=lambda item: item[0], reverse=True)
    best_score, schema, matches = scored_schemas[0]

    if len(scored_schemas) > 1 and scored_schemas[1][0] == best_score and best_score > 0:
        s1, s2 = scored_schemas[0][1], scored_schemas[1][1]
        m1_sheet = _canonical_text(s1.sheet) in canon_q if s1.sheet else False
        m2_sheet = _canonical_text(s2.sheet) in canon_q if s2.sheet else False
        has_all_intent = _has_all_sheets_intent(canon_q)
        if not (m1_sheet or m2_sheet) and not has_all_intent:
            return ExcelQueryPlanningResult(False, reason="ambiguous_sheet_table")

    aggregate_function = _find_aggregate(normalized)
    list_intent = any(term in normalized for term in _LIST_TERMS)
    top_intent = any(term in normalized for term in _TOP_TERMS)
    bottom_intent = any(term in normalized for term in _BOTTOM_TERMS)

    try:
        filters = _extract_planned_filters(question, schema.columns, matches)
    except StructuredQueryValidationError as exc:
        return ExcelQueryPlanningResult(False, reason=f"filter_validation_error: {exc}")

    mentioned = tuple(
        column for column, _span in sorted(matches.items(), key=lambda item: item[1][0])
    )
    if not (aggregate_function or list_intent or top_intent or bottom_intent or filters):
        return ExcelQueryPlanningResult(False, reason="unstructured_intent")
    if not mentioned and aggregate_function != "count":
        return ExcelQueryPlanningResult(False, reason="structured_column_not_identified")

    group_by: tuple[str, ...] = ()
    target_column = mentioned[-1] if mentioned else "*"
    for column in mentioned:
        start = matches[column][0]
        prefix = normalized[max(0, start - 18):start]
        if any(marker in prefix for marker in _GROUP_MARKERS):
            group_by = (column,)
            break

    aggregates: tuple[StructuredAggregate, ...] = ()
    select_columns: tuple[str, ...] = mentioned
    if aggregate_function:
        aggregate_candidates = tuple(column for column in mentioned if column not in group_by)
        target_column = aggregate_candidates[-1] if aggregate_candidates else target_column
        aggregate_column = "*" if aggregate_function == "count" and not mentioned else target_column
        safe_column = re.sub(r"\W+", "_", _normalize_question(aggregate_column)).strip("_") or "rows"
        alias = f"{aggregate_function}_{safe_column}"
        aggregates = (StructuredAggregate(aggregate_function, aggregate_column, alias),)
        select_columns = group_by

    order_by: tuple[StructuredOrder, ...] = ()
    if top_intent or bottom_intent:
        order_column = aggregates[0].alias if aggregates else target_column
        order_by = (StructuredOrder(order_column, "desc" if top_intent else "asc"),)

    explicit_limit = re.search(r"\b(?:top|bottom|đầu|cuối)\s+(\d+)\b", normalized)
    limit = int(explicit_limit.group(1)) if explicit_limit else int(default_limit)
    limit = max(1, min(limit, 50))
    if (top_intent or bottom_intent) and explicit_limit is None:
        limit = min(limit, 10)

    target_regions: list[str] = []
    has_all_intent = _has_all_sheets_intent(canon_q)
    if not schema.sheet or has_all_intent:
        matching_schemas = [
            s for _sc, s, _m in scored_schemas if s.columns == schema.columns
        ]
        if len(matching_schemas) > 1:
            target_regions = [
                f"{s.sheet}!{s.cell_range}" if s.cell_range else s.sheet
                for s in matching_schemas
            ]

    return ExcelQueryPlanningResult(
        True,
        plan=StructuredQueryPlan(
            sheet=schema.sheet,
            target_regions=tuple(target_regions),
            select_columns=select_columns,
            filters=filters,
            group_by=group_by,
            aggregates=aggregates,
            order_by=order_by,
            limit=limit,
        ),
        reason="structured_plan_created",
    )


def _has_all_sheets_intent(canonical_question: str) -> bool:
    return bool(re.search(r"(?:^|\s)(?:tat ca|all)(?:\s|$)", canonical_question))


def _normalize_question(value: str) -> str:
    return " ".join(str(value or "").casefold().split())


def _column_matches(text: str, columns: Sequence[str]) -> dict[str, tuple[int, int]]:
    matches: dict[str, tuple[int, int]] = {}
    occupied: list[tuple[int, int]] = []
    text_canon = _canonical_text(text)
    text_norm = _normalize_question(text)

    for column in sorted(columns, key=lambda item: len(_normalize_question(item)), reverse=True):
        candidate = _normalize_question(column.replace("_", " "))
        candidate_canon = _canonical_text(column.replace("_", " "))
        if not candidate and not candidate_canon:
            continue

        match = re.search(rf"(?<!\w){re.escape(candidate)}(?!\w)", text_norm, re.UNICODE)
        if match is None and candidate_canon:
            match = re.search(rf"(?<!\w){re.escape(candidate_canon)}(?!\w)", text_canon, re.UNICODE)

        if match is None:
            continue
        span = match.span()
        if any(span[0] < end and start < span[1] for start, end in occupied):
            continue
        matches[column] = span
        occupied.append(span)

    for column in columns:
        if column in matches:
            continue
        col_canon = _canonical_text(column.replace("_", " "))
        group_terms: tuple[str, ...] = ()
        for _group_key, terms in _SQL_LEXICON_MAP:
            if col_canon in terms or any(term in col_canon for term in terms):
                group_terms = terms
                break
        if not group_terms:
            continue
        for term in group_terms:
            match = re.search(rf"(?<!\w){re.escape(term)}(?!\w)", text_canon, re.UNICODE)
            if match:
                span = match.span()
                if not any(span[0] < end and start < span[1] for start, end in occupied):
                    matches[column] = span
                    occupied.append(span)
                    break

    return matches


def _find_aggregate(text: str) -> str:
    canon = _canonical_text(text)
    for function, terms in _AGGREGATE_TERMS:
        if any(_canonical_text(term) in canon for term in terms):
            return function
    return ""


def _extract_date_filters(text: str, date_columns: Sequence[str]) -> list[StructuredFilter]:
    filters: list[StructuredFilter] = []
    if not date_columns:
        return filters
    target_col = date_columns[0]

    m_month = re.search(r"tháng\s+(\d{1,2})(?:\s*(?:năm|/)\s*(\d{4}))?", text, re.IGNORECASE)
    if m_month:
        month = int(m_month.group(1))
        year = int(m_month.group(2)) if m_month.group(2) else 2026
        if 1 <= month <= 12 and 1900 <= year <= 2100:
            last_day = calendar.monthrange(year, month)[1]
            start_str = f"{year:04d}-{month:02d}-01"
            end_str = f"{year:04d}-{month:02d}-{last_day:02d}"
            filters.append(StructuredFilter(target_col, ">=", start_str))
            filters.append(StructuredFilter(target_col, "<=", end_str))
            return filters

    m_year = re.search(r"năm\s+(\d{4})", text, re.IGNORECASE)
    if m_year:
        year = int(m_year.group(1))
        if 1900 <= year <= 2100:
            filters.append(StructuredFilter(target_col, ">=", f"{year:04d}-01-01"))
            filters.append(StructuredFilter(target_col, "<=", f"{year:04d}-12-31"))
            return filters

    m_range = re.search(
        r"(?:từ|between)\s+(\d{1,4}[-/]\d{1,2}[-/]\d{1,4})\s+(?:đến|năm|to|and)\s+(\d{1,4}[-/]\d{1,2}[-/]\d{1,4})",
        text, re.IGNORECASE
    )
    if m_range:
        d1 = _parse_iso_date(m_range.group(1))
        d2 = _parse_iso_date(m_range.group(2))
        if d1 and d2:
            filters.append(StructuredFilter(target_col, ">=", d1))
            filters.append(StructuredFilter(target_col, "<=", d2))
            return filters

    m_before = re.search(r"(?:trước|before)\s+(\d{1,4}[-/]\d{1,2}[-/]\d{1,4})", text, re.IGNORECASE)
    if m_before:
        d = _parse_iso_date(m_before.group(1))
        if d:
            filters.append(StructuredFilter(target_col, "<=", d))
            return filters

    m_after = re.search(r"(?:sau|after)\s+(\d{1,4}[-/]\d{1,2}[-/]\d{1,4})", text, re.IGNORECASE)
    if m_after:
        d = _parse_iso_date(m_after.group(1))
        if d:
            filters.append(StructuredFilter(target_col, ">=", d))
            return filters

    m_on = re.search(r"(?:ngày|on)\s+(\d{1,4}[-/]\d{1,2}[-/]\d{1,4})", text, re.IGNORECASE)
    if m_on:
        d = _parse_iso_date(m_on.group(1))
        if d:
            filters.append(StructuredFilter(target_col, "=", d))
            return filters

    return filters


def _extract_planned_filters(
    text: str,
    columns: Sequence[str],
    matches: dict[str, tuple[int, int]],
) -> tuple[StructuredFilter, ...]:
    canon_text = _canonical_text(text)
    if re.search(r"\b(hoac|or)\b", canon_text):
        raise StructuredQueryValidationError("OR filter logic is unsupported")

    result: list[StructuredFilter] = []
    symbolic = re.compile(r"^\s*(>=|<=|!=|=|>|<)\s*(-?\d+(?:[.,]\d+)?)")
    word_operators = (
        ("lớn hơn hoặc bằng", ">="), ("greater than or equal", ">="),
        ("nhỏ hơn hoặc bằng", "<="), ("less than or equal", "<="),
        ("lớn hơn", ">"), ("greater than", ">"), ("trên", ">"),
        ("nhỏ hơn", "<"), ("less than", "<"), ("dưới", "<"),
        ("khác", "!="), ("not equal", "!="), ("bằng", "="), ("is", "="), ("là", "="),
    )

    date_cols = [
        col for col in columns
        if any(term in _canonical_text(col) for term in ("date", "ngay", "thoi gian", "created_at"))
    ]
    date_filters = _extract_date_filters(text, date_cols)
    result.extend(date_filters)

    for column in columns:
        if column not in matches:
            continue
        tail = text[matches[column][1]:matches[column][1] + 80]
        match = symbolic.match(tail)
        if match:
            result.append(StructuredFilter(column, match.group(1), _planned_number(match.group(2))))
            continue
        matched_word = False
        for term, operator in word_operators:
            word_match = re.match(rf"\s*{re.escape(term)}\s+(-?\d+(?:[.,]\d+)?)", tail, re.IGNORECASE)
            if word_match:
                result.append(StructuredFilter(column, operator, _planned_number(word_match.group(1))))
                matched_word = True
                break
            str_match = re.match(rf"\s*{re.escape(term)}\s+[\"']?([\w\-\s]+?)[\"']?(?:\s+và|\s+and|\s*$|\s*,)", tail, re.IGNORECASE)
            if str_match and operator == "=":
                val = str_match.group(1).strip()
                if val and not _NUMBER_RE.fullmatch(val.replace(" ", "")):
                    result.append(StructuredFilter(column, "=", val))
                    matched_word = True
                    break
        if matched_word:
            continue
        contains = re.match(r"\s*(?:chứa|contains)\s+[\"']?([^\"',;?]+)", tail, re.IGNORECASE)
        if contains:
            value = contains.group(1).strip()
            if value:
                result.append(StructuredFilter(column, "contains", value))
    return tuple(result)


def _planned_number(value: str) -> int | float:
    number = float(value.replace(",", "."))
    return int(number) if number.is_integer() else number
