"""
Generic converter adapters for RAG v2.
"""
import os
import hashlib
import datetime
import zipfile
import xml.etree.ElementTree as ET
from html.parser import HTMLParser
from typing import List, Optional, Dict, Any, Tuple

from .schema import DocumentElement, ExtractionStatus, ElementType, TableData, TableCell
from .adapters import DocumentConverterAdapter, BaseDocumentConverterAdapter, ConversionContext, AdapterCapabilities

class TextDocumentConverterAdapter(BaseDocumentConverterAdapter):
    def supports(self, path: str, file_type: Optional[str] = None, mime: Optional[str] = None) -> bool:
        ext = os.path.splitext(path)[1].lower()
        return ext in {".txt", ".md", ".csv"}

    def convert(self, path: str, context: ConversionContext) -> List[DocumentElement]:
        if not self.supports(path):
            if context.fail_soft:
                return [self._create_failed_element(path, f"Unsupported file: {path}", context, "TextDocumentConverterAdapter")]
            raise ValueError(f"Unsupported file: {path}")

        if not os.path.exists(path):
            if context.fail_soft:
                return [self._create_failed_element(path, f"File not found: {path}", context, "TextDocumentConverterAdapter")]
            raise FileNotFoundError(f"File not found: {path}")

        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()
        except Exception as e:
            if context.fail_soft:
                return [self._create_failed_element(path, f"Read failed: {e}", context, "TextDocumentConverterAdapter")]
            raise e

        # Element-first: split by paragraphs
        paragraphs = [p.strip() for p in content.split("\n\n") if p.strip()]
        if not paragraphs:
            paragraphs = [content.strip()] if content.strip() else [""]

        elements = []
        path_hash = hashlib.sha256(path.encode("utf-8")).hexdigest()[:12]
        doc_id = context.document_id or f"doc_{path_hash}"

        for idx, para in enumerate(paragraphs):
            element_id = f"{doc_id}_txt_{path_hash}_{idx}"
            is_heading = para.startswith("#")
            elem_type = ElementType.HEADING if is_heading else ElementType.TEXT
            
            elements.append(DocumentElement(
                element_id=element_id,
                document_id=doc_id,
                source_path=path,
                source_name=os.path.basename(path),
                file_type=os.path.splitext(path)[1].lower()[1:],
                extractor="TextDocumentConverterAdapter",
                extraction_status=ExtractionStatus.SUCCESS,
                element_type=elem_type,
                text=para,
                privacy_labels=context.privacy_labels,
                source_fingerprint=context.source_fingerprint,
                created_at=datetime.datetime.now(datetime.timezone.utc).isoformat()
            ))
        return elements

    def capabilities(self) -> AdapterCapabilities:
        return AdapterCapabilities(
            adapter_name="TextDocumentConverterAdapter",
            supported_file_types=[".txt", ".md", ".csv"],
            supports_metadata=True
        )


class HTMLReadableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.parts: List[str] = []
        self.skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() in {"script", "style", "noscript"}:
            self.skip_depth += 1
        if tag.lower() in {"p", "div", "br", "li", "tr", "h1", "h2", "h3", "h4", "h5", "h6"}:
            self.parts.append("\n\n")

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"script", "style", "noscript"} and self.skip_depth:
            self.skip_depth -= 1

    def handle_data(self, data: str) -> None:
        if not self.skip_depth and data.strip():
            self.parts.append(data)

    def get_text(self) -> str:
        return "".join(self.parts)


class HTMLDocumentConverterAdapter(BaseDocumentConverterAdapter):
    def supports(self, path: str, file_type: Optional[str] = None, mime: Optional[str] = None) -> bool:
        ext = os.path.splitext(path)[1].lower()
        return ext in {".html", ".htm"}

    def convert(self, path: str, context: ConversionContext) -> List[DocumentElement]:
        if not self.supports(path):
            if context.fail_soft:
                return [self._create_failed_element(path, f"Unsupported file: {path}", context, "HTMLDocumentConverterAdapter")]
            raise ValueError(f"Unsupported file: {path}")

        if not os.path.exists(path):
            if context.fail_soft:
                return [self._create_failed_element(path, f"File not found: {path}", context, "HTMLDocumentConverterAdapter")]
            raise FileNotFoundError(f"File not found: {path}")

        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                html_content = f.read()
        except Exception as e:
            if context.fail_soft:
                return [self._create_failed_element(path, f"Read failed: {e}", context, "HTMLDocumentConverterAdapter")]
            raise e

        parser = HTMLReadableParser()
        try:
            parser.feed(html_content)
            parser.close()
            text = parser.get_text()
        except Exception as e:
            if context.fail_soft:
                return [self._create_failed_element(path, f"HTML parsing failed: {e}", context, "HTMLDocumentConverterAdapter")]
            raise e

        paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
        if not paragraphs:
            paragraphs = [""]

        elements = []
        path_hash = hashlib.sha256(path.encode("utf-8")).hexdigest()[:12]
        doc_id = context.document_id or f"doc_{path_hash}"

        for idx, para in enumerate(paragraphs):
            element_id = f"{doc_id}_html_{path_hash}_{idx}"
            elements.append(DocumentElement(
                element_id=element_id,
                document_id=doc_id,
                source_path=path,
                source_name=os.path.basename(path),
                file_type=os.path.splitext(path)[1].lower()[1:],
                extractor="HTMLDocumentConverterAdapter",
                extraction_status=ExtractionStatus.SUCCESS,
                element_type=ElementType.TEXT,
                text=para,
                privacy_labels=context.privacy_labels,
                source_fingerprint=context.source_fingerprint,
                created_at=datetime.datetime.now(datetime.timezone.utc).isoformat()
            ))
        return elements

    def capabilities(self) -> AdapterCapabilities:
        return AdapterCapabilities(
            adapter_name="HTMLDocumentConverterAdapter",
            supported_file_types=[".html", ".htm"],
            supports_metadata=True
        )


class PDFDocumentConverterAdapter(BaseDocumentConverterAdapter):
    def supports(self, path: str, file_type: Optional[str] = None, mime: Optional[str] = None) -> bool:
        ext = os.path.splitext(path)[1].lower()
        return ext == ".pdf"

    def convert(self, path: str, context: ConversionContext) -> List[DocumentElement]:
        if not self.supports(path):
            if context.fail_soft:
                return [self._create_failed_element(path, f"Unsupported file: {path}", context, "PDFDocumentConverterAdapter")]
            raise ValueError(f"Unsupported file: {path}")

        if not os.path.exists(path):
            if context.fail_soft:
                return [self._create_failed_element(path, f"File not found: {path}", context, "PDFDocumentConverterAdapter")]
            raise FileNotFoundError(f"File not found: {path}")

        # Check dependency
        try:
            import fitz
            has_fitz = True
        except ImportError:
            has_fitz = False

        if not has_fitz:
            msg = "Missing dependency: PyMuPDF (fitz)"
            if context.fail_soft:
                return [self._create_failed_element(path, msg, context, "PDFDocumentConverterAdapter")]
            raise ImportError(msg)

        elements = []
        path_hash = hashlib.sha256(path.encode("utf-8")).hexdigest()[:12]
        doc_id = context.document_id or f"doc_{path_hash}"

        try:
            doc = fitz.open(str(path))
            for page_idx, page in enumerate(doc, start=1):
                try:
                    text = page.get_text("text") or ""
                except Exception:
                    text = ""
                
                paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
                if not paragraphs:
                    paragraphs = [""]

                for idx, para in enumerate(paragraphs):
                    element_id = f"{doc_id}_pdf_{path_hash}_p{page_idx}_{idx}"
                    elements.append(DocumentElement(
                        element_id=element_id,
                        document_id=doc_id,
                        source_path=path,
                        source_name=os.path.basename(path),
                        file_type="pdf",
                        extractor="PDFDocumentConverterAdapter",
                        extraction_status=ExtractionStatus.SUCCESS,
                        element_type=ElementType.TEXT,
                        text=para,
                        page=page_idx,
                        privacy_labels=context.privacy_labels,
                        source_fingerprint=context.source_fingerprint,
                        created_at=datetime.datetime.now(datetime.timezone.utc).isoformat()
                    ))
            doc.close()
        except Exception as e:
            if context.fail_soft:
                return [self._create_failed_element(path, f"PDF read failed: {e}", context, "PDFDocumentConverterAdapter")]
            raise e

        return elements

    def capabilities(self) -> AdapterCapabilities:
        try:
            import fitz
            has_fitz = True
        except ImportError:
            has_fitz = False
        return AdapterCapabilities(
            adapter_name="PDFDocumentConverterAdapter",
            supported_file_types=[".pdf"],
            supports_layout=True,
            supports_metadata=True,
            requires_external_dependency=True,
            dependency_status="ok" if has_fitz else "missing"
        )


class ExcelDocumentConverterAdapter(BaseDocumentConverterAdapter):
    def supports(self, path: str, file_type: Optional[str] = None, mime: Optional[str] = None) -> bool:
        ext = os.path.splitext(path)[1].lower()
        return ext in {".xlsx", ".xlsm"}

    def convert(self, path: str, context: ConversionContext) -> List[DocumentElement]:
        if not self.supports(path):
            if context.fail_soft:
                return [self._create_failed_element(path, f"Unsupported file: {path}", context, "ExcelDocumentConverterAdapter")]
            raise ValueError(f"Unsupported file: {path}")

        if not os.path.exists(path):
            if context.fail_soft:
                return [self._create_failed_element(path, f"File not found: {path}", context, "ExcelDocumentConverterAdapter")]
            raise FileNotFoundError(f"File not found: {path}")

        # Check dependency
        try:
            import openpyxl
            has_openpyxl = True
        except ImportError:
            has_openpyxl = False

        if not has_openpyxl:
            msg = "Missing dependency: openpyxl"
            if context.fail_soft:
                return [self._create_failed_element(path, msg, context, "ExcelDocumentConverterAdapter")]
            raise ImportError(msg)

        elements = []
        path_hash = hashlib.sha256(path.encode("utf-8")).hexdigest()[:12]
        doc_id = context.document_id or f"doc_{path_hash}"

        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
            for sheet_idx, sheet_name in enumerate(workbook.sheetnames[:10]):
                sheet = workbook[sheet_name]
                
                rows_data: List[List[str]] = []
                cells_data: List[TableCell] = []
                headers: List[str] = []

                max_r = min(sheet.max_row or 100, 100)
                max_c = min(sheet.max_column or 20, 20)

                for r_idx, row in enumerate(sheet.iter_rows(max_row=max_r, max_col=max_c, values_only=True), start=1):
                    row_vals = [str(val) if val is not None else "" for val in row]
                    if any(val.strip() for val in row_vals):
                        rows_data.append(row_vals)
                        for c_idx, val in enumerate(row_vals, start=1):
                            if val.strip():
                                is_h = (r_idx == 1)
                                cells_data.append(TableCell(
                                    row_index=r_idx - 1,
                                    column_index=c_idx - 1,
                                    text=val,
                                    is_header=is_h
                                ))

                if rows_data:
                    headers = rows_data[0] if len(rows_data) > 0 else []
                    table_data = TableData(
                        headers=headers,
                        rows=rows_data,
                        cells=cells_data
                    )
                    
                    element_id = f"{doc_id}_xls_{path_hash}_{sheet_idx}"
                    elements.append(DocumentElement(
                        element_id=element_id,
                        document_id=doc_id,
                        source_path=path,
                        source_name=os.path.basename(path),
                        file_type=os.path.splitext(path)[1].lower()[1:],
                        extractor="ExcelDocumentConverterAdapter",
                        extraction_status=ExtractionStatus.SUCCESS,
                        element_type=ElementType.TABLE,
                        table=table_data,
                        sheet=sheet_name,
                        row_range=(1, len(rows_data)),
                        column_range=(1, len(headers)),
                        privacy_labels=context.privacy_labels,
                        source_fingerprint=context.source_fingerprint,
                        created_at=datetime.datetime.now(datetime.timezone.utc).isoformat()
                    ))
            workbook.close()
        except Exception as e:
            if context.fail_soft:
                return [self._create_failed_element(path, f"Excel read failed: {e}", context, "ExcelDocumentConverterAdapter")]
            raise e

        return elements

    def capabilities(self) -> AdapterCapabilities:
        try:
            import openpyxl
            has_openpyxl = True
        except ImportError:
            has_openpyxl = False
        return AdapterCapabilities(
            adapter_name="ExcelDocumentConverterAdapter",
            supported_file_types=[".xlsx", ".xlsm"],
            supports_tables=True,
            supports_metadata=True,
            requires_external_dependency=True,
            dependency_status="ok" if has_openpyxl else "missing"
        )


class WordDocumentConverterAdapter(BaseDocumentConverterAdapter):
    def supports(self, path: str, file_type: Optional[str] = None, mime: Optional[str] = None) -> bool:
        ext = os.path.splitext(path)[1].lower()
        return ext == ".docx"

    def convert(self, path: str, context: ConversionContext) -> List[DocumentElement]:
        if not self.supports(path):
            if context.fail_soft:
                return [self._create_failed_element(path, f"Unsupported file: {path}", context, "WordDocumentConverterAdapter")]
            raise ValueError(f"Unsupported file: {path}")

        if not os.path.exists(path):
            if context.fail_soft:
                return [self._create_failed_element(path, f"File not found: {path}", context, "WordDocumentConverterAdapter")]
            raise FileNotFoundError(f"File not found: {path}")

        elements = []
        path_hash = hashlib.sha256(path.encode("utf-8")).hexdigest()[:12]
        doc_id = context.document_id or f"doc_{path_hash}"

        try:
            with zipfile.ZipFile(path, "r") as archive:
                if "word/document.xml" not in archive.namelist():
                    if context.fail_soft:
                        return [self._create_failed_element(path, "Missing word/document.xml", context, "WordDocumentConverterAdapter")]
                    raise ValueError("Invalid docx zip container")

                content_xml = archive.read("word/document.xml")
                root = ET.fromstring(content_xml)
                
                paragraphs = []
                for p in root.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p'):
                    texts = []
                    for t in p.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t'):
                        if t.text:
                            texts.append(t.text)
                    para_text = "".join(texts).strip()
                    if para_text:
                        paragraphs.append(para_text)

                if not paragraphs:
                    paragraphs = [""]

                for idx, para in enumerate(paragraphs):
                    element_id = f"{doc_id}_docx_{path_hash}_{idx}"
                    elements.append(DocumentElement(
                        element_id=element_id,
                        document_id=doc_id,
                        source_path=path,
                        source_name=os.path.basename(path),
                        file_type="docx",
                        extractor="WordDocumentConverterAdapter",
                        extraction_status=ExtractionStatus.SUCCESS,
                        element_type=ElementType.TEXT,
                        text=para,
                        privacy_labels=context.privacy_labels,
                        source_fingerprint=context.source_fingerprint,
                        created_at=datetime.datetime.now(datetime.timezone.utc).isoformat()
                    ))
        except Exception as e:
            if context.fail_soft:
                return [self._create_failed_element(path, f"Word read failed: {e}", context, "WordDocumentConverterAdapter")]
            raise e

        return elements

    def capabilities(self) -> AdapterCapabilities:
        return AdapterCapabilities(
            adapter_name="WordDocumentConverterAdapter",
            supported_file_types=[".docx"],
            supports_metadata=True
        )


class PowerPointDocumentConverterAdapter(BaseDocumentConverterAdapter):
    def supports(self, path: str, file_type: Optional[str] = None, mime: Optional[str] = None) -> bool:
        ext = os.path.splitext(path)[1].lower()
        return ext == ".pptx"

    def convert(self, path: str, context: ConversionContext) -> List[DocumentElement]:
        if not self.supports(path):
            if context.fail_soft:
                return [self._create_failed_element(path, f"Unsupported file: {path}", context, "PowerPointDocumentConverterAdapter")]
            raise ValueError(f"Unsupported file: {path}")

        if not os.path.exists(path):
            if context.fail_soft:
                return [self._create_failed_element(path, f"File not found: {path}", context, "PowerPointDocumentConverterAdapter")]
            raise FileNotFoundError(f"File not found: {path}")

        elements = []
        path_hash = hashlib.sha256(path.encode("utf-8")).hexdigest()[:12]
        doc_id = context.document_id or f"doc_{path_hash}"

        try:
            with zipfile.ZipFile(path, "r") as archive:
                slide_names = sorted(
                    [n for n in archive.namelist() if n.lower().startswith("ppt/slides/slide") and n.lower().endswith(".xml")],
                    key=lambda x: int(''.join(c for c in x if c.isdigit()))
                )

                if not slide_names:
                    if context.fail_soft:
                        return [self._create_failed_element(path, "No slide XMLs found", context, "PowerPointDocumentConverterAdapter")]
                    raise ValueError("Invalid pptx container or no slides found")

                for slide_idx, slide_name in enumerate(slide_names, start=1):
                    content_xml = archive.read(slide_name)
                    root = ET.fromstring(content_xml)
                    
                    texts = []
                    for t in root.iter('{http://schemas.openxmlformats.org/drawingml/2006/main}t'):
                        if t.text:
                            texts.append(t.text)
                    
                    slide_text = " ".join(texts).strip()
                    paragraphs = [p.strip() for p in slide_text.split("\n\n") if p.strip()]
                    if not paragraphs:
                        paragraphs = [""]

                    for idx, para in enumerate(paragraphs):
                        element_id = f"{doc_id}_pptx_{path_hash}_s{slide_idx}_{idx}"
                        elements.append(DocumentElement(
                            element_id=element_id,
                            document_id=doc_id,
                            source_path=path,
                            source_name=os.path.basename(path),
                            file_type="pptx",
                            extractor="PowerPointDocumentConverterAdapter",
                            extraction_status=ExtractionStatus.SUCCESS,
                            element_type=ElementType.TEXT,
                            text=para,
                            slide=slide_idx,
                            privacy_labels=context.privacy_labels,
                            source_fingerprint=context.source_fingerprint,
                            created_at=datetime.datetime.now(datetime.timezone.utc).isoformat()
                        ))
        except Exception as e:
            if context.fail_soft:
                return [self._create_failed_element(path, f"PowerPoint read failed: {e}", context, "PowerPointDocumentConverterAdapter")]
            raise e

        return elements

    def capabilities(self) -> AdapterCapabilities:
        return AdapterCapabilities(
            adapter_name="PowerPointDocumentConverterAdapter",
            supported_file_types=[".pptx"],
            supports_metadata=True
        )
