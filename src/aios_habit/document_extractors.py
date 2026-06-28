from __future__ import annotations

import html
import os
import re
import shutil
import subprocess
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

MAX_CHUNKS_PER_FILE = 30
MAX_OCR_IMAGE_BYTES = 8 * 1024 * 1024
MAX_PDF_OCR_PAGES = 3
OCR_TIMEOUT_SECONDS = 20
DEFAULT_OCR_LANG = "eng"
COMMON_TESSERACT_PATHS = (
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    r"D:\Tools\Tesseract-OCR\tesseract.exe",
    r"D:\Sandbox\tools\Tesseract-OCR\tesseract.exe",
)
USABLE_STATUSES = {"success", "extracted_success", "extracted_partial", "ocr_success", "ocr_partial", "extracted"}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}


@dataclass
class ExtractionResult:
    text: str
    file_type: str
    extractor_name: str
    extraction_status: str
    warning: str = ""
    section: str = ""
    page: str = ""
    slide: str = ""
    sheet: str = ""
    row_range: str = ""
    ocr_engine: str = ""
    ocr_lang: str = ""
    element_type: str = ""


class _ReadableHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self._parts: list[str] = []
        self._skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() in {"script", "style", "noscript"}:
            self._skip_depth += 1
        if tag.lower() in {"p", "div", "br", "li", "tr", "h1", "h2", "h3", "h4", "h5", "h6"}:
            self._parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"script", "style", "noscript"} and self._skip_depth:
            self._skip_depth -= 1
        if tag.lower() in {"p", "div", "li", "tr", "h1", "h2", "h3", "h4", "h5", "h6"}:
            self._parts.append("\n")

    def handle_data(self, data: str) -> None:
        if not self._skip_depth and data.strip():
            self._parts.append(data)

    def text(self) -> str:
        return "\n".join(_clean_lines("".join(self._parts).splitlines()))


def _clean_lines(lines: list[str], *, limit: int = 200) -> list[str]:
    cleaned: list[str] = []
    seen: set[str] = set()
    for line in lines:
        text = html.unescape(str(line or ""))
        text = re.sub(r"<[^>\n]{1,160}>", " ", text)
        text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]+", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(text)
        if len(cleaned) >= limit:
            break
    return cleaned


def _has_meaningful_text(text: str) -> bool:
    tokens = re.findall(r"[A-Za-z0-9À-ỹ]{2,}", text or "")
    return len(tokens) >= 2


def _xml_local_name(tag: str) -> str:
    return str(tag or "").rsplit("}", 1)[-1]


def _xml_root_from_zip(archive: zipfile.ZipFile, name: str):
    try:
        return ET.fromstring(archive.read(name))
    except (KeyError, ET.ParseError):
        return None


def _extract_xml_text(xml_text: str) -> list[str]:
    values = re.findall(r"<[^>]*t[^>]*>(.*?)</[^>]*t>", str(xml_text or ""), flags=re.IGNORECASE | re.DOTALL)
    return _clean_lines([html.unescape(value) for value in values], limit=200)


def _text_nodes(element) -> str:
    parts: list[str] = []
    for node in element.iter():
        if _xml_local_name(node.tag) == "t" and node.text:
            parts.append(node.text)
    return " ".join(_clean_lines(parts, limit=200)).strip()


def _chunk_result(result: ExtractionResult, path: Path, root: Path | None, max_chars_per_chunk: int) -> list[dict[str, Any]]:
    text = result.text.strip()
    rel = path.relative_to(root).as_posix() if root else path.name
    base = {
        "source_file": path.name,
        "relative_path": rel,
        "file_type": result.file_type,
        "section": result.section,
        "page": result.page,
        "slide": result.slide,
        "sheet": result.sheet,
        "row_range": result.row_range,
        "privacy_level": "local_only",
        "extractor_name": result.extractor_name,
        "extraction_status": result.extraction_status,
        "warning": result.warning,
        "ocr_engine": result.ocr_engine,
        "ocr_lang": result.ocr_lang,
        "element_type": result.element_type,
    }
    if result.extraction_status not in USABLE_STATUSES or not text:
        return [{"text": "", **base}]
    chunks: list[dict[str, Any]] = []
    for idx, start in enumerate(range(0, len(text), max_chars_per_chunk)):
        if idx >= MAX_CHUNKS_PER_FILE:
            break
        part = text[start:start + max_chars_per_chunk]
        section = result.section or f"chars {start}-{start + len(part)}"
        chunks.append({"text": part, **{**base, "section": section}})
    return chunks


def _extract_html(path: Path) -> ExtractionResult:
    parser = _ReadableHTMLParser()
    parser.feed(path.read_text(encoding="utf-8-sig", errors="ignore"))
    parser.close()
    text = parser.text()
    if not _has_meaningful_text(text):
        return ExtractionResult("", path.suffix.lower(), "html_parser", "failed_with_reason", "html has no visible text")
    return ExtractionResult(text, path.suffix.lower(), "html_parser", "extracted_success", section="visible text")


def _extract_pptx(path: Path) -> ExtractionResult:
    slide_lines: list[str] = []
    note_lines: list[str] = []
    media_count = 0
    try:
        with zipfile.ZipFile(path, "r") as archive:
            for name in archive.namelist():
                lowered = name.lower()
                if lowered.startswith("ppt/media/"):
                    media_count += 1
                    continue
                if not lowered.endswith(".xml"):
                    continue
                try:
                    xml_text = archive.read(name).decode("utf-8", errors="ignore")
                except KeyError:
                    continue
                tokens = _extract_xml_text(xml_text)
                if lowered.startswith("ppt/slides/"):
                    slide_lines.extend(tokens)
                elif lowered.startswith("ppt/notesslides/"):
                    note_lines.extend(tokens)
    except zipfile.BadZipFile:
        return ExtractionResult("", ".pptx", "pptx_zip_xml", "failed_with_reason", "invalid pptx zip container")
    sections: list[str] = []
    if slide_lines:
        sections.append("Slide text:\n" + "\n".join(slide_lines[:120]))
    if note_lines:
        sections.append("Speaker notes:\n" + "\n".join(note_lines[:60]))
    if media_count:
        sections.append(f"Embedded media/images: {media_count}")
    if not sections:
        return ExtractionResult("", ".pptx", "pptx_zip_xml", "unsupported_no_local_tool", "pptx has no extractable text payload")
    return ExtractionResult("\n\n".join(sections), ".pptx", "pptx_zip_xml", "extracted_success", section="slides/notes")


def _extract_excel(path: Path) -> list[ExtractionResult]:
    try:
        import openpyxl
    except Exception as exc:  # noqa: BLE001
        return [ExtractionResult("", path.suffix.lower(), "openpyxl", "unsupported_no_local_tool", f"openpyxl unavailable: {exc}")]
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    except Exception as exc:  # noqa: BLE001
        return [ExtractionResult("", path.suffix.lower(), "openpyxl", "failed_with_reason", f"excel read failed: {exc}")]
    results: list[ExtractionResult] = []
    try:
        for sheet_name in workbook.sheetnames[:12]:
            sheet = workbook[sheet_name]
            lines: list[str] = [f"Excel sheet: {sheet_name}"]
            row_count = 0
            for row in sheet.iter_rows(max_row=25, values_only=True):
                values = [str(cell) for cell in row if cell is not None and str(cell).strip()]
                if values:
                    row_count += 1
                    lines.append(" | ".join(values))
            if row_count:
                results.append(ExtractionResult("\n".join(lines), path.suffix.lower(), "openpyxl", "extracted_success", section=f"sheet {sheet_name} preview", sheet=sheet_name, row_range=f"1-{row_count}"))
    finally:
        workbook.close()
        
    try:
        with zipfile.ZipFile(path, "r") as archive:
            drawings = [n for n in archive.namelist() if n.lower().startswith("xl/drawings/drawing") and n.lower().endswith(".xml")]
            shape_texts: list[str] = []
            for drawing in drawings:
                try:
                    xml_text = archive.read(drawing).decode("utf-8", errors="ignore")
                except KeyError:
                    continue
                tokens = _extract_xml_text(xml_text)
                if tokens:
                    shape_texts.extend(tokens)
            if shape_texts:
                clean_shapes = _clean_lines(shape_texts, limit=300)
                if clean_shapes:
                    results.append(ExtractionResult(
                        "Excel Shapes/Text Boxes:\n" + "\n".join(clean_shapes),
                        path.suffix.lower(),
                        "openpyxl+zipfile",
                        "extracted_success",
                        section="shapes and text boxes"
                    ))
    except Exception:
        pass

    if not results:
        return [ExtractionResult("", path.suffix.lower(), "openpyxl", "failed_with_reason", "excel workbook has no readable values or shapes")]
    return results


def _extract_docx(path: Path) -> ExtractionResult:
    try:
        with zipfile.ZipFile(path, "r") as archive:
            names = ["word/document.xml"]
            names.extend(sorted(n for n in archive.namelist() if re.fullmatch(r"word/(header|footer)\d+\.xml", n, flags=re.IGNORECASE)))
            sections: list[str] = []
            for name in names:
                root = _xml_root_from_zip(archive, name)
                if root is None:
                    continue
                lines: list[str] = []
                for child in root.iter():
                    local = _xml_local_name(child.tag)
                    if local in {"p", "tbl"}:
                        text = _text_nodes(child)
                        if text:
                            lines.append(text)
                clean = _clean_lines(lines, limit=300)
                if clean:
                    label = "Document" if name == "word/document.xml" else Path(name).stem
                    sections.append(f"{label}:\n" + "\n".join(clean))
    except zipfile.BadZipFile:
        return ExtractionResult("", ".docx", "docx_zip_xml", "failed_with_reason", "invalid docx zip container")
    text = "\n\n".join(sections)
    if not _has_meaningful_text(text):
        return ExtractionResult("", ".docx", "docx_zip_xml", "failed_with_reason", "docx has no readable text")
    return ExtractionResult(text, ".docx", "docx_zip_xml", "extracted_success", section="word/document.xml")


def _discover_tesseract_cmd() -> str | None:
    env_cmd = os.environ.get("AIOS_TESSERACT_CMD", "").strip().strip('"')
    if env_cmd and Path(env_cmd).exists():
        return env_cmd
    path_cmd = shutil.which("tesseract")
    if path_cmd:
        return path_cmd
    for candidate in COMMON_TESSERACT_PATHS:
        if Path(candidate).exists():
            return candidate
    return None


def _ocr_lang() -> str:
    return os.environ.get("AIOS_OCR_LANG", DEFAULT_OCR_LANG).strip() or DEFAULT_OCR_LANG


def _tesseract_available() -> tuple[bool, str, str]:
    cmd = _discover_tesseract_cmd()
    lang = _ocr_lang()
    if not cmd:
        return False, "none", "local OCR unavailable: tesseract executable not found; set AIOS_TESSERACT_CMD or add Tesseract to PATH"
    try:
        import pytesseract
        pytesseract.pytesseract.tesseract_cmd = cmd
        pytesseract.get_tesseract_version()
        return True, "tesseract", cmd
    except Exception as exc:  # noqa: BLE001
        return False, "none", f"local OCR unavailable with {cmd}: {exc}; OCR lang={lang}"


def _ocr_image(path: Path, *, page: str = "") -> ExtractionResult:
    if path.stat().st_size > MAX_OCR_IMAGE_BYTES:
        return ExtractionResult("", path.suffix.lower(), "local_ocr", "unsupported_no_local_ocr", f"image exceeds OCR size guard: {MAX_OCR_IMAGE_BYTES} bytes", page=page, ocr_engine="none", ocr_lang=_ocr_lang())
    available, engine, detail = _tesseract_available()
    lang = _ocr_lang()
    if not available:
        return ExtractionResult("", path.suffix.lower(), "local_ocr", "unsupported_no_local_ocr", detail, page=page, ocr_engine="none", ocr_lang=lang)
    try:
        from PIL import Image
        import pytesseract
        if detail and Path(detail).exists():
            pytesseract.pytesseract.tesseract_cmd = detail
        with Image.open(path) as img:
            text = pytesseract.image_to_string(img, lang=lang, timeout=OCR_TIMEOUT_SECONDS)
    except Exception as exc:  # noqa: BLE001
        return ExtractionResult("", path.suffix.lower(), "local_ocr", "failed_with_reason", f"OCR failed: {exc}", page=page, ocr_engine=engine, ocr_lang=lang)
    lines = _clean_lines(text.splitlines(), limit=120)
    if not lines:
        return ExtractionResult("", path.suffix.lower(), "local_ocr", "ocr_partial", "OCR ran but returned no readable text", page=page, ocr_engine=engine, ocr_lang=lang)
    warning = "" if len(" ".join(lines)) >= 12 else "OCR output is very short; review manually"
    return ExtractionResult("\n".join(lines), path.suffix.lower(), "local_ocr", "ocr_success", warning=warning, page=page, section="ocr text", ocr_engine=engine, ocr_lang=lang)


def _extract_pdf(path: Path) -> list[ExtractionResult]:
    try:
        import pymupdf4llm
        has_pymupdf4llm = True
    except ImportError:
        has_pymupdf4llm = False

    try:
        import fitz
        has_fitz = True
    except ImportError:
        has_fitz = False

    if not has_pymupdf4llm and not has_fitz:
        return [ExtractionResult("", ".pdf", "pdf_optional_missing", "dependency_missing", "pymupdf4llm and fitz are not installed", element_type="pdf_page_text")]

    results: list[ExtractionResult] = []
    
    if has_pymupdf4llm:
        try:
            import pymupdf4llm
            md_text = pymupdf4llm.to_markdown(str(path))
            if md_text and md_text.strip():
                lines = _clean_lines(md_text.splitlines(), limit=500)
                results.append(ExtractionResult("\n".join(lines), ".pdf", "pymupdf4llm", "extracted", section="markdown doc", element_type="pdf_markdown_page"))
                return results
        except Exception:
            pass # fallback to fitz

    if has_fitz:
        try:
            import fitz
            doc = fitz.open(str(path))
            for index, page in enumerate(doc, start=1):
                try:
                    text = page.get_text("text") or ""
                except Exception:
                    text = ""
                lines = _clean_lines(text.splitlines(), limit=120)
                if lines:
                    results.append(ExtractionResult("\n".join(lines), ".pdf", "pymupdf", "extracted", section=f"page {index}", page=str(index), element_type="pdf_page_text"))
            doc.close()
        except Exception as exc:
            return [ExtractionResult("", ".pdf", "pymupdf", "parse_failed", f"pdf read failed: {exc}", element_type="pdf_page_text")]

    if not results:
        return [ExtractionResult("", ".pdf", "pymupdf" if has_fitz else "pymupdf4llm", "empty_text", "pdf has no extractable text", element_type="pdf_page_text")]
    
    return results


def extract_text_chunks_from_file(path: str | Path, *, root: str | Path | None = None, max_chars_per_chunk: int = 1200) -> list[dict[str, Any]]:
    file_path = Path(path)
    root_path = Path(root).resolve() if root else None
    ext = file_path.suffix.lower()
    if ext in {".html", ".htm"}:
        results = [_extract_html(file_path)]
    elif ext == ".pptx":
        results = [_extract_pptx(file_path)]
    elif ext in {".xlsx", ".xlsm"}:
        results = _extract_excel(file_path)
    elif ext == ".docx":
        results = [_extract_docx(file_path)]
    elif ext in IMAGE_EXTS:
        results = [_ocr_image(file_path)]
    elif ext == ".pdf":
        results = _extract_pdf(file_path)
    else:
        rel = file_path.relative_to(root_path).as_posix() if root_path else file_path.name
        return [{
            "text": "",
            "source_file": file_path.name,
            "relative_path": rel,
            "file_type": ext or "[no_ext]",
            "section": "",
            "page": "",
            "slide": "",
            "sheet": "",
            "row_range": "",
            "privacy_level": "local_only",
            "extractor_name": "document_extractors",
            "extraction_status": "unsupported_no_local_tool",
            "warning": "unsupported file type",
            "ocr_engine": "",
            "ocr_lang": "",
        }]
    chunks: list[dict[str, Any]] = []
    for result in results:
        chunks.extend(_chunk_result(result, file_path, root_path, max_chars_per_chunk))
    return chunks


def is_potentially_extractable(ext: str) -> bool:
    return ext.lower() in {".html", ".htm", ".pptx", ".pdf", ".xlsx", ".xlsm", ".docx"} | IMAGE_EXTS


def local_capabilities() -> dict[str, Any]:
    ocr_available, ocr_engine, detail = _tesseract_available()
    return {
        "ocr_available": ocr_available,
        "ocr_engine": ocr_engine,
        "tesseract_cmd": detail if ocr_available else "",
        "ocr_lang": _ocr_lang(),
        "ocr_warning": "" if ocr_available else detail,
        "pdf_render_available": __import__("importlib").util.find_spec("fitz") is not None,
        "image_size_guard_bytes": MAX_OCR_IMAGE_BYTES,
        "pdf_ocr_page_guard": MAX_PDF_OCR_PAGES,
        "cloud_ocr_used": False,
    }
