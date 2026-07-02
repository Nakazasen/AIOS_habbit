from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from aios_habit.workspace_chat_models import (
    WORKSPACE_CHAT_SOURCE_PREVIEW_LIMIT,
    WORKSPACE_CHAT_SOURCE_TEXT_LIMIT_BYTES,
)

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_ZIP_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_SHEETS = 12
MAX_ROWS_PER_SHEET = 1000
MAX_NON_EMPTY_CELLS = 20_000

XLS_UNSUPPORTED_MESSAGE = "Định dạng file không được hỗ trợ. Vui lòng lưu file dưới dạng .xlsx và thử lại."
GENERIC_READ_ERROR_MESSAGE = "Không thể đọc nội dung file Excel. File có thể bị hỏng hoặc có mật khẩu. Vui lòng kiểm tra lại."
EMPTY_WORKBOOK_MESSAGE = "File Excel này không có ô dữ liệu nào có thể đọc."
TRUNCATED_MESSAGE = "File Excel lớn nên AIOS chỉ đọc một phần nội dung. Bạn có thể chia file nhỏ hơn nếu cần."


@dataclass(frozen=True)
class ExtractedWorkspaceSource:
    ok: bool
    filename: str
    file_type: str = "xlsx"
    sheet_names: tuple[str, ...] = ()
    text: str = ""
    preview: str = ""
    truncated: bool = False
    owner_message: str = ""


def _safe_filename(filename: str) -> str:
    name = Path(filename or "").name.strip()
    return name or "workbook.xlsx"


def _failure(filename: str, message: str, truncated: bool = False) -> ExtractedWorkspaceSource:
    return ExtractedWorkspaceSource(ok=False, filename=_safe_filename(filename), truncated=truncated, owner_message=message)


def _cap_utf8(text: str, limit: int) -> tuple[str, bool]:
    encoded = text.encode("utf-8")
    if len(encoded) <= limit:
        return text, False
    return encoded[:limit].decode("utf-8", errors="ignore"), True


def _normalize_cell_value(value: Any) -> str:
    if isinstance(value, datetime):
        raw = value.isoformat(sep=" ")
    elif isinstance(value, (date, time)):
        raw = value.isoformat()
    else:
        raw = str(value)
    return re.sub(r"\s+", " ", raw).strip()


def _preflight_zip(file_bytes: bytes, filename: str) -> ExtractedWorkspaceSource | None:
    try:
        with zipfile.ZipFile(BytesIO(file_bytes)) as archive:
            total_uncompressed = sum(info.file_size for info in archive.infolist())
    except zipfile.BadZipFile:
        return _failure(filename, GENERIC_READ_ERROR_MESSAGE)
    except Exception:
        return _failure(filename, GENERIC_READ_ERROR_MESSAGE)
    if total_uncompressed > MAX_ZIP_UNCOMPRESSED_BYTES:
        return _failure(filename, GENERIC_READ_ERROR_MESSAGE, truncated=True)
    return None


def extract_xlsx_text(file_bytes: bytes, filename: str) -> ExtractedWorkspaceSource:
    safe_name = _safe_filename(filename)
    lower_name = safe_name.lower()

    if lower_name.endswith(".xls"):
        return _failure(safe_name, XLS_UNSUPPORTED_MESSAGE)
    if not lower_name.endswith(".xlsx"):
        return _failure(safe_name, GENERIC_READ_ERROR_MESSAGE)
    if not file_bytes:
        return _failure(safe_name, GENERIC_READ_ERROR_MESSAGE)
    if len(file_bytes) > MAX_UPLOAD_BYTES:
        return _failure(safe_name, GENERIC_READ_ERROR_MESSAGE, truncated=True)

    zip_failure = _preflight_zip(file_bytes, safe_name)
    if zip_failure is not None:
        return zip_failure

    workbook = None
    truncated = False
    non_empty_cells = 0
    output_lines: list[str] = [f"File Excel: {safe_name}"]
    sheet_names: list[str] = []
    all_sheet_names: tuple[str, ...] = ()

    try:
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=False, keep_links=False)
        all_sheet_names = tuple(workbook.sheetnames)

        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            if sheet_index >= MAX_SHEETS:
                truncated = True
                break
            sheet_names.append(sheet_name)
            ws = workbook[sheet_name]
            output_lines.extend(["", f"Trang tính: {sheet_name}"])
            sheet_has_data = False

            for row_index, row in enumerate(ws.iter_rows(), start=1):
                if row_index > MAX_ROWS_PER_SHEET:
                    truncated = True
                    break
                row_items: list[str] = []
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    normalized = _normalize_cell_value(value)
                    if not normalized:
                        continue
                    non_empty_cells += 1
                    if non_empty_cells > MAX_NON_EMPTY_CELLS:
                        truncated = True
                        break
                    coordinate = getattr(cell, "coordinate", f"{get_column_letter(cell.column)}{cell.row}")
                    row_items.append(f"{coordinate}={normalized}")
                    sheet_has_data = True
                if row_items:
                    output_lines.append(" | ".join(row_items))
                if truncated and non_empty_cells > MAX_NON_EMPTY_CELLS:
                    break

            if not sheet_has_data:
                output_lines.pop()
                output_lines.pop()
            if truncated:
                break

        if non_empty_cells == 0:
            return ExtractedWorkspaceSource(ok=False, filename=safe_name, sheet_names=all_sheet_names, truncated=truncated, owner_message=EMPTY_WORKBOOK_MESSAGE)

        text = "\n".join(output_lines).strip() + "\n"
        text, capped = _cap_utf8(text, WORKSPACE_CHAT_SOURCE_TEXT_LIMIT_BYTES)
        truncated = truncated or capped
        if truncated:
            text, _ = _cap_utf8(text.rstrip() + "\n" + TRUNCATED_MESSAGE + "\n", WORKSPACE_CHAT_SOURCE_TEXT_LIMIT_BYTES)

        preview = text[:WORKSPACE_CHAT_SOURCE_PREVIEW_LIMIT]
        owner_message = TRUNCATED_MESSAGE if truncated else "Đã đọc nội dung Excel và thêm vào nguồn tạm của cuộc trò chuyện."
        return ExtractedWorkspaceSource(ok=True, filename=safe_name, sheet_names=tuple(sheet_names), text=text, preview=preview, truncated=truncated, owner_message=owner_message)
    except Exception:
        return _failure(safe_name, GENERIC_READ_ERROR_MESSAGE)
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
