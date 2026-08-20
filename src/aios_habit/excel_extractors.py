from __future__ import annotations

import datetime
import importlib.util
import re
import zipfile
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO


@dataclass(frozen=True)
class ExcelExtractionConfig:
    max_file_bytes: int = 50 * 1024 * 1024
    max_uncompressed_bytes: int = 200 * 1024 * 1024
    max_sheets: int = 30
    max_rows_per_sheet: int | None = None
    max_non_empty_cells: int | None = None
    max_columns_per_region: int = 256
    max_images: int = 24
    max_image_bytes: int = 8 * 1024 * 1024
    max_total_image_bytes: int = 24 * 1024 * 1024
    max_image_pixels: int = 24_000_000
    max_charts: int = 48
    max_header_rows: int = 3
    chunk_row_size: int = 500
    enable_row_chunking: bool = True
    repeat_headers_in_chunks: bool = True


@dataclass(frozen=True)
class ExcelCell:
    row: int
    column: int
    coordinate: str
    text: str
    is_header: bool = False
    row_span: int = 1
    col_span: int = 1
    merge_range: str = ""


@dataclass(frozen=True)
class ExcelTableRegion:
    sheet: str
    cell_range: str
    row_range: tuple[int, int]
    column_range: tuple[int, int]
    rows: tuple[tuple[str, ...], ...]
    cells: tuple[ExcelCell, ...]
    header_rows: tuple[tuple[str, ...], ...] = ()
    headers: tuple[str, ...] = ()
    merged_ranges: tuple[str, ...] = ()
    chunk_index: int = 0
    total_chunks: int = 1


@dataclass(frozen=True)
class ExcelEmbeddedImage:
    sheet: str
    anchor: str
    index: int
    data: bytes
    extension: str = ".png"
    width: int | None = None
    height: int | None = None


@dataclass(frozen=True)
class ExcelChartMetadata:
    sheet: str
    anchor: str
    index: int
    chart_type: str
    title: str = ""
    series: tuple[str, ...] = ()
    references: tuple[str, ...] = ()

    def as_text(self) -> str:
        lines = [f"Excel chart: {self.chart_type}", f"Sheet: {self.sheet}", f"Anchor: {self.anchor}"]
        if self.title:
            lines.append(f"Title: {self.title}")
        if self.series:
            lines.append("Series: " + " | ".join(self.series))
        if self.references:
            lines.append("Source ranges: " + " | ".join(self.references))
        return "\n".join(lines)


@dataclass
class ExcelExtraction:
    file_type: str
    sheet_names: tuple[str, ...] = ()
    regions: list[ExcelTableRegion] = field(default_factory=list)
    images: list[ExcelEmbeddedImage] = field(default_factory=list)
    charts: list[ExcelChartMetadata] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    truncated_reasons: list[str] = field(default_factory=list)
    error: str = ""
    dependency_missing: str = ""

    @property
    def partial(self) -> bool:
        return bool(self.warnings or self.truncated_reasons)

    @property
    def succeeded(self) -> bool:
        return bool(self.regions or self.images or self.charts) and not self.error


def legacy_xls_available() -> bool:
    return importlib.util.find_spec("xlrd") is not None


def normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        raw = value.isoformat(sep=" ")
    elif isinstance(value, (datetime.date, datetime.time)):
        raw = value.isoformat()
    else:
        raw = str(value)
    return re.sub(r"\s+", " ", raw).strip()


def _column_name(number: int) -> str:
    value = ""
    while number > 0:
        number, remainder = divmod(number - 1, 26)
        value = chr(65 + remainder) + value
    return value or "A"


def _coordinate(row: int, column: int) -> str:
    return f"{_column_name(column)}{row}"


def _anchor(anchor: Any) -> str:
    marker = getattr(anchor, "_from", anchor)
    row, column = getattr(marker, "row", None), getattr(marker, "col", None)
    if isinstance(row, int) and isinstance(column, int):
        return _coordinate(row + 1, column + 1)
    return anchor if isinstance(anchor, str) else ""


def _looks_numeric(value: str) -> bool:
    try:
        float(value.replace(",", ""))
        return True
    except (TypeError, ValueError):
        return False


def _blocks(values: list[int]) -> list[tuple[int, int]]:
    if not values:
        return []
    output: list[tuple[int, int]] = []
    start = previous = values[0]
    for value in values[1:]:
        if value > previous + 1:
            output.append((start, previous))
            start = value
        previous = value
    return [*output, (start, previous)]


def _header_depth(rows: list[tuple[int, dict[int, str]]], first: int, last: int, merges: list[Any], limit: int) -> int:
    depth = 0
    for offset, (row_number, values) in enumerate(rows[:limit]):
        line = [values.get(column, "") for column in range(first, last + 1)]
        has_merge = any(item.min_row <= row_number <= item.max_row and item.min_col <= last and item.max_col >= first for item in merges)
        textual = sum(bool(value) and not _looks_numeric(value) for value in line)
        numeric = sum(bool(value) and _looks_numeric(value) for value in line)
        next_numeric = sum(_looks_numeric(value) for value in (rows[offset + 1][1].values() if offset + 1 < len(rows) else ()))
        if has_merge or (textual and (not numeric or next_numeric > numeric)):
            depth += 1
        else:
            break
    return depth


def _headers(rows: tuple[tuple[str, ...], ...], width: int) -> tuple[str, ...]:
    output: list[str] = []
    seen: dict[str, int] = {}
    for column in range(width):
        parts: list[str] = []
        for row in rows:
            value = row[column].strip() if column < len(row) else ""
            if value and (not parts or value.casefold() != parts[-1].casefold()):
                parts.append(value)
        base = " > ".join(parts) or f"Column {_column_name(column + 1)}"
        key = base.casefold()
        seen[key] = seen.get(key, 0) + 1
        output.append(base if seen[key] == 1 else f"{base} ({seen[key]})")
    return tuple(output)


def _regions(sheet: str, rows: list[tuple[int, dict[int, str]]], merges: list[Any], config: ExcelExtractionConfig) -> list[ExcelTableRegion]:
    row_groups: list[list[tuple[int, dict[int, str]]]] = []
    current: list[tuple[int, dict[int, str]]] = []
    previous = 0
    for item in rows:
        if current and item[0] > previous + 1:
            row_groups.append(current)
            current = []
        current.append(item)
        previous = item[0]
    if current:
        row_groups.append(current)

    output: list[ExcelTableRegion] = []
    for group in row_groups:
        columns = sorted({column for _, values in group for column in values})
        for first, last in _blocks(columns):
            last = min(last, first + config.max_columns_per_region - 1)
            selected = [(number, {column: value for column, value in values.items() if first <= column <= last}) for number, values in group]
            selected = [(number, values) for number, values in selected if values]
            if not selected:
                continue
            table_start_row, table_end_row = selected[0][0], selected[-1][0]
            depth = _header_depth(selected, first, last, merges, config.max_header_rows)
            header_selected = selected[:depth]
            data_selected = selected[depth:]
            width = last - first + 1
            header_rows = tuple(tuple(values.get(column, "") for column in range(first, last + 1)) for _, values in header_selected)
            headers = _headers(header_rows, width) if depth else ()
            relevant = [item for item in merges if not (item.max_row < table_start_row or item.min_row > table_end_row or item.max_col < first or item.min_col > last)]
            merge_anchors = {(item.min_row, item.min_col): item for item in relevant}

            chunk_size = config.chunk_row_size if (config.chunk_row_size is not None and config.chunk_row_size > 0) else (len(data_selected) or 1)
            if config.enable_row_chunking and data_selected and len(data_selected) > chunk_size:
                chunk_slices = [data_selected[i:i + chunk_size] for i in range(0, len(data_selected), chunk_size)]
            else:
                chunk_slices = [data_selected]

            total_chunks = len(chunk_slices)
            for chunk_index, chunk_data in enumerate(chunk_slices):
                if chunk_index == 0:
                    chunk_start_row = table_start_row
                    chunk_end_row = chunk_data[-1][0] if chunk_data else table_end_row
                else:
                    chunk_start_row = chunk_data[0][0]
                    chunk_end_row = chunk_data[-1][0]

                chunk_relevant = [item for item in relevant if not (item.max_row < chunk_start_row or item.min_row > chunk_end_row or item.max_col < first or item.min_col > last)]

                if config.repeat_headers_in_chunks and depth > 0:
                    rows_for_chunk = header_selected + chunk_data
                else:
                    rows_for_chunk = (header_selected if chunk_index == 0 else []) + chunk_data

                cells: list[ExcelCell] = []
                matrix: list[tuple[str, ...]] = []
                for row_number, values in rows_for_chunk:
                    matrix.append(tuple(values.get(column, "") for column in range(first, last + 1)))
                    is_hdr = (row_number < table_start_row + depth)
                    if chunk_index == 0 or not is_hdr:
                        for column, text in values.items():
                            merged = merge_anchors.get((row_number, column))
                            cells.append(ExcelCell(
                                row_number, column, _coordinate(row_number, column), text,
                                is_hdr,
                                merged.max_row - merged.min_row + 1 if merged else 1,
                                merged.max_col - merged.min_col + 1 if merged else 1,
                                str(merged) if merged else "",
                            ))

                output.append(ExcelTableRegion(
                    sheet, f"{_coordinate(chunk_start_row, first)}:{_coordinate(chunk_end_row, last)}",
                    (chunk_start_row, chunk_end_row), (first, last), tuple(matrix), tuple(cells),
                    header_rows, headers, tuple(map(str, chunk_relevant)),
                    chunk_index=chunk_index, total_chunks=total_chunks,
                ))
    return output


def _chart_title(chart: Any) -> str:
    try:
        parts: list[str] = []
        for paragraph in chart.title.tx.rich.p:
            for item in (*getattr(paragraph, "r", ()), *getattr(paragraph, "fld", ())):
                if getattr(item, "t", None):
                    parts.append(str(item.t))
        return " ".join(parts).strip()
    except Exception:
        return ""


def _chart_series(chart: Any) -> tuple[tuple[str, ...], tuple[str, ...]]:
    names: list[str] = []
    references: list[str] = []
    for series in getattr(chart, "ser", ()):
        title = getattr(series, "tx", None)
        title_ref = getattr(title, "strRef", None)
        literal, formula = getattr(title, "v", None), getattr(title_ref, "f", None)
        if literal or formula:
            names.append(str(literal or formula))
        if formula:
            references.append(str(formula))
        for attr in ("cat", "val", "xVal", "yVal"):
            holder = getattr(series, attr, None)
            for ref_name in ("numRef", "strRef", "multiLvlStrRef"):
                ref = getattr(getattr(holder, ref_name, None), "f", None)
                if ref:
                    references.append(str(ref))
    return tuple(dict.fromkeys(names)), tuple(dict.fromkeys(references))


def _preflight(data: bytes, config: ExcelExtractionConfig) -> str:
    try:
        with zipfile.ZipFile(BytesIO(data)) as archive:
            if sum(item.file_size for item in archive.infolist()) > config.max_uncompressed_bytes:
                return f"uncompressed workbook exceeds {config.max_uncompressed_bytes} bytes"
    except zipfile.BadZipFile:
        return "invalid OOXML zip container"
    return ""


def extract_excel(
    source: str | Path | bytes | bytearray | BinaryIO,
    *, filename: str | None = None,
    config: ExcelExtractionConfig | None = None,
    include_images: bool = True,
    include_charts: bool = True,
) -> ExcelExtraction:
    cfg = config or ExcelExtractionConfig()
    if isinstance(source, (str, Path)):
        path = Path(source)
        name, data = filename or path.name, path.read_bytes()
    elif isinstance(source, (bytes, bytearray)):
        name, data = filename or "workbook.xlsx", bytes(source)
    else:
        name, data = filename or "workbook.xlsx", source.read()
    suffix = Path(name).suffix.lower()
    if len(data) > cfg.max_file_bytes:
        return ExcelExtraction(suffix, error=f"workbook exceeds {cfg.max_file_bytes} bytes")
    if suffix == ".xls":
        return _extract_xls(data, cfg)
    if suffix not in {".xlsx", ".xlsm"}:
        return ExcelExtraction(suffix, error=f"unsupported Excel format: {suffix or '[none]'}")
    error = _preflight(data, cfg)
    return ExcelExtraction(suffix, error=error) if error else _extract_openpyxl(data, suffix, cfg, include_images, include_charts)


def _extract_openpyxl(data: bytes, suffix: str, config: ExcelExtractionConfig, include_images: bool, include_charts: bool) -> ExcelExtraction:
    try:
        import openpyxl
    except ImportError:
        return ExcelExtraction(suffix, dependency_missing="openpyxl")
    result = ExcelExtraction(suffix)
    workbook = None
    try:
        workbook = openpyxl.load_workbook(BytesIO(data), read_only=False, data_only=False, keep_links=False, keep_vba=False)
        result.sheet_names = tuple(workbook.sheetnames)
        cell_count = image_bytes = 0
        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            if config.max_sheets is not None and sheet_index >= config.max_sheets:
                result.truncated_reasons.append(f"sheet limit: {config.max_sheets}")
                break
            sheet = workbook[sheet_name]
            rows: list[tuple[int, dict[int, str]]] = []
            stop = False
            for row_number, row in enumerate(sheet.iter_rows(), 1):
                if config.max_rows_per_sheet is not None and row_number > config.max_rows_per_sheet:
                    result.truncated_reasons.append(f"row limit on {sheet_name}: {config.max_rows_per_sheet}")
                    break
                values: dict[int, str] = {}
                for cell in row:
                    text = normalize_cell_value(cell.value)
                    if not text:
                        continue
                    cell_count += 1
                    if config.max_non_empty_cells is not None and cell_count > config.max_non_empty_cells:
                        result.truncated_reasons.append(f"cell limit: {config.max_non_empty_cells}")
                        stop = True
                        break
                    values[int(cell.column)] = text
                if values:
                    rows.append((row_number, values))
                if stop:
                    break
            result.regions.extend(_regions(sheet_name, rows, list(sheet.merged_cells.ranges), config))
            if include_charts:
                for chart in getattr(sheet, "_charts", ()):
                    if config.max_charts is not None and len(result.charts) >= config.max_charts:
                        result.truncated_reasons.append(f"chart limit: {config.max_charts}")
                        break
                    names, references = _chart_series(chart)
                    result.charts.append(ExcelChartMetadata(sheet_name, _anchor(chart.anchor), len(result.charts) + 1, type(chart).__name__, _chart_title(chart), names, references))
            if include_images:
                for image in getattr(sheet, "_images", ()):
                    if config.max_images is not None and len(result.images) >= config.max_images:
                        result.truncated_reasons.append(f"image limit: {config.max_images}")
                        break
                    width, height = int(getattr(image, "width", 0) or 0), int(getattr(image, "height", 0) or 0)
                    if width and height and config.max_image_pixels is not None and width * height > config.max_image_pixels:
                        result.warnings.append(f"skipped oversized image on {sheet_name}: {width}x{height}")
                        continue
                    try:
                        payload = bytes(image._data())
                    except Exception as exc:
                        result.warnings.append(f"image read failed on {sheet_name}: {exc}")
                        continue
                    if (config.max_image_bytes is not None and len(payload) > config.max_image_bytes) or (config.max_total_image_bytes is not None and image_bytes + len(payload) > config.max_total_image_bytes):
                        result.truncated_reasons.append("embedded image byte guard")
                        continue
                    image_bytes += len(payload)
                    extension = "." + str(getattr(image, "format", "png") or "png").lower().lstrip(".")
                    result.images.append(ExcelEmbeddedImage(sheet_name, _anchor(image.anchor), len(result.images) + 1, payload, extension, width or None, height or None))
            if stop:
                break
    except Exception as exc:
        result.error = f"Excel read failed: {exc}"
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
    if not result.error and not (result.regions or result.images or result.charts):
        result.error = "excel workbook has no readable values, images, or charts"
    return result


@dataclass(frozen=True)
class _LegacyMerge:
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    def __str__(self) -> str:
        return f"{_coordinate(self.min_row, self.min_col)}:{_coordinate(self.max_row, self.max_col)}"


def _extract_xls(data: bytes, config: ExcelExtractionConfig) -> ExcelExtraction:
    if not legacy_xls_available():
        return ExcelExtraction(".xls", dependency_missing="xlrd")
    import xlrd

    result = ExcelExtraction(".xls")
    workbook = None
    try:
        workbook = xlrd.open_workbook(file_contents=data, on_demand=True, formatting_info=True)
        result.sheet_names = tuple(workbook.sheet_names())
        cell_count = 0
        sheet_count = workbook.nsheets if config.max_sheets is None else min(workbook.nsheets, config.max_sheets)
        for sheet_index in range(sheet_count):
            sheet = workbook.sheet_by_index(sheet_index)
            rows: list[tuple[int, dict[int, str]]] = []
            row_limit = sheet.nrows if config.max_rows_per_sheet is None else min(sheet.nrows, config.max_rows_per_sheet)
            for row_index in range(row_limit):
                values: dict[int, str] = {}
                col_limit = sheet.ncols if config.max_columns_per_region is None else min(sheet.ncols, config.max_columns_per_region)
                for column_index in range(col_limit):
                    text = normalize_cell_value(sheet.cell_value(row_index, column_index))
                    if text:
                        cell_count += 1
                        if config.max_non_empty_cells is not None and cell_count > config.max_non_empty_cells:
                            result.truncated_reasons.append(f"cell limit: {config.max_non_empty_cells}")
                            break
                        values[column_index + 1] = text
                if values:
                    rows.append((row_index + 1, values))
                if config.max_non_empty_cells is not None and cell_count > config.max_non_empty_cells:
                    break
            merges = [_LegacyMerge(rlo + 1, rhi, clo + 1, chi) for rlo, rhi, clo, chi in getattr(sheet, "merged_cells", ())]
            result.regions.extend(_regions(sheet.name, rows, merges, config))
        if config.max_sheets is not None and workbook.nsheets > config.max_sheets:
            result.truncated_reasons.append(f"sheet limit: {config.max_sheets}")
    except Exception as exc:
        result.error = f"legacy Excel read failed: {exc}"
    finally:
        if workbook is not None:
            workbook.release_resources()
    if not result.error and not result.regions:
        result.error = "legacy Excel workbook has no readable values"
    return result

