"""Prepare deterministic NotebookLM-compatible representations of the 70-file corpus.

NotebookLM's local-file uploader accepts text/PDF/media formats but not Excel or
HTML. This script keeps the raw corpus untouched and creates one transparent,
deterministic derivative per unsupported source so a qualification notebook can
contain the same 70 logical sources. The manifest records the raw/derivative
hashes and must travel with any sealed reference.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
from pathlib import Path
import re
import zipfile
import xml.etree.ElementTree as ET
from typing import Any


SUPPORTED_UPLOAD_EXTENSIONS = frozenset({
    ".aac", ".csv", ".docx", ".epub", ".gif", ".jpeg", ".jpg", ".m4a",
    ".md", ".mp3", ".mp4", ".ogg", ".opus", ".pdf", ".png", ".txt",
    ".wav", ".webp",
})


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def json_hash(value: Any) -> str:
    encoded = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return str(value)


def workbook_to_csv(source: Path, target: Path) -> dict[str, Any]:
    import openpyxl

    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=False, keep_links=False)
    row_count = 0
    sheet_count = 0
    with target.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        for sheet_count, sheet in enumerate(workbook.worksheets, start=1):
            writer.writerow(["__SHEET__", sheet.title])
            for row in sheet.iter_rows(values_only=True):
                writer.writerow([scalar(value) for value in row])
                row_count += 1
    workbook.close()
    return {"conversion": "workbook_to_csv", "sheet_count": sheet_count, "row_count": row_count}


def workbook_to_text(source: Path, target: Path) -> dict[str, Any]:
    """Write a line-oriented representation for workbooks rejected as CSV."""
    import openpyxl

    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=False, keep_links=False)
    row_count = 0
    sheet_count = 0
    with target.open("w", encoding="utf-8") as handle:
        for sheet_count, sheet in enumerate(workbook.worksheets, start=1):
            handle.write(f"__SHEET__ {sheet.title}\n")
            for row in sheet.iter_rows(values_only=True):
                handle.write(" | ".join(scalar(value) for value in row) + "\n")
                row_count += 1
    workbook.close()
    return {"conversion": "workbook_to_text", "sheet_count": sheet_count, "row_count": row_count}


def html_to_text(source: Path, target: Path) -> dict[str, Any]:
    raw = source.read_text(encoding="utf-8", errors="replace")
    text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", raw)
    text = re.sub(r"(?s)<[^>]+>", "\n", text)
    text = html.unescape(text)
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
    cleaned = "\n".join(line for line in lines if line)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(cleaned + "\n", encoding="utf-8")
    return {"conversion": "html_to_text", "line_count": len(cleaned.splitlines())}


def pptx_to_text(source: Path, target: Path) -> dict[str, Any]:
    """Extract visible slide text without depending on a desktop renderer."""
    slides: list[tuple[str, list[str]]] = []
    with zipfile.ZipFile(source) as archive:
        for name in sorted(archive.namelist()):
            if not (name.startswith("ppt/slides/slide") and name.endswith(".xml")):
                continue
            root = ET.fromstring(archive.read(name))
            values = [node.text or "" for node in root.findall(".//{*}t")]
            slides.append((name, values))
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", encoding="utf-8") as handle:
        for name, values in slides:
            handle.write(f"__SLIDE__ {name}\n")
            handle.write(" ".join(value.strip() for value in values if value.strip()))
            handle.write("\n")
    return {"conversion": "pptx_to_text", "slide_count": len(slides)}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", type=Path, required=True)
    parser.add_argument("--local-manifest", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest = json.loads(args.local_manifest.read_text(encoding="utf-8"))
    rows = list(manifest.get("files", []))
    if len(rows) != 70 or int(manifest.get("business_file_count", 0)) != 70:
        raise SystemExit(f"Expected the sealed 70-file corpus, found {len(rows)} rows")
    args.output_dir.mkdir(parents=True, exist_ok=True)
    records: list[dict[str, Any]] = []
    for ordinal, row in enumerate(rows, start=1):
        relative = str(row["relative_path"])
        source = args.source_root / relative
        if not source.is_file():
            raise SystemExit(f"Missing corpus source: {source}")
        extension = source.suffix.casefold()
        raw_sha = sha256_file(source)
        title = source.name
        record: dict[str, Any] = {
            "ordinal": ordinal,
            "relative_path": relative,
            "title": title,
            "raw_extension": extension,
            "raw_sha256": raw_sha,
        }
        if extension in SUPPORTED_UPLOAD_EXTENSIONS:
            record.update({"upload_path": str(source.resolve()), "upload_extension": extension, "conversion": "none"})
        elif extension in {".xlsx", ".xlsm"}:
            if "agv通信仕様" in source.stem.casefold():
                target = args.output_dir / f"{ordinal:03d}__{source.stem}.txt"
                record.update(workbook_to_text(source, target))
                record.update({"upload_path": str(target.resolve()), "upload_extension": ".txt"})
            else:
                target = args.output_dir / f"{ordinal:03d}__{source.stem}.csv"
                record.update(workbook_to_csv(source, target))
                record.update({"upload_path": str(target.resolve()), "upload_extension": ".csv"})
        elif extension in {".pptx"}:
            target = args.output_dir / f"{ordinal:03d}__{source.stem}.txt"
            record.update(pptx_to_text(source, target))
            record.update({"upload_path": str(target.resolve()), "upload_extension": ".txt"})
        elif extension in {".html", ".htm"}:
            target = args.output_dir / f"{ordinal:03d}__{source.stem}.txt"
            record.update(html_to_text(source, target))
            record.update({"upload_path": str(target.resolve()), "upload_extension": ".txt"})
        else:
            raise SystemExit(f"No deterministic NotebookLM conversion for {relative} ({extension})")
        record["upload_sha256"] = sha256_file(Path(record["upload_path"]))
        records.append(record)
    output = {
        "schema_version": 1,
        "status": "READY_FOR_UPLOAD",
        "source_root": str(args.source_root.resolve()),
        "local_manifest_path": str(args.local_manifest.resolve()),
        "local_manifest_sha256": sha256_file(args.local_manifest),
        "logical_source_count": len(records),
        "records": records,
    }
    output["conversion_manifest_hash"] = json_hash(output)
    (args.output_dir / "conversion_manifest.json").write_text(
        json.dumps(output, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(json.dumps({
        "status": output["status"],
        "logical_source_count": output["logical_source_count"],
        "converted_count": sum(row["conversion"] != "none" for row in records),
        "conversion_manifest_hash": output["conversion_manifest_hash"],
        "output_dir": str(args.output_dir.resolve()),
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
